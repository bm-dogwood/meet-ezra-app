"""
Excel Report Generator for Sales Reports.

Generates daily and weekly sales report downloads matching the new template format:
  Center Name | Date | Revenue Target | Total Net | Service Net | Product Net |
  Average Ticket | Labor Target | Labor Hours | Tip Amount | CC | Cash
"""

from io import BytesIO
from datetime import date
from typing import List, Dict, Any
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:
    """Generates Excel workbooks from sales data."""

    SALES_COLUMNS = [
        'Center Name', 'Date', 'Revenue Target', 'Total Net', 'Service Net',
        'Product Net', 'Average Ticket', 'Labor Target', 'Labor Hours',
        'Tip Amount', 'CC', 'Cash',
    ]

    COLUMN_WIDTHS = {
        'Center Name': 30,
        'Date': 14,
        'Revenue Target': 16,
        'Total Net': 14,
        'Service Net': 14,
        'Product Net': 14,
        'Average Ticket': 16,
        'Labor Target': 14,
        'Labor Hours': 14,
        'Tip Amount': 14,
        'CC': 14,
        'Cash': 14,
    }

    CURRENCY_COLUMNS = [
        'Revenue Target', 'Total Net', 'Service Net', 'Product Net',
        'Average Ticket', 'Tip Amount', 'CC', 'Cash',
    ]

    HOURS_COLUMNS = ['Labor Target', 'Labor Hours']

    def _create_workbook_with_headers(self) -> Workbook:
        """Create a new workbook with formatted header row."""
        wb = Workbook()
        ws = wb.active

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, column_name in enumerate(self.SALES_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.COLUMN_WIDTHS.get(column_name, 14)

        return wb

    def _format_value(self, value: Any, column_name: str) -> Any:
        if value is None:
            return ''
        if column_name in self.CURRENCY_COLUMNS or column_name in self.HOURS_COLUMNS:
            if isinstance(value, (int, float, Decimal)):
                return round(float(value), 2)
        return value

    def _add_data_row(self, ws, row_idx: int, row_data: Dict[str, Any]) -> None:
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, column_name in enumerate(self.SALES_COLUMNS, start=1):
            value = row_data.get(column_name, '')
            formatted_value = self._format_value(value, column_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=formatted_value)
            cell.border = thin_border

            if column_name in self.CURRENCY_COLUMNS and formatted_value != '':
                cell.number_format = '#,##0.00'
            elif column_name in self.HOURS_COLUMNS and formatted_value != '':
                cell.number_format = '#,##0.00'

    def _add_summary_row(self, ws, row_idx: int, data_start_row: int) -> None:
        """Add a summary row with SUM/AVERAGE formulas."""
        summary_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        last_data_row = row_idx - 1

        for col_idx, column_name in enumerate(self.SALES_COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)

            if column_name == 'Center Name':
                value = 'Total'
            elif column_name == 'Date':
                value = ''
            elif column_name == 'Average Ticket':
                # Average, not sum
                value = f'=AVERAGE({col_letter}{data_start_row}:{col_letter}{last_data_row})'
            elif column_name in self.CURRENCY_COLUMNS or column_name in self.HOURS_COLUMNS:
                value = f'=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})'
            else:
                value = ''

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = summary_font
            cell.border = thin_border

            if column_name in self.CURRENCY_COLUMNS:
                cell.number_format = '#,##0.00'
            elif column_name in self.HOURS_COLUMNS:
                cell.number_format = '#,##0.00'

    def _add_no_data_message(self, ws) -> None:
        ws.cell(row=2, column=1, value='No data available for the requested date range')
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(self.SALES_COLUMNS))

    def generate_daily_report(self, data: List[Dict[str, Any]], report_date: date) -> BytesIO:
        wb = self._create_workbook_with_headers()
        ws = wb.active
        ws.title = f"Daily Sales {report_date.strftime('%Y-%m-%d')}"

        if not data:
            self._add_no_data_message(ws)
        else:
            for row_idx, row_data in enumerate(data, start=2):
                self._add_data_row(ws, row_idx, row_data)
            # Summary row
            self._add_summary_row(ws, len(data) + 2, data_start_row=2)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generate_weekly_report(
        self,
        data: List[Dict[str, Any]],
        start_date: date,
        end_date: date,
    ) -> BytesIO:
        wb = self._create_workbook_with_headers()
        ws = wb.active
        ws.title = f"Weekly Sales {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y')}"

        if not data:
            self._add_no_data_message(ws)
        else:
            for row_idx, row_data in enumerate(data, start=2):
                self._add_data_row(ws, row_idx, row_data)
            self._add_summary_row(ws, len(data) + 2, data_start_row=2)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
