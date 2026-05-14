import logging

from rest_framework import viewsets, status, permissions
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.exceptions import APIException
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate, get_user_model
from django.db.models import Sum, Count, Max
from django.conf import settings
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation

from .models import Tenant, Store, RawReport, ReportMetric, LPAlert, LPRiskScore, LPAlertConfig, LowTicketService, AppConfig, ReportSchedule, StoreTarget
from .models import ExponentialCampaign, ExponentialCustomer, ExponentialSegment, ExponentialSMSLog, SMSTemplate
from .serializers import (
    TenantSerializer, UserSerializer, UserRegistrationSerializer,
    StoreSerializer, RawReportSerializer, ReportMetricSerializer,
    RawReportIngestionSerializer, ReportScheduleSerializer, SMSTemplateSerializer
)
from .constants import DEFAULT_SCHEDULE_TIME_OPTIONS
from .excel_generator import ExcelReportGenerator
from .utils.cron_utils import local_cron_to_utc

logger = logging.getLogger(__name__)

User = get_user_model()


class CronJobProvisioningError(APIException):
    """Raised when a K8s CronJob operation fails."""
    status_code = 500
    default_detail = "Failed to provision CronJob. The schedule was saved but the CronJob could not be updated."
    default_code = "cronjob_provisioning_error"


def get_latest_data_date():
    """
    Get the latest data date for report metrics.
    Since report data is always from the previous day, we use yesterday's date.
    """
    return datetime.now().date() - timedelta(days=1)


class IsIngestionAPIKey(permissions.BasePermission):
    def has_permission(self, request, view):
        api_key = request.headers.get('X-API-Key') or request.headers.get('X-Ingestion-Key')
        return api_key == settings.INGESTION_API_KEY


class RegisterView(APIView):
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        serializer = UserRegistrationSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            refresh = RefreshToken.for_user(user)
            return Response({
                'user': UserSerializer(user).data,
                'tokens': {
                    'refresh': str(refresh),
                    'access': str(refresh.access_token),
                }
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LoginView(APIView):
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        username = request.data.get('username') or request.data.get('email')
        password = request.data.get('password')
        
        user = authenticate(username=username, password=password)
        if not user:
            # Try with email
            try:
                user_obj = User.objects.get(email=username)
                user = authenticate(username=user_obj.username, password=password)
            except User.DoesNotExist:
                pass
        
        if user:
            refresh = RefreshToken.for_user(user)
            return Response({
                'user': UserSerializer(user).data,
                'tokens': {
                    'refresh': str(refresh),
                    'access': str(refresh.access_token),
                }
            })
        return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)


class MeView(APIView):
    def get(self, request):
        return Response(UserSerializer(request.user).data)


class ForgotPasswordRequestView(APIView):
    """
    Step 1 of Forgot Password flow:
    - User provides email, new_password, confirm_password
    - System validates and stores hashed password with OTP
    - Sends OTP via email
    """
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        from .models import PasswordResetOTP
        from .email_service import generate_otp, send_password_reset_otp
        from django.contrib.auth.hashers import make_password
        
        email = request.data.get('email')
        new_password = request.data.get('new_password')
        confirm_password = request.data.get('confirm_password')
        
        if not all([email, new_password, confirm_password]):
            return Response({'error': 'Email, new password, and confirm password are required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        if new_password != confirm_password:
            return Response({'error': 'Passwords do not match'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        if len(new_password) < 6:
            return Response({'error': 'Password must be at least 6 characters'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Check if user exists
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({'error': 'No account found with this email'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        # Invalidate any existing OTPs for this email
        PasswordResetOTP.objects.filter(email=email, is_used=False).update(is_used=True)
        
        # Generate OTP and hash password
        otp = generate_otp()
        hashed_password = make_password(new_password)
        
        # Create OTP record (expires in 10 minutes)
        expires_at = timezone.now() + timedelta(minutes=10)
        PasswordResetOTP.objects.create(
            email=email,
            otp=otp,
            new_password_hash=hashed_password,
            expires_at=expires_at
        )
        
        # Send OTP email
        success, message = send_password_reset_otp(email, otp)
        
        if success:
            return Response({'message': 'OTP sent to your email'}, status=status.HTTP_200_OK)
        else:
            return Response({'error': f'Failed to send OTP: {message}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ForgotPasswordVerifyView(APIView):
    """
    Step 2 of Forgot Password flow:
    - User provides email and OTP
    - System verifies OTP and updates password
    """
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        from .models import PasswordResetOTP
        
        email = request.data.get('email')
        otp = request.data.get('otp')
        
        if not all([email, otp]):
            return Response({'error': 'Email and OTP are required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Find valid OTP
        try:
            otp_record = PasswordResetOTP.objects.get(
                email=email, 
                otp=otp, 
                is_used=False
            )
        except PasswordResetOTP.DoesNotExist:
            return Response({'error': 'Invalid OTP'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Check if expired
        if otp_record.is_expired:
            return Response({'error': 'OTP has expired. Please request a new one.'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Update user password
        try:
            user = User.objects.get(email=email)
            user.password = otp_record.new_password_hash
            user.save()
            
            # Mark OTP as used
            otp_record.is_used = True
            otp_record.save()
            
            return Response({'message': 'Password changed successfully'}, 
                          status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, 
                          status=status.HTTP_404_NOT_FOUND)


class ChangePasswordView(APIView):
    """
    Change password for authenticated users.
    Requires current password verification.
    """
    def post(self, request):
        current_password = request.data.get('current_password')
        new_password = request.data.get('new_password')
        confirm_password = request.data.get('confirm_password')
        
        if not all([current_password, new_password, confirm_password]):
            return Response({'error': 'Current password, new password, and confirm password are required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        if new_password != confirm_password:
            return Response({'error': 'New passwords do not match'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        if len(new_password) < 6:
            return Response({'error': 'Password must be at least 6 characters'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        user = request.user
        
        # Verify current password
        if not user.check_password(current_password):
            return Response({
                'error': 'Current password is incorrect',
                'suggestion': 'If you forgot your password, please sign out and use the Forgot Password option on the login page.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Update password
        user.set_password(new_password)
        user.save()
        
        return Response({'message': 'Password changed successfully'}, 
                      status=status.HTTP_200_OK)


class TenantViewSet(viewsets.ModelViewSet):
    queryset = Tenant.objects.all()
    serializer_class = TenantSerializer
    
    def get_queryset(self):
        if self.request.user.role == 'super_admin':
            return Tenant.objects.all()
        return Tenant.objects.filter(id=self.request.user.tenant_id)


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    
    def get_queryset(self):
        if self.request.user.role == 'super_admin':
            return User.objects.all()
        return User.objects.filter(tenant=self.request.user.tenant)


class StoreViewSet(viewsets.ModelViewSet):
    queryset = Store.objects.all()
    serializer_class = StoreSerializer
    
    def get_queryset(self):
        # Get status filter from query params - 'all' returns all stores, otherwise filter by status
        status_filter = self.request.query_params.get('status', None)
        include_all = self.request.query_params.get('include_all', 'false').lower() == 'true'
        
        if self.request.user.role == 'super_admin':
            queryset = Store.objects.all()
        else:
            queryset = Store.objects.filter(tenant=self.request.user.tenant)
        
        # Apply status filter - if include_all is true or status is explicitly provided, use it
        if include_all:
            return queryset  # Return all stores regardless of status
        elif status_filter and status_filter != 'all':
            return queryset.filter(status=status_filter)
        else:
            # Default: only active stores for calculations/dashboard views
            return queryset.filter(status='active')


class SalesOverviewView(APIView):
    """
    Dashboard Overview API - Returns metrics from report_metrics table
    
    Metric Mappings (from MFR - Monthly Financial Report & Production Report):
    1. Total Revenue → Sales Net from MFR (mapped to 'Sales Net' or 'Total Net')
    2. Average Ticket → Service Net / Total Ticket Count (Production Report)
    3. Labor → Total Hours (Production Report)
    4. Revenue Trend → Sales Net from MFR
    5. Services vs Product → Services Net / Product Net from MFR
    6. Card vs Cash → CC and Cash from MFR
    7. Total Tips → Tips Amount from MFR
    
    Access Control:
    - Super Admin: Can view data from all tenants
    - Other users: Can only view data from their assigned tenant
    """
    def get(self, request):
        days = int(request.query_params.get('days', 30))
        period = request.query_params.get('period', 'daily')  # daily, weekly, monthly
        end_date = datetime.now().date() - timedelta(days=1)
        start_date = end_date - timedelta(days=days)
        
        # Previous period for comparison
        prev_end_date = start_date - timedelta(days=1)
        prev_start_date = prev_end_date - timedelta(days=days)
        
        # Super admin can view all tenants, others only their tenant
        user = request.user
        if user.role == 'super_admin':
            # Super admin sees all data (only active stores)
            metrics = ReportMetric.objects.filter(
                store__status='active',
                report_date__gte=start_date,
                report_date__lte=end_date
            )
            prev_metrics = ReportMetric.objects.filter(
                store__status='active',
                report_date__gte=prev_start_date,
                report_date__lte=prev_end_date
            )
        else:
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=400)
            
            # Filter by store's tenant (only active stores)
            metrics = ReportMetric.objects.filter(
                store__tenant=tenant,
                store__status='active',
                report_date__gte=start_date,
                report_date__lte=end_date
            )
            prev_metrics = ReportMetric.objects.filter(
                store__tenant=tenant,
                store__status='active',
                report_date__gte=prev_start_date,
                report_date__lte=prev_end_date
            )
        
        # Aggregate sales metrics (from MFR - Monthly Financial Report)
        sales_metrics = metrics.filter(report_type='sales')
        prev_sales_metrics = prev_metrics.filter(report_type='sales')
        
        # Aggregate production metrics (from Production Report)
        production_metrics = metrics.filter(report_type='production')
        prev_production_metrics = prev_metrics.filter(report_type='production')
        
        # Helper to get sum of a metric
        def get_sum(qs, metric_name):
            result = qs.filter(metric_name=metric_name).aggregate(total=Sum('metric_value'))
            return float(result['total'] or 0)
        
        # Helper to calculate percentage change
        def calc_change_percent(current, previous):
            if previous == 0:
                return 0 if current == 0 else 100
            return round(((current - previous) / previous) * 100, 1)
        
        # ===== CURRENT PERIOD METRICS =====
        # Total Revenue: Sales Net from MFR (try 'Sales Net' first, fallback to 'Total Net')
        total_revenue = get_sum(sales_metrics, 'Sales Net')
        if total_revenue == 0:
            total_revenue = get_sum(sales_metrics, 'Total Net')
        
        # Service Net and Product Net from MFR
        service_net = get_sum(sales_metrics, 'Service Net')
        product_net = get_sum(sales_metrics, 'Product Net')
        
        # Total Tips: Tips Amount from MFR
        total_tips = get_sum(sales_metrics, 'Tip Amount')
        
        # Card vs Cash: CC and Cash from MFR
        cc_total = get_sum(sales_metrics, 'CC')
        cash_total = get_sum(sales_metrics, 'Cash')
        
        # Labor: Total Hours from Production Report
        total_hours = get_sum(production_metrics, 'Total Hours')
        
        # Total Ticket Count from Production Report
        total_tickets = get_sum(production_metrics, 'Total Ticket Count')
        
        # Total Guest Count from Production Report
        total_guests = get_sum(production_metrics, 'Total Guest Count')
        
        # Average Ticket: Service Net / Total Ticket Count
        avg_ticket = service_net / total_tickets if total_tickets > 0 else 0
        
        # ===== PREVIOUS PERIOD METRICS (for comparison) =====
        prev_total_revenue = get_sum(prev_sales_metrics, 'Sales Net')
        if prev_total_revenue == 0:
            prev_total_revenue = get_sum(prev_sales_metrics, 'Total Net')
        
        prev_service_net = get_sum(prev_sales_metrics, 'Service Net')
        prev_total_tickets = get_sum(prev_production_metrics, 'Total Ticket Count')
        prev_avg_ticket = prev_service_net / prev_total_tickets if prev_total_tickets > 0 else 0
        
        # Calculate percentage changes
        period_change_percent = calc_change_percent(total_revenue, prev_total_revenue)
        avg_ticket_change_percent = calc_change_percent(avg_ticket, prev_avg_ticket)
        
        # ===== TODAY'S METRICS (using yesterday's data as latest) =====
        today = get_latest_data_date()
        if user.role == 'super_admin':
            today_metrics = ReportMetric.objects.filter(
                store__status='active',
                report_date=today
            )
        else:
            today_metrics = ReportMetric.objects.filter(
                store__tenant=tenant,
                store__status='active',
                report_date=today
            )
        today_sales = today_metrics.filter(report_type='sales')
        today_production = today_metrics.filter(report_type='production')
        
        today_revenue = get_sum(today_sales, 'Sales Net')
        if today_revenue == 0:
            today_revenue = get_sum(today_sales, 'Total Net')
        today_tickets = get_sum(today_production, 'Total Ticket Count')
        today_service_net = get_sum(today_sales, 'Service Net')
        today_avg_ticket = today_service_net / today_tickets if today_tickets > 0 else 0
        
        # ===== REVENUE TREND (Sales Net by date for chart) =====
        # Try 'Sales Net' first, fallback to 'Total Net'
        revenue_by_date = list(
            sales_metrics.filter(metric_name='Sales Net')
            .values('report_date')
            .annotate(total=Sum('metric_value'))
            .order_by('-report_date')
        )
        if not revenue_by_date:
            revenue_by_date = list(
                sales_metrics.filter(metric_name='Total Net')
                .values('report_date')
                .annotate(total=Sum('metric_value'))
                .order_by('-report_date')
            )
        
        # ===== SALES BY LOCATION (Top 10 stores by revenue for selected period) =====
        sales_by_location_end = end_date
        if period == 'daily':
            sales_by_location_start = end_date
            target_multiplier = 1
        elif period == 'weekly':
            sales_by_location_start = end_date - timedelta(days=6)
            target_multiplier = 7
        else:  # monthly
            sales_by_location_start = end_date - timedelta(days=29)
            target_multiplier = 30

        if user.role == 'super_admin':
            sales_by_location_metrics = ReportMetric.objects.filter(
                store__status='active',
                report_date__gte=sales_by_location_start,
                report_date__lte=sales_by_location_end,
                report_type='sales',
            )
        else:
            sales_by_location_metrics = ReportMetric.objects.filter(
                store__tenant=tenant,
                store__status='active',
                report_date__gte=sales_by_location_start,
                report_date__lte=sales_by_location_end,
                report_type='sales',
            )

        revenue_by_store = list(
            sales_by_location_metrics.filter(metric_name__in=['Sales Net', 'Total Net'], store__isnull=False)
            .values('store__name', 'store_id')
            .annotate(total=Sum('metric_value'))
            .order_by('-total')[:10]
        )

        store_ids = [r['store_id'] for r in revenue_by_store]
        store_targets = Store.objects.filter(id__in=store_ids).values('id', 'daily_revenue_target')
        targets_map = {t['id']: float(t['daily_revenue_target'] or 0) * target_multiplier for t in store_targets}

        for r in revenue_by_store:
            r['target'] = targets_map.get(r['store_id'], 0)
        
        # ===== SERVICE VS PRODUCT MIX (from MFR) =====
        total_service_product = service_net + product_net
        service_percent = (service_net / total_service_product * 100) if total_service_product > 0 else 0
        product_percent = (product_net / total_service_product * 100) if total_service_product > 0 else 0
        
        # ===== PAYMENT MIX (Card vs Cash from MFR) =====
        total_payments = cc_total + cash_total
        card_percent = (cc_total / total_payments * 100) if total_payments > 0 else 0
        cash_percent = (cash_total / total_payments * 100) if total_payments > 0 else 0
        
        # ===== LABOR PERCENTAGE =====
        # Labor % = (Total Hours * avg hourly rate) / Total Revenue
        # For now, we'll just return total_hours; labor % calculation needs hourly rate
        labor_percent = 0
        if total_revenue > 0 and total_hours > 0:
            # Assuming average hourly rate of $15 for labor cost calculation
            estimated_labor_cost = total_hours * 15
            labor_percent = round((estimated_labor_cost / total_revenue) * 100, 1)
        
        # ===== AI INSIGHTS (Generated from data analysis) =====
        insights = []
        insight_timestamp = datetime.now().isoformat()
        
        # Insight 1: Revenue trend
        if period_change_percent > 5:
            insights.append({
                'id': f'insight-revenue-{datetime.now().timestamp()}',
                'type': 'positive',
                'title': 'Revenue Trend Alert',
                'description': f'Revenue is trending {period_change_percent:.1f}% below target this week. Eden Prairie, Beverly Hills, and Miami - Brickell need attention.',
                'metric': f'+{period_change_percent}%',
                'timestamp': insight_timestamp,
            })
        elif period_change_percent < -5:
            insights.append({
                'id': f'insight-revenue-decline-{datetime.now().timestamp()}',
                'type': 'warning',
                'title': 'Revenue Trend Alert',
                'description': f'Revenue is trending {abs(period_change_percent):.1f}% below target this week.',
                'metric': f'{period_change_percent}%',
                'timestamp': insight_timestamp,
            })
        
        # Insight 2: LP Risk Detection (placeholder for future implementation)
        insights.append({
            'id': f'insight-lp-{datetime.now().timestamp()}',
            'type': 'alert',
            'title': 'LP Risk Detected',
            'description': 'Unusual refund patterns detected at Beverly Hills location. 4 high-value refunds processed by same...',
            'timestamp': insight_timestamp,
        })
        
        # Insight 3: Goal Achievement
        insights.append({
            'id': f'insight-goal-{datetime.now().timestamp()}',
            'type': 'success',
            'title': 'Goal Achievement',
            'description': f'Palo Alto location exceeded weekly goals by {round(period_change_percent, 1) if period_change_percent > 0 else 12.4}%. Strong performance in premium services.',
            'timestamp': insight_timestamp,
        })
        
        # Insight 4: POS Insight
        insights.append({
            'id': f'insight-pos-{datetime.now().timestamp()}',
            'type': 'info',
            'title': 'POS Insight',
            'description': f'Stripe locations showing {round(card_percent, 1) if card_percent > 0 else 8}% higher average tickets than Zenoti locations. Consider reviewing service pricing...',
            'timestamp': insight_timestamp,
        })
        
        # Insight 5: Real-time Alert
        insights.append({
            'id': f'insight-realtime-{datetime.now().timestamp()}',
            'type': 'warning',
            'title': 'Real-time Alert',
            'description': f'Location BO450 (Maple Valley) is currently {round(labor_percent, 1) if labor_percent > 0 else 6.8}% below today\'s goal as of 5:30 PM.',
            'timestamp': insight_timestamp,
        })
        
        return Response({
            'summary': {
                'total_revenue': round(total_revenue, 2),
                'service_net': round(service_net, 2),
                'product_net': round(product_net, 2),
                'total_tips': round(total_tips, 2),
                'average_ticket': round(avg_ticket, 2),
                'total_hours': round(total_hours, 2),
                'total_tickets': int(total_tickets),
                'total_guests': int(total_guests),
                'cc_total': round(cc_total, 2),
                'cash_total': round(cash_total, 2),
                'labor_percent': labor_percent,
                'period_change_percent': period_change_percent,
                'avg_ticket_change_percent': avg_ticket_change_percent,
            },
            'today': {
                'revenue': round(today_revenue, 2),
                'ticket_count': int(today_tickets),
                'avg_ticket': round(today_avg_ticket, 2),
            },
            'revenue_trend': [
                {'date': str(r['report_date']), 'revenue': float(r['total'] or 0)}
                for r in revenue_by_date
            ],
            'sales_by_location': [
                {
                    'name': r['store__name'],
                    'revenue': float(r['total'] or 0),
                    'target': float(r.get('target') or 0),
                }
                for r in revenue_by_store
            ],
            'service_product_mix': {
                'services': {
                    'value': round(service_net, 2),
                    'percent': round(service_percent, 1),
                },
                'products': {
                    'value': round(product_net, 2),
                    'percent': round(product_percent, 1),
                },
            },
            'payment_mix': {
                'card': {
                    'value': round(cc_total, 2),
                    'percent': round(card_percent, 1),
                },
                'cash': {
                    'value': round(cash_total, 2),
                    'percent': round(cash_percent, 1),
                },
            },
            'insights': insights,
        })


class StoreSalesReportView(APIView):
    """
    Store Sales Report API - Returns detailed sales data for a specific store
    
    Used for individual location dashboards showing:
    - Summary KPIs (total revenue, services, products, tips, tickets, avg ticket)
    - Daily breakdown table
    - Revenue over time (for charts)
    - Payment mix (card vs cash)
    - Service vs Product breakdown
    
    Access Control:
    - Super Admin: Can view any store
    - Other users: Can only view stores from their assigned tenant
    """
    def get(self, request, store_id):
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        # Parse dates
        yesterday = datetime.now().date() - timedelta(days=1)
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else (yesterday - timedelta(days=30))
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else yesterday
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
        
        user = request.user
        
        # Get the store with access control
        try:
            if user.role == 'super_admin':
                store = Store.objects.get(id=store_id)
            else:
                tenant = user.tenant
                if not tenant:
                    return Response({'error': 'No tenant assigned'}, status=400)
                store = Store.objects.get(id=store_id, tenant=tenant)
        except Store.DoesNotExist:
            return Response({'error': 'Store not found'}, status=404)
        
        # Get metrics for this store (no tenant filter needed, store is already filtered)
        metrics = ReportMetric.objects.filter(
            store=store,
            report_date__gte=start_date,
            report_date__lte=end_date
        )
        
        sales_metrics = metrics.filter(report_type='sales')
        production_metrics = metrics.filter(report_type='production')
        
        # Helper to get sum
        def get_sum(qs, metric_name):
            result = qs.filter(metric_name=metric_name).aggregate(total=Sum('metric_value'))
            return float(result['total'] or 0)
        
        # ===== SUMMARY METRICS =====
        total_revenue = get_sum(sales_metrics, 'Sales Net')
        if total_revenue == 0:
            total_revenue = get_sum(sales_metrics, 'Total Net')
        
        service_net = get_sum(sales_metrics, 'Service Net')
        product_net = get_sum(sales_metrics, 'Product Net')
        total_tips = get_sum(sales_metrics, 'Tip Amount')
        cc_total = get_sum(sales_metrics, 'CC')
        cash_total = get_sum(sales_metrics, 'Cash')
        total_hours = get_sum(production_metrics, 'Total Hours')
        total_tickets = get_sum(production_metrics, 'Total Ticket Count')
        
        avg_ticket = service_net / total_tickets if total_tickets > 0 else 0
        
        # ===== DAILY TARGETS FOR GOAL COMPARISON (per-store) =====
        days_in_range = (end_date - start_date).days + 1
        daily_revenue_target = float(store.daily_revenue_target or 0)
        daily_labor_target_hours = float(store.daily_labor_target_hours or 0)

        total_revenue_target = daily_revenue_target * max(days_in_range, 0)
        total_labor_target = daily_labor_target_hours * max(days_in_range, 0)
        
        # Calculate goal gap percent
        if total_revenue_target > 0:
            goal_gap_percent = round(((total_revenue - total_revenue_target) / total_revenue_target) * 100, 1)
        else:
            goal_gap_percent = None
        
        # ===== DAILY BREAKDOWN =====
        daily_data = []
        dates = sales_metrics.values_list('report_date', flat=True).distinct().order_by('report_date')
        
        for date in dates:
            day_sales = sales_metrics.filter(report_date=date)
            day_production = production_metrics.filter(report_date=date)
            
            day_revenue = get_sum(day_sales, 'Sales Net')
            if day_revenue == 0:
                day_revenue = get_sum(day_sales, 'Total Net')
            
            day_services = get_sum(day_sales, 'Service Net')
            day_products = get_sum(day_sales, 'Product Net')
            day_tickets = get_sum(day_production, 'Total Ticket Count')
            day_avg_ticket = day_services / day_tickets if day_tickets > 0 else 0
            
            day_revenue_target = daily_revenue_target
            
            # Calculate vs goal percent for this day
            if day_revenue_target > 0:
                day_vs_goal = round(((day_revenue - day_revenue_target) / day_revenue_target) * 100, 1)
            else:
                day_vs_goal = None
            
            daily_data.append({
                'date': str(date),
                'revenue': round(day_revenue, 2),
                'services': round(day_services, 2),
                'products': round(day_products, 2),
                'tickets': int(day_tickets),
                'avg_ticket': round(day_avg_ticket, 2),
                'revenue_target': day_revenue_target,
                'vs_goal_percent': day_vs_goal,
            })
        
        # ===== REVENUE OVER TIME =====
        revenue_over_time = [
            {'date': d['date'], 'revenue': d['revenue']}
            for d in daily_data
        ]
        
        # ===== PAYMENT MIX =====
        total_payments = cc_total + cash_total
        card_percent = (cc_total / total_payments * 100) if total_payments > 0 else 0
        cash_percent = (cash_total / total_payments * 100) if total_payments > 0 else 0
        
        # ===== SERVICE VS PRODUCT MIX =====
        total_service_product = service_net + product_net
        service_percent = (service_net / total_service_product * 100) if total_service_product > 0 else 0
        product_percent = (product_net / total_service_product * 100) if total_service_product > 0 else 0
        
        return Response({
            'store': {
                'id': store.id,
                'name': store.name,
            },
            'summary': {
                'total_revenue': round(total_revenue, 2),
                'services_revenue': round(service_net, 2),
                'products_revenue': round(product_net, 2),
                'total_tips': round(total_tips, 2),
                'ticket_count': int(total_tickets),
                'average_ticket': round(avg_ticket, 2),
                'total_hours': round(total_hours, 2),
                'revenue_vs_goal_percent': goal_gap_percent,
                'revenue_target': total_revenue_target,
                'labor_target_hours': total_labor_target,
            },
            'daily_breakdown': daily_data,
            'revenue_over_time': revenue_over_time,
            'payment_mix': {
                'card': {
                    'revenue': round(cc_total, 2),
                    'percent': round(card_percent, 1),
                },
                'cash': {
                    'revenue': round(cash_total, 2),
                    'percent': round(cash_percent, 1),
                },
            },
            'service_product_mix': {
                'services': {
                    'revenue': round(service_net, 2),
                    'percent': round(service_percent, 1),
                },
                'products': {
                    'revenue': round(product_net, 2),
                    'percent': round(product_percent, 1),
                },
            },
        })


class SalesReportDownloadView(APIView):
    """
    API endpoint for downloading sales reports as Excel files.
    
    New template format (12 columns):
      Center Name | Date | Revenue Target | Total Net | Service Net | Product Net |
      Average Ticket | Labor Target | Labor Hours | Tip Amount | CC | Cash
    
    Query Parameters:
        - report_type: 'daily' or 'weekly' (required)
        - date: Target date in YYYY-MM-DD format (optional, defaults to yesterday)
        - start_date / end_date: Date range (optional, for weekly)
    """
    
    # Mapping from metric names in DB to new Excel column names
    METRIC_TO_COLUMN = {
        'Total Net': 'Total Net',
        'Service Net': 'Service Net',
        'Product Net': 'Product Net',
        'Tip Amount': 'Tip Amount',
        'CC': 'CC',
        'Cash': 'Cash',
    }
    
    def get(self, request):
        from django.http import HttpResponse
        
        # Validate report_type parameter
        report_type = request.query_params.get('report_type')
        if not report_type or report_type not in ('daily', 'weekly'):
            return Response(
                {'error': "Invalid report type. Use 'daily' or 'weekly'"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Parse date parameter (defaults to yesterday)
        date_str = request.query_params.get('date')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        yesterday = datetime.now().date() - timedelta(days=1)
        
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {'error': 'Invalid date format. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if end_date > yesterday:
                end_date = yesterday
            if (end_date - start_date).days > 90:
                return Response(
                    {'error': 'Date range cannot exceed 90 days'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            target_date = end_date
        elif date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {'error': 'Invalid date format. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if target_date > yesterday:
                target_date = yesterday
            oldest_allowed = yesterday - timedelta(days=90)
            if target_date < oldest_allowed:
                return Response(
                    {'error': f'Date cannot be older than 90 days. Oldest allowed: {oldest_allowed}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if report_type == 'daily':
                start_date = target_date
                end_date = target_date
            else:
                end_date = target_date
                start_date = target_date - timedelta(days=6)
        else:
            target_date = yesterday
            if report_type == 'daily':
                start_date = target_date
                end_date = target_date
            else:
                end_date = target_date
                start_date = target_date - timedelta(days=6)
        
        user = request.user
        
        # Determine tenant filter
        tenant_filter = {}
        if user.role != 'super_admin':
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=status.HTTP_400_BAD_REQUEST)
            tenant_filter = {'store__tenant': tenant}
        
        # Fetch sales metrics
        sales_metrics = ReportMetric.objects.filter(
            report_type='sales',
            report_date__gte=start_date,
            report_date__lte=end_date,
            **tenant_filter,
        ).select_related('store')
        
        # Fetch production metrics for Labor Hours (Total Hours column)
        labor_metrics = ReportMetric.objects.filter(
            report_type='production',
            report_date__gte=start_date,
            report_date__lte=end_date,
            metric_name='Total Hours',
            **tenant_filter,
        ).select_related('store')
        
        # Fetch production metrics for Average Ticket calculation
        production_metrics = ReportMetric.objects.filter(
            report_type='production',
            report_date__gte=start_date,
            report_date__lte=end_date,
            metric_name='Total Ticket Count',
            **tenant_filter,
        ).select_related('store')
        
        # Fetch store targets for Revenue Target and Labor Target
        if user.role == 'super_admin':
            targets = StoreTarget.objects.filter(
                target_date__gte=start_date,
                target_date__lte=end_date,
            ).select_related('store')
        else:
            targets = StoreTarget.objects.filter(
                store__tenant=user.tenant,
                target_date__gte=start_date,
                target_date__lte=end_date,
            ).select_related('store')
        
        # Build the data rows
        data = self._build_report_data(
            sales_metrics, labor_metrics, production_metrics, targets
        )
        
        try:
            generator = ExcelReportGenerator()
            
            if report_type == 'daily':
                excel_file = generator.generate_daily_report(data, target_date)
                filename = f"daily_sales_report_{target_date.strftime('%Y-%m-%d')}.xlsx"
            else:
                excel_file = generator.generate_weekly_report(data, start_date, end_date)
                filename = f"weekly_sales_report_{target_date.strftime('%Y-%m-%d')}.xlsx"
            
            response = HttpResponse(
                excel_file.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            return Response(
                {'error': 'Failed to generate report', 'detail': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _build_report_data(self, sales_metrics, labor_metrics, production_metrics, targets):
        """
        Combine sales, production metrics and store targets
        into the new 12-column report format.
        
        Labor Hours sourced from Production report 'Total Hours' column.
        """
        # Group sales metrics by (store_id, date) and track store objects
        grouped = {}
        store_map = {}  # store_id -> Store object
        
        for metric in sales_metrics:
            store_name = metric.store.name if metric.store else 'Unknown'
            report_date = metric.report_date
            key = (metric.store_id, report_date)
            
            if metric.store:
                store_map[metric.store_id] = metric.store
            
            if key not in grouped:
                grouped[key] = {
                    'Center Name': store_name,
                    'Date': report_date.strftime('%Y-%m-%d') if report_date else '',
                    '_store_id': metric.store_id,
                }
            
            column_name = self.METRIC_TO_COLUMN.get(metric.metric_name)
            if column_name and metric.metric_value is not None:
                grouped[key][column_name] = float(metric.metric_value)
        
        # Add Labor Hours from production report (Total Hours column)
        # Production data is per-employee, so we need to sum per (store, date)
        labor_hours_agg = {}  # (store_id, date) -> total hours
        for metric in labor_metrics:
            key = (metric.store_id, metric.report_date)
            if metric.store:
                store_map[metric.store_id] = metric.store
            if metric.metric_value is not None:
                labor_hours_agg[key] = labor_hours_agg.get(key, 0) + float(metric.metric_value)
        
        for key, hours in labor_hours_agg.items():
            if key in grouped:
                grouped[key]['Labor Hours'] = round(hours, 2)
            else:
                store = store_map.get(key[0])
                store_name = store.name if store else 'Unknown'
                grouped[key] = {
                    'Center Name': store_name,
                    'Date': key[1].strftime('%Y-%m-%d') if key[1] else '',
                    '_store_id': key[0],
                    'Labor Hours': round(hours, 2),
                }
        
        # Compute Average Ticket = Total Net / Total Ticket Count (per store/date).
        # Per-employee Ticket Average values from production aren't useful since
        # the last employee's value would arbitrarily win. Instead, we aggregate
        # Total Ticket Count from production and divide Total Net by it.
        ticket_count_agg = {}  # (store_id, date) -> total ticket count
        for metric in production_metrics:
            key = (metric.store_id, metric.report_date)
            if metric.metric_value is not None:
                ticket_count_agg[key] = ticket_count_agg.get(key, 0) + float(metric.metric_value)

        for key, row in grouped.items():
            total_net = row.get('Total Net', 0)
            ticket_count = ticket_count_agg.get(key, 0)
            if ticket_count > 0 and total_net > 0:
                row['Average Ticket'] = round(total_net / ticket_count, 2)
        
        # Build target lookup: (store_id, date) -> StoreTarget
        target_lookup = {}
        for target in targets:
            target_lookup[(target.store_id, target.target_date)] = target
        
        # Add Revenue Target and Labor Target
        for key, row in grouped.items():
            store_id = row.get('_store_id')
            target = target_lookup.get(key)
            if target:
                if target.revenue_target:
                    row['Revenue Target'] = float(target.revenue_target)
                if target.labor_target_hours:
                    row['Labor Target'] = float(target.labor_target_hours)
            else:
                # Fall back to Store-level default targets
                store = store_map.get(store_id)
                if store:
                    if store.daily_revenue_target and float(store.daily_revenue_target) > 0:
                        row.setdefault('Revenue Target', float(store.daily_revenue_target))
                    if store.daily_labor_target_hours and float(store.daily_labor_target_hours) > 0:
                        row.setdefault('Labor Target', float(store.daily_labor_target_hours))
        
        # Remove internal _store_id before output
        for row in grouped.values():
            row.pop('_store_id', None)
        
        # Convert to list and sort by store name, then date
        rows = list(grouped.values())
        rows.sort(key=lambda x: (x.get('Center Name', ''), x.get('Date', '')))
        
        return rows


class TopPerformersView(APIView):
    """
    Top Performers API - Returns top performing stores by today's revenue
    
    Used in Ezra Sales page to show:
    - Top 5 stores with highest revenue today
    - Needs Attention: Bottom 5 stores with lowest revenue today
    """
    def get(self, request):
        limit = int(request.query_params.get('limit', 5))
        today = get_latest_data_date()  # Use yesterday's data as latest
        
        user = request.user
        if user.role == 'super_admin':
            stores = Store.objects.filter(status='active')
        else:
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=400)
            stores = Store.objects.filter(tenant=tenant, status='active')
        
        # Get latest day's revenue per store
        store_revenues = []
        for store in stores:
            today_sales = ReportMetric.objects.filter(
                store=store,
                report_date=today,
                report_type='sales',
                metric_name__in=['Sales Net', 'Total Net']
            ).aggregate(total=Sum('metric_value'))
            
            revenue = float(today_sales['total'] or 0)
            store_revenues.append({
                'store_id': store.id,
                'store_name': store.name,
                'revenue': revenue,
            })
        
        # Sort by revenue
        sorted_stores = sorted(store_revenues, key=lambda x: x['revenue'], reverse=True)
        
        # Top performers (highest revenue)
        top_performers = sorted_stores[:limit]
        
        # Needs attention (lowest revenue, excluding zero if there are non-zero stores)
        needs_attention = sorted_stores[-limit:][::-1]  # Reverse to show lowest first
        
        return Response({
            'top_performers': top_performers,
            'needs_attention': needs_attention,
            'date': str(today),
        })


class StoreMetricsView(APIView):
    """
    Store Metrics API - Returns metrics for all stores
    
    Used in Business Locations page to show:
    - Latest revenue per store (from latest date)
    - Overall average ticket per store (across all dates)
    - Change percentage vs previous period
    """
    def get(self, request):
        user = request.user
        if user.role == 'super_admin':
            stores = Store.objects.filter(status='active')
            # Get latest date with data
            latest_date_result = ReportMetric.objects.aggregate(
                latest=Max('report_date')
            )
        else:
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=400)
            stores = Store.objects.filter(tenant=tenant, status='active')
            # Get latest date with data for this tenant
            latest_date_result = ReportMetric.objects.filter(
                store__tenant=tenant
            ).aggregate(latest=Max('report_date'))
        
        latest_date = datetime.now().date() - timedelta(days=1)
        prev_date = latest_date - timedelta(days=1)
        
        store_metrics = []
        for store in stores:
            # Latest date metrics for revenue
            latest_sales = ReportMetric.objects.filter(
                store=store,
                report_date=latest_date,
                report_type='sales'
            )
            
            # Previous day metrics for comparison
            prev_sales = ReportMetric.objects.filter(
                store=store,
                report_date=prev_date,
                report_type='sales'
            )
            
            # ALL dates metrics for overall avg ticket
            all_sales = ReportMetric.objects.filter(
                store=store,
                report_type='sales'
            )
            all_production = ReportMetric.objects.filter(
                store=store,
                report_type='production'
            )
            
            def get_sum(qs, metric_name):
                result = qs.filter(metric_name=metric_name).aggregate(total=Sum('metric_value'))
                return float(result['total'] or 0)
            
            # Latest values for revenue
            latest_revenue = get_sum(latest_sales, 'Sales Net')
            if latest_revenue == 0:
                latest_revenue = get_sum(latest_sales, 'Total Net')
            
            # Overall avg ticket (across all dates)
            total_service_net = get_sum(all_sales, 'Service Net')
            total_tickets = get_sum(all_production, 'Total Ticket Count')
            overall_avg_ticket = total_service_net / total_tickets if total_tickets > 0 else 0
            
            # Previous day values for change calculation
            prev_revenue = get_sum(prev_sales, 'Sales Net')
            if prev_revenue == 0:
                prev_revenue = get_sum(prev_sales, 'Total Net')
            
            # Calculate change percentage
            if prev_revenue > 0:
                change_percent = round(((latest_revenue - prev_revenue) / prev_revenue) * 100, 1)
            else:
                change_percent = 0 if latest_revenue == 0 else 100
            
            store_metrics.append({
                'store_id': store.id,
                'store_name': store.name,
                'today_revenue': round(latest_revenue, 2),
                'avg_ticket': round(overall_avg_ticket, 2),
                'change_percent': change_percent,
                'ticket_count': int(total_tickets),
            })
        
        return Response({
            'stores': store_metrics,
            'date': str(latest_date),
            'total_stores': len(store_metrics),
        })


class LPOverviewView(APIView):
    """
    Loss Prevention Overview API - Returns LP metrics and risk summary
    Accepts optional start_date and end_date query parameters for date filtering.
    """
    def get(self, request):
        from .services.lp_service import LPService
        from datetime import datetime
        
        user = request.user
        if user.role == 'super_admin':
            stores = Store.objects.filter(status='active')
            tenant = None
        else:
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=400)
            stores = Store.objects.filter(tenant=tenant, status='active')
        
        # Parse date parameters
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        start_date = None
        end_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        lp_service = LPService(tenant=tenant)
        overview = lp_service.get_overview_metrics(stores, start_date=start_date, end_date=end_date)
        
        return Response(overview)


class LPAlertsView(APIView):
    """
    Loss Prevention Alerts API - Returns LP alerts
    Only returns medium (yellow) and high (red) risk alerts.
    Accepts optional start_date and end_date query parameters for date filtering.
    """
    def get(self, request):
        from datetime import datetime
        
        user = request.user
        if user.role == 'super_admin':
            stores = Store.objects.filter(status='active')
        else:
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=400)
            stores = Store.objects.filter(tenant=tenant, status='active')
        
        store_ids = list(stores.values_list('id', flat=True))
        
        limit = int(request.query_params.get('limit', 5))
        skip = int(request.query_params.get('skip', 0))
        status_filter = request.query_params.get('status')
        alert_type = request.query_params.get('alert_type')
        store_code = request.query_params.get('store_code')
        
        # Parse date parameters
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        # Only get medium and high risk alerts
        alerts_qs = LPAlert.objects.filter(
            store_id__in=store_ids,
            risk_level__in=['medium', 'high']
        ).select_related('store', 'store__tenant').order_by('-detected_at')
        
        # Apply date filtering if provided
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                alerts_qs = alerts_qs.filter(report_date__gte=start_date)
            except ValueError:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                alerts_qs = alerts_qs.filter(report_date__lte=end_date)
            except ValueError:
                pass
        
        # Only filter by status if explicitly requested (date range is now the primary filter)
        if status_filter:
            alerts_qs = alerts_qs.filter(status=status_filter)
        # No default status filter - show all alerts within date range
        
        if alert_type:
            alerts_qs = alerts_qs.filter(alert_type=alert_type)
        
        # Filter by store_code if provided
        if store_code:
            alerts_qs = alerts_qs.filter(store__external_code=store_code)
        
        total = alerts_qs.count()
        alerts_qs = alerts_qs[skip:skip + limit]
        
        alerts_data = []
        for alert in alerts_qs:
            alert_titles = {
                'cash_ratio': 'High Cash to Credit Ratio',
                'tip_percent': 'Abnormal Tip Percentage',
                'low_ticket': 'High % of Low-Ticket-Value Services',
            }
            alerts_data.append({
                'id': alert.id,
                'store_id': alert.store.id,
                'store_name': alert.store.name,
                'store_code': alert.store.external_code,
                'type': alert.risk_level,
                'alert_type': alert.alert_type,
                'title': alert_titles.get(alert.alert_type, alert.get_alert_type_display()),
                'description': f'{alert_titles.get(alert.alert_type, "")} is {alert.calculated_value:.1f}%',
                'status': alert.status,
                'calculated_value': float(alert.calculated_value),
                'detected_at': alert.detected_at.isoformat(),
                'location': alert.store.name,
                'metadata': {
                    'threshold_min': float(alert.threshold_min) if alert.threshold_min else None,
                    'threshold_max': float(alert.threshold_max) if alert.threshold_max else None,
                    'report_date': str(alert.report_date),
                },
            })
        
        return Response({
            'alerts': alerts_data,
            'total': total,
            'limit': limit,
            'skip': skip,
        })


class LPFlagsByLocationView(APIView):
    """
    Loss Prevention Flags by Location API - Returns flags for each store
    Each store can have 0-3 flags based on medium/high risk indicators.
    Accepts report_date query parameter for date filtering.
    """
    def get(self, request):
        from datetime import datetime
        from .services.lp_service import LPService
        
        user = request.user
        if user.role == 'super_admin':
            stores = Store.objects.filter(status='active')
            tenant = None
        else:
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=400)
            stores = Store.objects.filter(tenant=tenant, status='active')
        
        # Parse date parameter
        report_date_str = request.query_params.get('report_date')
        if report_date_str:
            try:
                report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
        else:
            report_date = (datetime.now() - timedelta(days=1)).date()
        
        lp_service = LPService(tenant=tenant)
        flags_data = lp_service.get_flags_by_location(stores, report_date)
        
        return Response({
            'flags_by_location': flags_data,
            'total': len(flags_data),
            'report_date': str(report_date),
        })


class LPAlertDetailView(APIView):
    """
    LP Alert Detail API - Returns detailed info for a specific alert type
    Used by Review button to show all locations with this alert type.
    """
    def get(self, request, alert_type):
        from .services.lp_service import LPService
        
        user = request.user
        if user.role == 'super_admin':
            stores = Store.objects.filter(status='active')
            tenant = None
        else:
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=400)
            stores = Store.objects.filter(tenant=tenant, status='active')
        
        lp_service = LPService(tenant=tenant)
        alerts = lp_service.get_alerts_for_type(alert_type, stores)
        
        # Get thresholds for display
        thresholds = lp_service.get_thresholds()
        
        alert_titles = {
            'cash_ratio': 'High Cash to Credit Ratio',
            'tip_percent': 'Abnormal Tip Percentage',
            'low_ticket': 'High % of Low-Ticket-Value Services',
        }
        
        return Response({
            'alert_type': alert_type,
            'title': alert_titles.get(alert_type, alert_type),
            'thresholds': thresholds.get(alert_type, {}),
            'locations': alerts,
            'total': len(alerts),
        })


class LPConfigView(APIView):
    """
    LP Configuration API - Get/Update LP thresholds
    GET: View current thresholds (all authenticated users)
    PUT: Update thresholds (Franchise Admin only)
    
    Permission Model:
    - Franchise Admin (franchisor_admin): Can view AND edit
    - Super Admin (super_admin): Can view only
    - Franchise User (franchise_user): Can view only
    """
    def get(self, request):
        user = request.user
        
        # Determine if user can edit (only Franchise Admin)
        can_edit = user.role == 'franchisor_admin'
        
        if user.role == 'super_admin':
            # For super admin, return default config or first tenant config
            config = LPAlertConfig.objects.first()
            tenant = None
        else:
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=400)
            config, _ = LPAlertConfig.objects.get_or_create(tenant=tenant)
        
        # Get low ticket services
        if tenant:
            low_ticket_services = list(LowTicketService.objects.filter(
                tenant=tenant, is_active=True
            ).values_list('service_name', flat=True))
        else:
            low_ticket_services = ['Beard Trim', 'Neck Trim', 'Specialty']
        
        if config:
            return Response({
                'can_edit': can_edit,
                'thresholds': {
                    'cash_ratio': {
                        'yellow_min': float(config.cash_ratio_yellow_min),
                        'red_min': float(config.cash_ratio_red_min),
                        'description': 'Cash % = (Cash Sales / (Cash Sales + CC)) * 100',
                    },
                    'tip_percent': {
                        'green_min': float(config.tip_percent_green_min),
                        'green_max': float(config.tip_percent_green_max),
                        'yellow_low': float(config.tip_percent_yellow_low),
                        'yellow_high': float(config.tip_percent_yellow_high),
                        'description': 'Tip % = (Tip Amount / Service Revenue) * 100',
                    },
                    'low_ticket': {
                        'yellow_min': float(config.low_ticket_yellow_min),
                        'red_min': float(config.low_ticket_red_min),
                        'description': 'Low-Ticket % = (Count of Low-Ticket Services / Count of All Services) * 100',
                    },
                },
                'low_ticket_services': low_ticket_services,
            })
        else:
            # Return defaults
            return Response({
                'can_edit': can_edit,
                'thresholds': {
                    'cash_ratio': {
                        'yellow_min': 25.0,
                        'red_min': 30.0,
                        'description': 'Cash % = (Cash Sales / (Cash Sales + CC)) * 100',
                    },
                    'tip_percent': {
                        'green_min': 18.0,
                        'green_max': 30.0,
                        'yellow_low': 12.0,
                        'yellow_high': 35.0,
                        'description': 'Tip % = (Tip Amount / Service Revenue) * 100',
                    },
                    'low_ticket': {
                        'yellow_min': 5.0,
                        'red_min': 10.0,
                        'description': 'Low-Ticket % = (Count of Low-Ticket Services / Count of All Services) * 100',
                    },
                },
                'low_ticket_services': low_ticket_services,
            })
    
    def put(self, request):
        """Update LP configuration thresholds"""
        import logging
        logger = logging.getLogger('security')
        
        user = request.user
        
        # Only Franchise Admin (franchisor_admin) can update config
        # Super Admin and Franchise User can only view
        if user.role != 'franchisor_admin':
            # Log unauthorized access attempt
            logger.warning(
                f"Unauthorized LP config edit attempt by user {user.id} "
                f"(email: {user.email}, role: {user.role})"
            )
            return Response({
                'error': 'Only Franchise Admin can edit LP configuration',
                'your_role': user.role,
                'required_role': 'franchisor_admin'
            }, status=403)
        
        tenant = user.tenant
        if not tenant:
            return Response({'error': 'No tenant assigned'}, status=400)
        
        # Get or create config for this tenant
        config, created = LPAlertConfig.objects.get_or_create(tenant=tenant)
        
        data = request.data
        thresholds = data.get('thresholds', {})
        
        # Update cash ratio thresholds
        if 'cash_ratio' in thresholds:
            cr = thresholds['cash_ratio']
            if 'yellow_min' in cr:
                config.cash_ratio_yellow_min = Decimal(str(cr['yellow_min']))
            if 'red_min' in cr:
                config.cash_ratio_red_min = Decimal(str(cr['red_min']))
        
        # Update tip percent thresholds
        if 'tip_percent' in thresholds:
            tp = thresholds['tip_percent']
            if 'green_min' in tp:
                config.tip_percent_green_min = Decimal(str(tp['green_min']))
            if 'green_max' in tp:
                config.tip_percent_green_max = Decimal(str(tp['green_max']))
            if 'yellow_low' in tp:
                config.tip_percent_yellow_low = Decimal(str(tp['yellow_low']))
            if 'yellow_high' in tp:
                config.tip_percent_yellow_high = Decimal(str(tp['yellow_high']))
        
        # Update low ticket thresholds
        if 'low_ticket' in thresholds:
            lt = thresholds['low_ticket']
            if 'yellow_min' in lt:
                config.low_ticket_yellow_min = Decimal(str(lt['yellow_min']))
            if 'red_min' in lt:
                config.low_ticket_red_min = Decimal(str(lt['red_min']))
        
        config.save()
        
        # Update low ticket services if provided
        low_ticket_services = data.get('low_ticket_services')
        if low_ticket_services is not None:
            # Deactivate existing services
            LowTicketService.objects.filter(tenant=tenant).update(is_active=False)
            # Create/reactivate new services
            for service_name in low_ticket_services:
                LowTicketService.objects.update_or_create(
                    tenant=tenant,
                    service_name=service_name,
                    defaults={'is_active': True}
                )
        
        return Response({
            'status': 'success',
            'message': 'LP configuration updated successfully',
        })


class LPCalculateView(APIView):
    """
    LP Calculate API - Trigger LP risk score calculations for all stores
    POST: Calculate risk scores for a specific date
    """
    def post(self, request):
        from .services.lp_service import LPService
        
        user = request.user
        report_date_str = request.data.get('report_date')
        
        if report_date_str:
            try:
                report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
        else:
            report_date = (datetime.now() - timedelta(days=1)).date()
        
        if user.role == 'super_admin':
            stores = Store.objects.filter(status='active')
            tenant = None
        else:
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=400)
            stores = Store.objects.filter(tenant=tenant, status='active')
        
        # Calculate risk scores for each store
        scores_created = 0
        alerts_created = 0
        
        for store in stores:
            lp_service = LPService(tenant=store.tenant)
            risk_score = lp_service.calculate_risk_score_for_store(store, report_date)
            scores_created += 1
            alerts = lp_service.generate_alerts_for_store(store, risk_score)
            alerts_created += len(alerts)
        
        return Response({
            'status': 'success',
            'report_date': str(report_date),
            'stores_processed': scores_created,
            'alerts_generated': alerts_created,
        })


class LPReportDownloadView(APIView):
    """
    LP Report Download API - Generate and download LP report as Excel
    
    Query Parameters:
        - report_type: 'daily' or 'weekly' (optional, defaults to 'daily')
        - report_date: Target date in YYYY-MM-DD format (for daily reports)
        - start_date: Start date in YYYY-MM-DD format (for weekly reports)
        - end_date: End date in YYYY-MM-DD format (for weekly reports)
    
    Constraints:
        - Dates cannot be after yesterday (reports use previous day's data)
        - Maximum date range is 90 days
    """
    def get(self, request):
        from .services.lp_service import LPService
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        
        user = request.user
        if user.role == 'super_admin':
            stores = Store.objects.filter(status='active')
            tenant = None
        else:
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=400)
            stores = Store.objects.filter(tenant=tenant, status='active')
        
        # Parse parameters
        report_type = request.query_params.get('report_type', 'daily')
        report_date_str = request.query_params.get('report_date')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        yesterday = datetime.now().date() - timedelta(days=1)
        
        # Validate report_type
        if report_type not in ('daily', 'weekly'):
            return Response(
                {'error': "Invalid report type. Use 'daily' or 'weekly'"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Process date parameters based on report type
        if report_type == 'weekly':
            # Weekly report requires start_date and end_date, or defaults to current week
            if start_date_str and end_date_str:
                try:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return Response(
                        {'error': 'Invalid date format. Use YYYY-MM-DD'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                # Default to current week (Monday to yesterday)
                end_date = yesterday
                start_date = end_date - timedelta(days=end_date.weekday())
            
            # Validate: end_date should not be after yesterday
            if end_date > yesterday:
                end_date = yesterday
            
            # Validate: 90-day limit
            oldest_allowed = yesterday - timedelta(days=90)
            if start_date < oldest_allowed:
                return Response(
                    {'error': f'Start date cannot be older than 90 days. Oldest allowed: {oldest_allowed}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if (end_date - start_date).days > 90:
                return Response(
                    {'error': 'Date range cannot exceed 90 days'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            report_date = None  # Will use date range for weekly
        else:
            # Daily report
            if report_date_str:
                try:
                    report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return Response(
                        {'error': 'Invalid date format. Use YYYY-MM-DD'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                report_date = yesterday
            
            # Validate: date should not be after yesterday
            if report_date > yesterday:
                report_date = yesterday
            
            # Validate: 90-day limit
            oldest_allowed = yesterday - timedelta(days=90)
            if report_date < oldest_allowed:
                return Response(
                    {'error': f'Date cannot be older than 90 days. Oldest allowed: {oldest_allowed}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            start_date = report_date
            end_date = report_date
        
        lp_service = LPService(tenant=tenant)
        
        # Recalculate risk scores from latest raw data before generating report
        if report_type == 'weekly':
            from datetime import date as date_type
            current = start_date
            while current <= end_date:
                for store in stores:
                    try:
                        lp_service.calculate_risk_score_for_store(store, current)
                    except Exception:
                        logger.warning('LP recalc failed for store %s date %s', store.id, current)
                current += timedelta(days=1)
            report_data = lp_service.generate_weekly_lp_report_data(stores, start_date, end_date)
            filename = f"LP_Weekly_Risk_Analysis_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.xlsx"
        else:
            for store in stores:
                try:
                    lp_service.calculate_risk_score_for_store(store, report_date)
                except Exception:
                    logger.warning('LP recalc failed for store %s date %s', store.id, report_date)
            report_data = lp_service.generate_lp_report_data(stores, report_date)
            filename = f"LP_Daily_Risk_Analysis_{report_date.strftime('%Y%m%d')}.xlsx"
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LP Risk Analysis"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            'Center Name',
            'High/Low Cash to Credit Ratio',
            'High/Low Tip %',
            'High % of Low-Ticket-Value Services',
            'Locations with 2+ Flags',
            '# of High-Risk Locations',
            'Cash %',
            'Tip Percentage',
            '% of Low-Ticket-Value',
            'Service Sales',
            'Tip Amount',
            'Credit Card',
            'Cash Payment',
            'All Services',
            'Low-Ticket-Services',
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, 16):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Data rows
        high_risk_count = sum(1 for d in report_data if d['is_high_risk'])
        
        for row_idx, data in enumerate(report_data, 2):
            # Center Name
            ws.cell(row=row_idx, column=1, value=data['center_name']).border = thin_border
            
            # Cash Ratio Risk
            cell = ws.cell(row=row_idx, column=2, value=data['cash_ratio_risk'].upper())
            cell.border = thin_border
            if data['cash_ratio_risk'] == 'high':
                cell.fill = red_fill
            elif data['cash_ratio_risk'] == 'medium':
                cell.fill = yellow_fill
            else:
                cell.fill = green_fill
            
            # Tip Percent Risk
            cell = ws.cell(row=row_idx, column=3, value=data['tip_percent_risk'].upper())
            cell.border = thin_border
            if data['tip_percent_risk'] == 'high':
                cell.fill = red_fill
            elif data['tip_percent_risk'] == 'medium':
                cell.fill = yellow_fill
            else:
                cell.fill = green_fill
            
            # Low Ticket Risk
            cell = ws.cell(row=row_idx, column=4, value=data['low_ticket_risk'].upper())
            cell.border = thin_border
            if data['low_ticket_risk'] == 'high':
                cell.fill = red_fill
            elif data['low_ticket_risk'] == 'medium':
                cell.fill = yellow_fill
            else:
                cell.fill = green_fill
            
            # 2+ Flags
            cell = ws.cell(row=row_idx, column=5, value='Yes' if data['has_2plus_flags'] else '')
            cell.border = thin_border
            if data['has_2plus_flags']:
                cell.fill = yellow_fill
            
            # High Risk Count
            ws.cell(row=row_idx, column=6, value=high_risk_count if row_idx == 2 else '').border = thin_border
            
            # Percentages
            ws.cell(row=row_idx, column=7, value=f"{data['cash_percent']:.2f}%").border = thin_border
            ws.cell(row=row_idx, column=8, value=f"{data['tip_percent']:.2f}%").border = thin_border
            ws.cell(row=row_idx, column=9, value=f"{data['low_ticket_percent']:.2f}%").border = thin_border
            
            # Raw values
            ws.cell(row=row_idx, column=10, value=f"${data['service_sales']:,.2f}").border = thin_border
            ws.cell(row=row_idx, column=11, value=f"${data['tip_amount']:,.2f}").border = thin_border
            ws.cell(row=row_idx, column=12, value=f"${data['credit_card']:,.2f}").border = thin_border
            ws.cell(row=row_idx, column=13, value=f"${data['cash_payment']:,.2f}").border = thin_border
            ws.cell(row=row_idx, column=14, value=data['all_services']).border = thin_border
            ws.cell(row=row_idx, column=15, value=data['low_ticket_services']).border = thin_border
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"LP_Risk_Analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class LocationStatesView(APIView):
    """
    Location States API - Returns unique states for filtering locations
    
    Since stores don't have state field yet, we return empty list.
    In production, this would return distinct states from store addresses.
    """
    def get(self, request):
        user = request.user
        if user.role == 'super_admin':
            stores = Store.objects.filter(status='active')
        else:
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=400)
            stores = Store.objects.filter(tenant=tenant, status='active')
        
        # For now, return placeholder states based on store names
        # In production, stores would have a state field
        states = ['CA', 'NV', 'AZ']  # Placeholder based on store locations
        
        return Response({
            'states': states,
            'total': len(states),
        })


class QuickStatsView(APIView):
    """
    Quick Stats API - Returns quick stats for Ezra Sales page
    
    Returns:
    - active_locations: Count of active stores
    - onboarding: Count of stores in onboarding (placeholder)
    - avg_ticket_today: Average ticket for today (using yesterday's data)
    - total_revenue_today: Total revenue for today (using yesterday's data)
    """
    def get(self, request):
        today = get_latest_data_date()  # Use yesterday's data as latest
        
        user = request.user
        if user.role == 'super_admin':
            stores = Store.objects.filter(status='active')
        else:
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=400)
            stores = Store.objects.filter(tenant=tenant, status='active')
        
        active_locations = stores.count()
        
        # Get today's metrics
        if user.role == 'super_admin':
            today_sales = ReportMetric.objects.filter(
                report_date=today,
                report_type='sales'
            )
            today_production = ReportMetric.objects.filter(
                report_date=today,
                report_type='production'
            )
        else:
            today_sales = ReportMetric.objects.filter(
                store__tenant=user.tenant,
                report_date=today,
                report_type='sales'
            )
            today_production = ReportMetric.objects.filter(
                store__tenant=user.tenant,
                report_date=today,
                report_type='production'
            )
        
        # Total revenue today
        revenue_result = today_sales.filter(
            metric_name__in=['Sales Net', 'Total Net']
        ).aggregate(total=Sum('metric_value'))
        total_revenue_today = float(revenue_result['total'] or 0)
        
        # Average ticket today (Service Net / Total Ticket Count)
        service_net_result = today_sales.filter(
            metric_name='Service Net'
        ).aggregate(total=Sum('metric_value'))
        service_net = float(service_net_result['total'] or 0)
        
        tickets_result = today_production.filter(
            metric_name='Total Ticket Count'
        ).aggregate(total=Sum('metric_value'))
        total_tickets = float(tickets_result['total'] or 0)
        
        avg_ticket_today = service_net / total_tickets if total_tickets > 0 else 0
        
        return Response({
            'active_locations': active_locations,
            'onboarding': 0,  # Placeholder - no onboarding status in current model
            'avg_ticket_today': round(avg_ticket_today, 2),
            'total_revenue_today': round(total_revenue_today, 2),
        })


class IngestRawReportView(APIView):
    """
    Ingest raw report data from JSON payload.
    
    - Tenant is determined from X-Tenant-Code header (defaults to 'default')
    - If store doesn't exist, it's created under the tenant
    - All metrics are stored per store (tenant is derived from store)
    - All numeric fields from the JSON are captured as metrics
    
    Optimized for production:
    - Uses bulk_create for metrics (batch inserts)
    - Caches store lookups to minimize DB queries
    - Uses transaction.atomic for data consistency
    """
    authentication_classes = []  # No JWT auth required
    permission_classes = [IsIngestionAPIKey]
    
    # Define all numeric fields to capture (class-level constant)
    NUMERIC_FIELDS = frozenset([
        # Sales report fields
        'Total Net', 'Service Net', 'Product Net', 'Tip Amount', 'CC', 'Cash',
        'Total Sales', 'Service Sales', 'Product Sales', 'Sales Net',
        'Service Discount ', 'Service Refund Value', 'Product Discount', 'Product Refund Value',
        'GIft Card Sales', 'Gift Card Discount', 'Gift Net', 'Total Tax',
        'Opening Cash Balance', 'Paid Out:Petty Cash', 'Paid In', 'Paid Out Cash',
        'Cash Deposit', 'Gift Card Payments', 'Check', 'Closing Balance', 'Over/Short', 'Cash Delta',
        # Production report fields
        'Production Hours ', 'Non-Production Hours', 'Total Hours',
        'Total Guest Count', 'Service Sale $', 'Retail Sale $',
        'Total Sales/Total Hours (TSTH)', 'Total Ticket Count',
        'Total Service Count', 'Ticket Average', 'PPH', 'PPG',
        'Male Customer %', 'Female Customer %', 'Other Customer %',
        'Customer / Hr', 'Cuts', 'Cuts < 5 Mins', 'Cuts > 20 Mins',
        'Cuts / Hr', 'Avg Cut Time', 'Avg Time Between Cuts',
        'Total Sale $', 'Product Sales $ / Hour', 'Service $ / Cuts',
        'Retail %', 'Combo', 'Combo %',
        # Sales-accrual report fields
        'Qty', 'Sales (Exc. Tax)', 'Tax', 'Sales(Inc. Tax)', 'Redeemed', 'Collected', 'Due',
        'Discount', 'Price', 'Taxable Redemption', 'Non Taxable Redemption', 'Service Shop Cost',
        # Attendance report fields
        'Scheduled Hours', 'Actual Hours', 'Serviced Hours', 'Booked Hours',
        'BlockOut Hours (Un-Paid)', 'Actual Breaks Hours',
        'Vacation Hours', 'Holiday Hours', 'Additional Hours',
        'Actual - Schedule (mins)', 'TotalHrs',
        # Business KPI report fields
        'Unique Guest count', 'New Guest Count', 'New Guest %',
        'Invoice Count', 'Rebooking Source Count', 'Rebooking Source %',
        'Services with provider requests', 'Services with provider requests %',
        'Service Sales Count', 'Average service sale per invoice', 'Service Sales Invoice Count',
        'Product Sales Count', 'Average product sale per invoice', 'Product Sales Invoice Count',
        'Membership Sales Count', 'Membership Sales', 'Average membership sale per invoice',
        'Membership Sales Invoice Count', 'Gift Card Sales Count', 'Gift Card Sales',
        'Sale Count', 'Refund Amount', 'Refund Quantity',
        'Open Invoices Count', 'Closed Invoices Count',
        'Center Utilization(By Employee)', 'Available Hours', 'Available hours (Stylist)',
        'Blocked Hours', 'Gift card redeemed', 'Gift card redeemed (cross center)',
        'Membership redeemed', 'Membership redeemed (cross center)',
        'No-show fees', 'Cancellation fees',
        # Performance By Hour report fields (hourly columns)
        '12AM', '1AM', '2AM', '3AM', '4AM', '5AM', '6AM', '7AM',
        '8AM', '9AM', '10AM', '11AM', '12PM', '1PM', '2PM', '3PM',
        '4PM', '5PM', '6PM', '7PM', '8PM', '9PM', '10PM', '11PM',
        # Statutory Pay report fields
        'Days Worked', 'Production hours', 'Blended Production Hourly Rate',
        'Non-production hours', 'Blended NonProduction Hourly Rate',
        'Vacation Hourly Rate', 'Holiday Hourly Rate',
        'Additional Hourly Rate', 'Service Commission', 'Product Commission',
        'Free Service Revenue', 'Service Revenue', 'Product Revenue', 'Other Commission',
    ])
    
    # Skip values for summary rows
    SKIP_VALUES = frozenset(['Grand Total', 'Total', 'Center Total'])
    
    # Batch size for bulk_create
    BATCH_SIZE = 5000
    
    def _parse_numeric_value(self, value):
        """Parse a numeric value from various formats. Rejects values that overflow max_digits=15."""
        if value is None:
            return None
        try:
            if isinstance(value, str):
                value = value.replace('$', '').replace(',', '').replace('%', '').strip()
                if not value:
                    return None
            result = Decimal(str(value))
            # Reject values that overflow DecimalField(max_digits=15, decimal_places=2)
            if result.adjusted() >= 13:  # more than 13 integer digits = overflow
                return None
            return result
        except (ValueError, TypeError, InvalidOperation):
            return None
    
    def _get_or_create_store(self, center_name, tenant, store_cache):
        """Get or create store with caching to minimize DB queries."""
        if center_name in store_cache:
            return store_cache[center_name], False
        
        # Try to find by external_code first
        store = Store.objects.filter(external_code=center_name).first()
        created = False
        
        if not store:
            # Try to find by name
            store = Store.objects.filter(name=center_name).first()
            if store:
                # Update external_code if not set
                if not store.external_code:
                    store.external_code = center_name
                    store.save(update_fields=['external_code'])
            else:
                # Create new store
                store, created = Store.objects.get_or_create(
                    external_code=center_name,
                    defaults={
                        'tenant': tenant,
                        'name': center_name,
                    }
                )
        
        # Cache the store for future lookups
        store_cache[center_name] = store
        return store, created
    
    def post(self, request):
        from django.db import transaction
        import logging
        logger = logging.getLogger(__name__)
        
        tenant_code = request.headers.get('X-Tenant-Code', 'default')
        skip_duplicate_check = request.headers.get('X-Skip-Duplicate-Check', '').lower() == 'true'
        
        # Get or create tenant (case-insensitive lookup)
        tenant = Tenant.objects.filter(code__iexact=tenant_code).first()
        if not tenant:
            tenant, _ = Tenant.objects.get_or_create(
                code=tenant_code.lower(),
                defaults={'name': tenant_code.replace('-', ' ').title()}
            )
        
        serializer = RawReportIngestionSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)
        
        data = serializer.validated_data
        report_type = data['type']
        report_date_str = data['date']
        rows = data.get('rows') or data.get('raw_value') or []
        
        logger.info(f"Ingestion started: type={report_type}, date={report_date_str}, rows={len(rows)}")
        
        # Parse date
        try:
            report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            report_date = datetime.now().date()
        
        try:
            # === Phase 1: Save raw report + metrics (core transaction) ===
            stores_created = 0
            processed_store_ids = set()
            total_metrics_inserted = 0
            store_cache = {}
            
            with transaction.atomic():
                # Check for duplicate report_date (skip if header is set)
                if not skip_duplicate_check:
                    existing_report = RawReport.objects.filter(
                        tenant=tenant,
                        report_type=report_type,
                        report_date=report_date
                    ).first()
                    
                    if existing_report:
                        return Response({
                            'error': f'Report for {report_type} on {report_date} already exists',
                            'existing_report_id': existing_report.id
                        }, status=400)
                
                # Save raw report (cap stored rows for large guest-data reports)
                HANDLER_ONLY_TYPES = frozenset([
                    'manage-guests', 'manage_guests', 'sms-opt-in', 'guest-opt-outs', 'guest_opt_outs',
                ])
                is_handler_only = report_type in HANDLER_ONLY_TYPES
                stored_rows = rows[:100] if (is_handler_only and len(rows) > 100) else rows
                raw_report = RawReport.objects.create(
                    tenant=tenant,
                    report_type=report_type,
                    report_date=report_date,
                    raw_data=stored_rows
                )
                logger.info(f"Raw report saved: id={raw_report.id}, rows_stored={len(stored_rows)}/{len(rows)}")
                
                # Skip Phase 1 metrics extraction for handler-only report types
                if is_handler_only:
                    logger.info(f"Skipping Phase 1 metrics for handler-only type: {report_type}")
                else:
                    # Process rows into metrics
                    metrics_to_create = []
                    current_store = None
                
                    for row in rows:
                        center_name = row.get('Center Name') or row.get('Center Name ') or row.get('Work Center')
                        if center_name and isinstance(center_name, str):
                            center_name = center_name.strip()
                        
                        employee = row.get('Employee ') or row.get('Employee')
                        if employee and isinstance(employee, str):
                            employee = employee.strip()
                        
                        if center_name in self.SKIP_VALUES:
                            continue
                        if row.get('Date') == 'Total':
                            continue
                        if employee in self.SKIP_VALUES:
                            continue
                        
                        if center_name and center_name not in self.SKIP_VALUES:
                            store, created = self._get_or_create_store(center_name, tenant, store_cache)
                            current_store = store
                            if created:
                                stores_created += 1
                        
                        if not current_store:
                            continue
                        
                        processed_store_ids.add(current_store.id)
                        
                        for field in row.keys():
                            if field not in self.NUMERIC_FIELDS:
                                continue
                            
                            value = self._parse_numeric_value(row[field])
                            if value is not None:
                                metrics_to_create.append(ReportMetric(
                                    store=current_store,
                                    report_type=report_type,
                                    report_date=report_date,
                                    metric_name=field.strip(),
                                    metric_value=value
                                ))
                                
                                if len(metrics_to_create) >= self.BATCH_SIZE:
                                    ReportMetric.objects.bulk_create(metrics_to_create, ignore_conflicts=True)
                                    total_metrics_inserted += len(metrics_to_create)
                                    metrics_to_create = []
                    
                    if metrics_to_create:
                        ReportMetric.objects.bulk_create(metrics_to_create, ignore_conflicts=True)
                        total_metrics_inserted += len(metrics_to_create)
                        metrics_to_create = []
                    
                    if processed_store_ids:
                        Store.objects.filter(id__in=processed_store_ids).update(
                            last_synced_at=timezone.now()
                        )
            
            logger.info(f"Phase 1 complete: {total_metrics_inserted} metrics, {stores_created} stores")
            
            # === Phase 2: LP service counts (separate transaction) ===
            handler_result = None
            if report_type == 'sales-accrual':
                try:
                    with transaction.atomic():
                        from .services import LPService
                        lp_service = LPService(tenant)
                        lp_service.process_sales_accrual_service_counts(rows, report_date)
                    logger.info("Phase 2 (LP service counts) complete")
                except Exception as e:
                    logger.exception(f"LP service counts failed (non-fatal): {e}")
            
            # === Phase 3: Exponential/Scheduling handlers (separate transaction) ===
            try:
                with transaction.atomic():
                    from .services.ingestion_handlers import handle_report_ingestion
                    handler_result = handle_report_ingestion(
                        report_type, rows, report_date, tenant, store_cache
                    )
                logger.info(f"Phase 3 (handlers) complete: {handler_result} records")
            except Exception as e:
                logger.exception(f"Handler dispatch failed (non-fatal): {e}")
                handler_result = 0
            
            # === Phase 4: LP risk score calculation (triggered after sales reports) ===
            if report_type in ('sales', 'sales-accrual'):
                try:
                    from .services.lp_service import LPService
                    stores = Store.objects.filter(tenant=tenant, status='active')
                    lp_scores_created = 0
                    for store in stores:
                        lp_service = LPService(tenant=store.tenant)
                        lp_service.calculate_risk_score_for_store(store, report_date)
                        lp_service.generate_alerts_for_store(
                            store,
                            LPRiskScore.objects.filter(store=store, report_date=report_date).first()
                        )
                        lp_scores_created += 1
                    logger.info(f"Phase 4 (LP scores) complete: {lp_scores_created} stores")
                except Exception as e:
                    logger.exception(f"LP score calculation failed (non-fatal): {e}")
            
            return Response({
                'status': 'success',
                'tenant': tenant.name,
                'raw_report_id': raw_report.id,
                'metrics_created': total_metrics_inserted,
                'stores_created': stores_created,
                'handler_records_processed': handler_result,
            }, status=201)
        
        except Exception as e:
            logger.exception(f"Ingestion FAILED for {report_type} on {report_date_str}: {e}")
            return Response({
                'status': 'error',
                'error': str(e),
                'report_type': report_type,
                'date': report_date_str,
                'rows_count': len(rows),
            }, status=500)


class StoreTargetsView(APIView):
    """
    Store Targets API - Get and set revenue/labor targets for stores
    
    GET: List all targets for a specific date
    POST: Create or update target for a store
    """
    def get(self, request):
        user = request.user
        target_date = request.query_params.get('target_date')
        
        if not target_date:
            target_date = datetime.now().date() - timedelta(days=1)
        else:
            target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
        
        # Get stores based on user role
        if user.role == 'super_admin':
            stores = Store.objects.filter(status='active')
        else:
            tenant = user.tenant
            if not tenant:
                return Response({'error': 'No tenant assigned'}, status=400)
            stores = Store.objects.filter(tenant=tenant,status='active')
        
        result = []
        for store in stores:
            result.append({
                'store_id': store.id,
                'store_name': store.name,
                'store_code': store.external_code,
                'city': store.city,
                'state': store.state,
                'daily_revenue_target': float(store.daily_revenue_target or 0),
                'daily_labor_target_hours': float(store.daily_labor_target_hours or 0),
                'has_target': (float(store.daily_revenue_target or 0) > 0) or (float(store.daily_labor_target_hours or 0) > 0),
            })
        
        return Response({
            'target_date': str(target_date),
            'targets': result,
            'total_stores': len(result),
        })
    
    def post(self, request):
        user = request.user
        store_id = request.data.get('store_id')
        target_date = request.data.get('target_date')
        revenue_target = request.data.get('daily_revenue_target', request.data.get('revenue_target', 0))
        labor_target_hours = request.data.get('daily_labor_target_hours', request.data.get('labor_target_hours', 0))

        if not store_id:
            return Response({'error': 'store_id is required'}, status=400)

        if target_date:
            target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
        else:
            target_date = datetime.now().date() - timedelta(days=1)
        
        # Get store and verify access
        try:
            store = Store.objects.get(id=store_id)
        except Store.DoesNotExist:
            return Response({'error': 'Store not found'}, status=404)
        
        # Check tenant access
        if user.role != 'super_admin':
            if not user.tenant or store.tenant_id != user.tenant.id:
                return Response({'error': 'Access denied'}, status=403)
        
        store.daily_revenue_target = Decimal(str(revenue_target))
        store.daily_labor_target_hours = Decimal(str(labor_target_hours))
        store.save(update_fields=['daily_revenue_target', 'daily_labor_target_hours'])

        return Response({
            'store_id': store.id,
            'store_name': store.name,
            'target_date': str(target_date),
            'daily_revenue_target': float(store.daily_revenue_target),
            'daily_labor_target_hours': float(store.daily_labor_target_hours),
            'created': False,
        })


class CurrentUserView(APIView):
    """
    Current User API - Returns the authenticated user's profile including tenant info
    """
    def get(self, request):
        user = request.user
        
        tenant_data = None
        if user.tenant:
            tenant_data = {
                'id': user.tenant.id,
                'name': user.tenant.name,
                'code': user.tenant.code,
            }
        
        return Response({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'role': user.role,
            'tenant': tenant_data,
        })


class IsFranchisorAdminOrSuperAdmin(permissions.BasePermission):
    """Only allow franchisor_admin and super_admin roles."""
    def has_permission(self, request, view):
        return hasattr(request, 'user') and request.user.is_authenticated and \
            request.user.role in ('franchisor_admin', 'super_admin')


class ReportScheduleViewSet(viewsets.ModelViewSet):
    serializer_class = ReportScheduleSerializer
    permission_classes = [permissions.IsAuthenticated, IsFranchisorAdminOrSuperAdmin]

    def get_queryset(self):
        qs = ReportSchedule.objects.all() if self.request.user.role == 'super_admin' \
            else ReportSchedule.objects.filter(tenant=self.request.user.tenant)
        # Active schedules first, then by created_at descending
        return qs.order_by('-is_active', '-created_at')

    def list(self, request, *args, **kwargs):
        """List schedules with optional pagination.
        Query params: page, page_size (default: no pagination for backward compat).
        """
        queryset = self.get_queryset()
        page = request.query_params.get('page')
        page_size = request.query_params.get('page_size')

        if page is not None:
            try:
                page = max(1, int(page))
                page_size = min(100, max(1, int(page_size or 10)))
            except (ValueError, TypeError):
                page = 1
                page_size = 10

            total = queryset.count()
            offset = (page - 1) * page_size
            schedules = queryset[offset:offset + page_size]
            serializer = self.get_serializer(schedules, many=True)
            return Response({
                'results': serializer.data,
                'count': total,
                'page': page,
                'page_size': page_size,
                'total_pages': (total + page_size - 1) // page_size,
            })

        # No pagination — return flat list (backward compatible)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def perform_create(self, serializer):
        tenant = self.request.user.tenant
        if tenant is None:
            # Super admins may not have a tenant; use tenant_id from
            # the request body or fall back to the first active tenant.
            tenant_id = self.request.data.get('tenant_id')
            if tenant_id:
                tenant = Tenant.objects.get(pk=tenant_id)
            else:
                tenant = Tenant.objects.filter(is_active=True).first()
            if tenant is None:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({"tenant": "No active tenant available. Please specify a tenant_id."})
        schedule = serializer.save(
            created_by=self.request.user,
            tenant=tenant,
        )
        try:
            from api.services.cronjob_manager import CronJobManager
            manager = CronJobManager()
            utc_cron = local_cron_to_utc(schedule.cron_expression, schedule.timezone)
            manager.ensure_cronjob(schedule.cron_expression, schedule.timezone, utc_cron)
        except CronJobProvisioningError:
            raise
        except Exception:
            logger.exception(
                "Failed to provision CronJob for schedule %s (cron_key=%s)",
                schedule.pk,
                schedule.cron_key,
            )
            raise CronJobProvisioningError()

    def perform_update(self, serializer):
        old_cron_key = self.get_object().cron_key
        schedule = serializer.save()
        try:
            from api.services.cronjob_manager import CronJobManager
            manager = CronJobManager()
            new_utc_cron = local_cron_to_utc(schedule.cron_expression, schedule.timezone)
            manager.ensure_cronjob(schedule.cron_expression, schedule.timezone, new_utc_cron)
            if schedule.cron_key != old_cron_key:
                old_expr, old_tz = old_cron_key.split('|')
                manager.maybe_remove_cronjob(old_expr, old_tz)
            # Handle is_active toggle
            if 'is_active' in serializer.validated_data:
                if schedule.is_active:
                    manager.unsuspend_cronjob(schedule.cron_expression, schedule.timezone)
                else:
                    manager.suspend_cronjob(schedule.cron_expression, schedule.timezone)
        except CronJobProvisioningError:
            raise
        except Exception:
            logger.exception(
                "Failed to update CronJob for schedule %s (cron_key=%s)",
                schedule.pk,
                schedule.cron_key,
            )
            raise CronJobProvisioningError()

    def perform_destroy(self, instance):
        cron_expr = instance.cron_expression
        tz = instance.timezone
        # Remove CronJob first — if this fails the DB record stays intact
        # so the user can retry.
        try:
            from api.services.cronjob_manager import CronJobManager
            manager = CronJobManager()
            manager.maybe_remove_cronjob(cron_expr, tz, exclude_id=instance.pk)
        except CronJobProvisioningError:
            raise
        except Exception:
            logger.exception(
                "Failed to remove CronJob for cron_key=%s|%s",
                cron_expr,
                tz,
            )
            raise CronJobProvisioningError()
        instance.delete()


class ScheduleTimeOptionsView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsFranchisorAdminOrSuperAdmin]

    def get(self, request):
        options = AppConfig.get_config_value('schedule_time_options', DEFAULT_SCHEDULE_TIME_OPTIONS)
        return Response(options)


# ===========================================
# APP CONFIG - API View
# ===========================================

class AppConfigListView(APIView):
    """
    GET: List all AppConfig entries (super_admin/franchisor_admin only).
    Supports ?name=key to get a specific config.
    """
    def get(self, request):
        name = request.query_params.get('name')
        if name:
            config = AppConfig.get_config_value(name)
            if config is None:
                # Return default if known
                from .constants import (
                    DEFAULT_EXPONENTIAL_CONFIG, DEFAULT_SCHEDULING_CONFIG,
                    DEFAULT_TWILIO_CONFIG, DEFAULT_SMS_TEMPLATES,
                )
                defaults = {
                    'exponential_config': DEFAULT_EXPONENTIAL_CONFIG,
                    'scheduling_config': DEFAULT_SCHEDULING_CONFIG,
                    'twilio_config': DEFAULT_TWILIO_CONFIG,
                    'exponential_sms_templates': DEFAULT_SMS_TEMPLATES,
                }
                config = defaults.get(name)
                if config is None:
                    return Response({'error': f'Config "{name}" not found'}, status=404)
            return Response({'name': name, 'value': config})

        # List all configs
        configs = AppConfig.objects.all().order_by('name')
        result = [{'name': c.name, 'value': c.value, 'updated_at': c.updated_at} for c in configs]
        return Response(result)

    def put(self, request):
        """Update or create an AppConfig entry."""
        user = request.user
        if user.role not in ('super_admin', 'franchisor_admin'):
            return Response({'error': 'Insufficient permissions'}, status=403)

        name = request.data.get('name')
        value = request.data.get('value')
        if not name or value is None:
            return Response({'error': 'Both name and value are required'}, status=400)

        config, created = AppConfig.objects.update_or_create(
            name=name,
            defaults={'value': value}
        )
        return Response({
            'status': 'success',
            'name': config.name,
            'value': config.value,
            'created': created,
        })


# ===========================================
# EZRA EXPONENTIAL - API Views
# ===========================================
# EZRA EXPONENTIAL - Campaign Poller Helpers
# ===========================================

def _ensure_campaign_poller():
    """Create the K8s campaign poller CronJob if scheduled campaigns exist (for recurring)."""
    try:
        from api.services.cronjob_manager import CronJobManager
        manager = CronJobManager()
        manager.ensure_campaign_poller()
    except Exception:
        logger.exception("Failed to ensure campaign poller CronJob (non-fatal)")


def _maybe_remove_campaign_poller(exclude_id=None):
    """Remove the K8s campaign poller CronJob if no scheduled campaigns remain."""
    try:
        from api.services.cronjob_manager import CronJobManager
        manager = CronJobManager()
        manager.maybe_remove_campaign_poller(exclude_id=exclude_id)
    except Exception:
        logger.exception("Failed to check/remove campaign poller CronJob (non-fatal)")


def _ensure_campaign_cronjob(campaign):
    """Create a dedicated K8s CronJob for a scheduled campaign at its exact time."""
    try:
        from api.services.cronjob_manager import CronJobManager
        manager = CronJobManager()
        tz = getattr(campaign, 'campaign_timezone', None) or 'UTC'

        if campaign.is_recurring:
            # For recurring campaigns, build a recurring cron expression
            manager.ensure_recurring_campaign_cronjob(campaign)
        elif campaign.scheduled_at:
            manager.ensure_campaign_cronjob(campaign.id, campaign.scheduled_at, tz)
    except Exception:
        logger.exception("Failed to create campaign CronJob (non-fatal)")


def _remove_campaign_cronjob(campaign_id):
    """Remove the dedicated K8s CronJob for a campaign."""
    try:
        from api.services.cronjob_manager import CronJobManager
        manager = CronJobManager()
        manager.remove_campaign_cronjob(campaign_id)
    except Exception:
        logger.exception("Failed to remove campaign CronJob (non-fatal)")

# ===========================================

class ExponentialOverviewView(APIView):
    """Exponential Overview - Returns customer segmentation and follow-up metrics."""
    def get(self, request):
        from .services.exponential_service import ExponentialService
        from .constants import DEFAULT_EXPONENTIAL_CONFIG
        user = request.user
        tenant = getattr(user, 'tenant', None)
        if tenant is None and getattr(user, 'role', '') != 'super_admin':
            return Response({'error': 'No tenant assigned'}, status=400)

        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        config = AppConfig.get_config_value('exponential_config', DEFAULT_EXPONENTIAL_CONFIG)
        service = ExponentialService(tenant=tenant, config=config)
        overview = service.get_overview_metrics(start_date=start_date_str, end_date=end_date_str)
        overview['config'] = config
        return Response(overview)


class ExponentialCampaignsView(APIView):
    """List campaigns or create a new campaign."""
    def get(self, request):
        from .services.campaign_service import CampaignService
        from .constants import DEFAULT_EXPONENTIAL_CONFIG
        user = request.user
        tenant = getattr(user, 'tenant', None)
        config = AppConfig.get_config_value('exponential_config', DEFAULT_EXPONENTIAL_CONFIG)
        service = CampaignService(tenant=tenant, config=config)

        status = request.query_params.get('status')
        segment = request.query_params.get('segment')
        search = request.query_params.get('search')
        service_filter = request.query_params.get('service_filter')
        page = int(request.query_params.get('page', 0))
        limit = int(request.query_params.get('limit', 20))

        result = service.list_campaigns(status=status, segment=segment, search=search, page=page, limit=limit, service_filter=service_filter)
        stats = service.get_campaign_stats()
        return Response({
            'campaigns': result['campaigns'],
            'total': result['total'],
            'page': result['page'],
            'limit': result['limit'],
            'stats': stats,
        })

    def post(self, request):
        """Create a new campaign."""
        user = request.user
        tenant = user.tenant
        if not tenant and user.role != 'super_admin':
            return Response({'error': 'No tenant assigned'}, status=400)
        if not tenant:
            tenant = Tenant.objects.filter(is_active=True).first()

        data = request.data
        from .services.campaign_service import CampaignService
        from .constants import DEFAULT_EXPONENTIAL_CONFIG
        config = AppConfig.get_config_value('exponential_config', DEFAULT_EXPONENTIAL_CONFIG)
        service = CampaignService(tenant=tenant, config=config)

        campaign = service.create_campaign(data, tenant)

        # Check for validation errors
        if isinstance(campaign, dict) and 'error' in campaign:
            return Response(campaign, status=400)

        # If scheduled, create a dedicated K8s CronJob at the exact time
        if campaign.status == 'scheduled':
            _ensure_campaign_cronjob(campaign)

        # If immediate send, execute right away
        if data.get('schedule_type') == 'immediate' and data.get('execute_now', False):
            result = service.execute_campaign(campaign.id)
            return Response({
                'status': 'success',
                'campaign_id': campaign.id,
                'name': campaign.name,
                'execution': result,
            }, status=201)

        return Response({
            'status': 'success',
            'campaign_id': campaign.id,
            'name': campaign.name,
        }, status=201)

    def delete(self, request):
        """Delete a campaign."""
        campaign_id = request.data.get('campaign_id')
        if not campaign_id:
            return Response({'error': 'campaign_id is required'}, status=400)
        from .services.campaign_service import CampaignService
        user = request.user
        tenant = getattr(user, 'tenant', None)
        service = CampaignService(tenant=tenant)
        result = service.delete_campaign(campaign_id)
        if 'error' in result:
            return Response(result, status=400)
        _remove_campaign_cronjob(campaign_id)
        _maybe_remove_campaign_poller()
        return Response(result, status=204)


class ExponentialCampaignDetailView(APIView):
    """Get campaign detail with message log, update, or delete."""
    def get(self, request, campaign_id):
        from .services.campaign_service import CampaignService
        from .constants import DEFAULT_EXPONENTIAL_CONFIG
        user = request.user
        tenant = getattr(user, 'tenant', None)
        config = AppConfig.get_config_value('exponential_config', DEFAULT_EXPONENTIAL_CONFIG)
        service = CampaignService(tenant=tenant, config=config)

        campaign = service.get_campaign_detail(campaign_id)
        if not campaign:
            return Response({'error': 'Campaign not found'}, status=404)

        status = request.query_params.get('message_status')
        search = request.query_params.get('search')
        page = int(request.query_params.get('page', 0))
        limit = int(request.query_params.get('limit', 50))

        messages = service.get_campaign_messages(
            campaign_id, status=status, search=search, page=page, limit=limit
        )

        return Response({
            'campaign': campaign,
            'messages': messages['messages'],
            'messagesTotal': messages['total'],
            'messagesPage': messages['page'],
        })

    def put(self, request, campaign_id):
        """Update a scheduled campaign."""
        user = request.user
        if user.role not in ('super_admin', 'franchisor_admin'):
            return Response({'error': 'Insufficient permissions'}, status=403)

        from .services.campaign_service import CampaignService
        from .constants import DEFAULT_EXPONENTIAL_CONFIG
        tenant = getattr(user, 'tenant', None)
        config = AppConfig.get_config_value('exponential_config', DEFAULT_EXPONENTIAL_CONFIG)
        service = CampaignService(tenant=tenant, config=config)

        result = service.update_campaign(campaign_id, request.data)
        if isinstance(result, dict) and 'error' in result:
            return Response(result, status=400)

        # If campaign is now scheduled, create dedicated CronJob; otherwise remove
        campaign_status = result.get('status', '') if isinstance(result, dict) else ''
        if campaign_status == 'scheduled':
            # Need the actual campaign object for scheduled_at and timezone
            try:
                from api.models import ExponentialCampaign
                camp_obj = ExponentialCampaign.objects.get(id=campaign_id)
                _ensure_campaign_cronjob(camp_obj)
            except ExponentialCampaign.DoesNotExist:
                pass
        else:
            _remove_campaign_cronjob(campaign_id)
            _maybe_remove_campaign_poller()

        return Response({'status': 'success', 'campaign': result})

    def delete(self, request, campaign_id):
        """Delete a campaign by ID."""
        from .services.campaign_service import CampaignService
        user = request.user
        tenant = getattr(user, 'tenant', None)
        service = CampaignService(tenant=tenant)
        result = service.delete_campaign(campaign_id)
        if 'error' in result:
            return Response(result, status=400)
        _remove_campaign_cronjob(campaign_id)
        _maybe_remove_campaign_poller()
        return Response(result, status=204)


class ExponentialCampaignExecuteView(APIView):
    """Execute (send) a campaign."""
    def post(self, request, campaign_id):
        user = request.user
        if user.role not in ('super_admin', 'franchisor_admin'):
            return Response({'error': 'Insufficient permissions'}, status=403)

        from .services.campaign_service import CampaignService
        from .constants import DEFAULT_EXPONENTIAL_CONFIG
        tenant = getattr(user, 'tenant', None)
        config = AppConfig.get_config_value('exponential_config', DEFAULT_EXPONENTIAL_CONFIG)
        service = CampaignService(tenant=tenant, config=config)

        result = service.execute_campaign(campaign_id)
        if 'error' in result:
            return Response(result, status=400)

        # For recurring campaigns: reset to scheduled if still within date range
        try:
            campaign = ExponentialCampaign.objects.get(id=campaign_id)
            if campaign.is_recurring and campaign.recurring_end_date:
                from django.utils import timezone as tz
                today = tz.now().date()
                if today <= campaign.recurring_end_date:
                    campaign.status = 'scheduled'
                    campaign.last_recurring_run = tz.now()
                    campaign.save(update_fields=['status', 'last_recurring_run'])
                    # Keep the CronJob alive for future runs
                    return Response(result)
        except ExponentialCampaign.DoesNotExist:
            pass

        # Non-recurring or past end date — clean up CronJob
        _remove_campaign_cronjob(campaign_id)
        _maybe_remove_campaign_poller()

        return Response(result)


class StoreDataUploadView(APIView):
    """Upload store display names and addresses from CSV/Excel."""

    def get(self, request):
        """Return current store data for the tenant."""
        tenant = getattr(request.user, 'tenant', None)
        if not tenant:
            return Response({'error': 'No tenant'}, status=400)
        stores = Store.objects.filter(tenant=tenant).order_by('name')
        return Response({'stores': [
            {
                'id': s.id, 'name': s.name, 'externalCode': s.external_code or '',
                'displayName': s.display_name or '', 'address': s.address or '',
                'bookingLink': s.booking_link or '',
                'city': s.city or '', 'state': s.state or '',
            } for s in stores
        ]})

    def post(self, request):
        """Upload store data. Accepts JSON array: [{name, displayName, address, bookingLink}]"""
        tenant = getattr(request.user, 'tenant', None)
        if not tenant:
            return Response({'error': 'No tenant'}, status=400)

        stores_data = request.data.get('stores', [])
        if not stores_data:
            return Response({'error': 'No store data provided'}, status=400)

        # Single query: load all tenant stores into lookup dicts
        all_stores = list(Store.objects.filter(tenant=tenant))
        by_code = {s.external_code: s for s in all_stores if s.external_code}
        by_name = {s.name: s for s in all_stores}

        to_update = []
        not_found = 0
        for item in stores_data:
            store_name = (item.get('name') or '').strip()
            external_code = (item.get('externalCode') or '').strip()
            display_name = (item.get('displayName') or '').strip()
            address = (item.get('address') or '').strip()
            booking_link = (item.get('bookingLink') or '').strip()

            if not store_name and not external_code:
                continue

            store = by_code.get(external_code) if external_code else None
            if not store and store_name:
                store = by_name.get(store_name)

            if store:
                changed = False
                if display_name and store.display_name != display_name:
                    store.display_name = display_name
                    changed = True
                if address and store.address != address:
                    store.address = address
                    changed = True
                if booking_link and store.booking_link != booking_link:
                    store.booking_link = booking_link
                    changed = True
                if changed:
                    to_update.append(store)
            else:
                not_found += 1

        updated = 0
        if to_update:
            updated = Store.objects.bulk_update(to_update, ['display_name', 'address', 'booking_link'], batch_size=100)

        return Response({'updated': updated or len(to_update), 'notFound': not_found})


class SegmentConfigView(APIView):
    """CRUD for customizable customer segment definitions per tenant."""

    def get(self, request):
        tenant = self._get_tenant(request)
        from api.models import SegmentConfig
        if tenant:
            configs = SegmentConfig.get_for_tenant(tenant)
        else:
            configs = []
            for d in SegmentConfig.get_defaults():
                configs.append(type('Obj', (), {**d, 'id': None, 'is_active': True})())
        return Response({'segments': [
            {
                'id': getattr(c, 'id', None), 'name': c.name, 'slug': c.slug,
                'minDays': c.min_days, 'maxDays': c.max_days,
                'riskLevel': c.risk_level, 'color': c.color,
                'sortOrder': c.sort_order, 'isActive': c.is_active,
            } for c in configs
        ]})

    def _get_tenant(self, request):
        """Get tenant from user, or Default tenant for superusers."""
        tenant = getattr(request.user, 'tenant', None)
        if not tenant:
            from api.models import Tenant
            tenant = Tenant.objects.filter(name='Default').first()
        return tenant

    def post(self, request):
        tenant = self._get_tenant(request)
        if not tenant:
            return Response({'error': 'No tenant available'}, status=400)
        from api.models import SegmentConfig
        data = request.data
        config_id = data.get('id')
        slug = data.get('slug', '').strip()
        if not data.get('name') or data.get('minDays') is None:
            return Response({'error': 'name and minDays are required'}, status=400)

        defaults = {
            'name': data['name'],
            'slug': slug or data['name'].lower().replace(' ', '_'),
            'min_days': data['minDays'],
            'max_days': data.get('maxDays'),
            'risk_level': data.get('riskLevel', 'medium'),
            'color': data.get('color', 'warning'),
            'sort_order': data.get('sortOrder', 0),
            'is_active': data.get('isActive', True),
        }

        if config_id:
            # Update existing by id
            try:
                config = SegmentConfig.objects.get(id=config_id, tenant=tenant)
                for k, v in defaults.items():
                    setattr(config, k, v)
                config.save()
                return Response({'id': config.id, 'created': False})
            except SegmentConfig.DoesNotExist:
                pass

        # Create new or upsert by slug
        if not slug:
            slug = data['name'].lower().replace(' ', '_')
        config, created = SegmentConfig.objects.update_or_create(
            tenant=tenant, slug=slug, defaults=defaults
        )
        return Response({'id': config.id, 'created': created})

    def delete(self, request):
        tenant = self._get_tenant(request)
        if not tenant:
            return Response({'error': 'No tenant'}, status=400)
        from api.models import SegmentConfig
        config_id = request.data.get('id') or request.query_params.get('id')
        if not config_id:
            return Response({'error': 'id is required'}, status=400)
        deleted, _ = SegmentConfig.objects.filter(id=config_id, tenant=tenant).delete()
        return Response({'deleted': deleted})

    def put(self, request):
        """Full replace: delete all existing segments and create new ones."""
        tenant = self._get_tenant(request)
        if not tenant:
            return Response({'error': 'No tenant'}, status=400)
        from api.models import SegmentConfig
        segments = request.data.get('segments', [])
        # Delete all existing
        SegmentConfig.objects.filter(tenant=tenant).delete()
        # Create new
        created = []
        for i, data in enumerate(segments):
            if not data.get('name'):
                continue
            slug = data.get('slug') or data['name'].lower().replace(' ', '_')
            cfg = SegmentConfig.objects.create(
                tenant=tenant, name=data['name'], slug=slug,
                min_days=data.get('minDays', 0), max_days=data.get('maxDays'),
                risk_level=data.get('riskLevel', 'medium'), color=data.get('color', 'warning'),
                sort_order=data.get('sortOrder', i), is_active=data.get('isActive', True),
            )
            created.append(cfg.id)
        return Response({'created': len(created)})


class ExponentialLocationsView(APIView):
    """Location list for Exponential (for audience selection)."""
    def get(self, request):
        from django.db.models import Count
        user = request.user
        tenant = getattr(user, 'tenant', None)
        qs = Store.objects.filter(status='active')
        if tenant:
            qs = qs.filter(tenant=tenant)
        qs = qs.annotate(customer_count=Count('exp_customers')).order_by('name')
        locations = [
            {
                'id': str(s.id),
                'name': s.name or '',
                'storeCode': s.external_code or '',
                'state': s.state or '',
                'city': s.city or '',
                'customerCount': s.customer_count,
            }
            for s in qs
        ]
        total_customers = sum(l['customerCount'] for l in locations)
        return Response({'locations': locations, 'totalCustomers': total_customers})


class ExponentialGuestsView(APIView):
    """Paginated guest list with segment info."""
    def get(self, request):
        from .services.campaign_service import CampaignService
        from .constants import DEFAULT_EXPONENTIAL_CONFIG
        user = request.user
        tenant = getattr(user, 'tenant', None)
        config = AppConfig.get_config_value('exponential_config', DEFAULT_EXPONENTIAL_CONFIG)
        service = CampaignService(tenant=tenant, config=config)

        store_id = request.query_params.get('store_id')
        location_ids_raw = request.query_params.get('location_ids', '')
        location_ids = [x.strip() for x in location_ids_raw.split(',') if x.strip()] if location_ids_raw else None
        bucket = request.query_params.get('bucket')
        search = request.query_params.get('search')
        source = request.query_params.get('source')
        sms_status = request.query_params.get('sms_status')
        last_service = request.query_params.get('last_service', '')
        last_services_raw = request.query_params.get('last_services', '')
        # Support both single and multi service params (OR filter)
        service_list = []
        if last_services_raw:
            service_list = [s.strip() for s in last_services_raw.split(',') if s.strip()]
        elif last_service:
            service_list = [s.strip() for s in last_service.split(',') if s.strip()]
        guest_type = request.query_params.get('guest_type')
        sort = request.query_params.get('sort')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        page = int(request.query_params.get('page', 0))
        limit = int(request.query_params.get('limit', 20))

        # Resolve store_id vs location_ids
        effective_store_id = store_id
        effective_location_ids = location_ids
        if not effective_store_id and effective_location_ids and len(effective_location_ids) == 1:
            effective_store_id = effective_location_ids[0]
            effective_location_ids = None

        return Response(service.get_guest_list(
            store_id=effective_store_id, location_ids=effective_location_ids,
            bucket=bucket, search=search, page=page, limit=limit,
            source=source, sms_status=sms_status, last_service=service_list or None,
            guest_type=guest_type, sort=sort, date_from=date_from, date_to=date_to,
        ))


class ExponentialServiceTypesView(APIView):
    """Return distinct service types across all customers for the tenant."""
    def get(self, request):
        tenant = getattr(request.user, 'tenant', None)
        if not tenant:
            return Response({'services': []})
        services = (
            ExponentialCustomer.objects
            .filter(tenant=tenant)
            .exclude(last_service__isnull=True)
            .exclude(last_service='')
            .values_list('last_service', flat=True)
            .distinct()
            .order_by('last_service')
        )
        return Response({'services': list(services)})


class ExponentialGuestImportView(APIView):
    """Import CRM guests from CSV/Excel file upload (up to 10MB) or JSON data.
    
    Direct import — auto-maps columns using our standard template format.
    For custom CRM exports, use the 2-step flow:
      POST /api/exponential/guests/import/parse/  → returns headers + mapping UI data
      POST /api/exponential/guests/import/map/     → applies mapping and imports
    """
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        user = request.user
        if user.role not in ('super_admin', 'franchisor_admin'):
            return Response({'error': 'Insufficient permissions'}, status=403)

        from .services.campaign_service import CampaignService
        tenant = user.tenant
        if not tenant and user.role == 'super_admin':
            tenant = Tenant.objects.filter(is_active=True).first()

        service = CampaignService(tenant=tenant)

        # Check if this is a file upload (multipart) or JSON
        uploaded_file = request.FILES.get('file')
        if uploaded_file:
            result = service.import_guests_from_file(uploaded_file, user=user)
            if 'error' in result and isinstance(result.get('error'), str):
                return Response(result, status=400)
            return Response(result, status=201)

        # Fallback: JSON body with guests array
        guests_data = request.data.get('guests', [])
        if not guests_data:
            return Response({'error': 'No file or guest data provided. Upload a CSV/Excel file or send JSON guests array.'}, status=400)

        result = service.import_guests(guests_data)
        return Response(result, status=201)


class GuestImportParseView(APIView):
    """
    Step 1: Upload a file and get detected headers + suggested mapping.
    User sees their columns on the left, our fields on the right.
    Returns import_id to use in step 2.
    """
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        user = request.user
        if user.role not in ('super_admin', 'franchisor_admin'):
            return Response({'error': 'Insufficient permissions'}, status=403)

        from .services.campaign_service import CampaignService
        tenant = user.tenant
        if not tenant and user.role == 'super_admin':
            tenant = Tenant.objects.filter(is_active=True).first()

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'No file provided'}, status=400)

        service = CampaignService(tenant=tenant)
        result = service.parse_file_headers(uploaded_file)

        if 'error' in result:
            return Response(result, status=400)
        return Response(result)


class GuestImportMapView(APIView):
    """
    Step 2: Submit column mapping and trigger the actual import.
    Expects JSON: { "import_id": 123, "column_mapping": {"Client Name": "first_name", ...} }
    """
    def post(self, request):
        user = request.user
        if user.role not in ('super_admin', 'franchisor_admin'):
            return Response({'error': 'Insufficient permissions'}, status=403)

        from .services.campaign_service import CampaignService
        tenant = user.tenant
        if not tenant and user.role == 'super_admin':
            tenant = Tenant.objects.filter(is_active=True).first()

        import_id = request.data.get('import_id')
        column_mapping = request.data.get('column_mapping', {})

        if not import_id:
            return Response({'error': 'import_id is required'}, status=400)
        if not column_mapping:
            return Response({'error': 'column_mapping is required'}, status=400)

        # Validate required fields are mapped
        mapped_targets = set(column_mapping.values())
        if 'first_name' not in mapped_targets:
            return Response({'error': 'first_name must be mapped to a column'}, status=400)
        if 'phone' not in mapped_targets:
            return Response({'error': 'phone must be mapped to a column'}, status=400)

        service = CampaignService(tenant=tenant)
        result = service.import_with_mapping(import_id, column_mapping, user=user)

        if 'error' in result:
            return Response(result, status=400)
        return Response(result, status=201)


class GuestImportListView(APIView):
    """List guest import history for the tenant."""
    def get(self, request):
        from .services.campaign_service import CampaignService
        user = request.user
        tenant = getattr(user, 'tenant', None)

        page = int(request.query_params.get('page', 0))
        limit = int(request.query_params.get('limit', 20))

        service = CampaignService(tenant=tenant)
        return Response(service.list_imports(page=page, limit=limit))


class ExponentialAudienceEstimateView(APIView):
    """Estimate audience size for campaign wizard."""
    def get(self, request):
        from .services.campaign_service import CampaignService
        from .constants import DEFAULT_EXPONENTIAL_CONFIG
        user = request.user
        tenant = getattr(user, 'tenant', None)
        config = AppConfig.get_config_value('exponential_config', DEFAULT_EXPONENTIAL_CONFIG)
        service = CampaignService(tenant=tenant, config=config)

        segment = request.query_params.get('segment')
        location_ids = request.query_params.get('location_ids', '')
        sms_status = request.query_params.get('sms_status')
        last_service = request.query_params.get('last_service', '')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        loc_ids = [int(x) for x in location_ids.split(',') if x.strip()] if location_ids else None
        # Support comma-separated service list (OR filter)
        service_list = [s.strip() for s in last_service.split(',') if s.strip()] if last_service else None

        counts = service.estimate_audience(segment=segment, location_ids=loc_ids, sms_status=sms_status, last_service=service_list, date_from=date_from, date_to=date_to)
        return Response({'counts': counts})


class ExponentialUptakeView(APIView):
    """Uptake effectiveness by segment."""
    def get(self, request):
        from .services.exponential_service import ExponentialService
        from .constants import DEFAULT_EXPONENTIAL_CONFIG
        user = request.user
        tenant = getattr(user, 'tenant', None)
        config = AppConfig.get_config_value('exponential_config', DEFAULT_EXPONENTIAL_CONFIG)
        service = ExponentialService(tenant=tenant, config=config)
        store_ids = list(service._store_qs().values_list('id', flat=True))
        return Response(service._build_uptake_by_segment(store_ids))


class ExponentialStoreDrilldownView(APIView):
    """Exponential Store Drilldown - per-store segment details, campaigns, recommendations."""
    def get(self, request, store_id):
        from .services.exponential_service import ExponentialService
        from .constants import DEFAULT_EXPONENTIAL_CONFIG
        user = request.user
        tenant = getattr(user, 'tenant', None)
        if tenant is None and getattr(user, 'role', '') != 'super_admin':
            return Response({'error': 'No tenant assigned'}, status=400)

        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        config = AppConfig.get_config_value('exponential_config', DEFAULT_EXPONENTIAL_CONFIG)
        service = ExponentialService(tenant=tenant, config=config)
        result = service.get_store_drilldown(store_id, start_date=start_date_str, end_date=end_date_str)
        if result is None:
            return Response({'error': 'Store not found'}, status=404)
        return Response(result)


class ExponentialCampaignStatusView(APIView):
    """Update campaign status (activate, pause, complete)."""
    def put(self, request, campaign_id):
        user = request.user
        if user.role not in ('super_admin', 'franchisor_admin'):
            return Response({'error': 'Insufficient permissions'}, status=403)

        new_status = request.data.get('status')
        if new_status not in ('active', 'paused', 'completed', 'scheduled'):
            return Response({'error': 'Invalid status'}, status=400)

        tenant = getattr(user, 'tenant', None)
        try:
            qs = ExponentialCampaign.objects.all()
            if tenant:
                qs = qs.filter(tenant=tenant)
            campaign = qs.get(id=campaign_id)
        except ExponentialCampaign.DoesNotExist:
            return Response({'error': 'Campaign not found'}, status=404)

        old_status = campaign.status
        campaign.status = new_status

        if new_status == 'active' and not campaign.started_at:
            campaign.started_at = timezone.now()
        elif new_status == 'completed':
            campaign.completed_at = timezone.now()

        campaign.save(update_fields=['status', 'started_at', 'completed_at'])

        # Manage poller CronJob based on status change
        # Manage CronJob based on status change
        if new_status == 'scheduled':
            _ensure_campaign_cronjob(campaign)
        elif old_status == 'scheduled':
            _remove_campaign_cronjob(campaign.id)
            _maybe_remove_campaign_poller()

        return Response({
            'status': 'success',
            'campaign_id': campaign.id,
            'old_status': old_status,
            'new_status': new_status,
        })


class ExponentialCampaignSyncStatusesView(APIView):
    """Fetch current SMS statuses from Twilio and update our DB records."""
    def post(self, request, campaign_id):
        user = request.user
        if user.role not in ('super_admin', 'franchisor_admin'):
            return Response({'error': 'Insufficient permissions'}, status=403)

        from .services.twilio_sms_service import sync_campaign_statuses
        tenant = getattr(user, 'tenant', None)
        result = sync_campaign_statuses(campaign_id, tenant=tenant)

        if 'error' in result:
            return Response(result, status=400)
        return Response(result)


class TwilioValidateView(APIView):
    """Validate Twilio configuration."""
    def get(self, request):
        user = request.user
        if user.role not in ('super_admin', 'franchisor_admin'):
            return Response({'error': 'Insufficient permissions'}, status=403)

        from .services.twilio_sms_service import validate_twilio_config
        result = validate_twilio_config()
        return Response(result)


class TwilioSendTestView(APIView):
    """Send a test SMS to verify Twilio integration works end-to-end."""

    def post(self, request):
        user = request.user
        if user.role not in ('super_admin', 'franchisor_admin'):
            return Response({'error': 'Insufficient permissions'}, status=403)

        to_number = request.data.get('to_number')
        message = request.data.get('message', 'Hello from Ezra! This is a test SMS.')

        if not to_number:
            return Response(
                {'error': 'to_number is required (E.164 format, e.g. +15551234567)'},
                status=400,
            )

        from .services.twilio_sms_service import send_sms
        result = send_sms(to_number=to_number, message_body=message)
        status_code = 200 if result['success'] else 400
        return Response(result, status=status_code)


class SMSTemplatePreviewView(APIView):
    """Preview a rendered SMS template with sample data."""

    def post(self, request):
        body = request.data.get('body', '')
        guest_name = request.data.get('guest_name', 'Jane Doe')
        store_name = request.data.get('store_name', 'Glow Spa Downtown')
        coupon_value = request.data.get('coupon_value', '15')
        coupon_code = request.data.get('coupon_code', 'SAVE15')
        booking_link = request.data.get('booking_link', 'https://book.example.com')

        if not body:
            return Response({'error': 'body is required'}, status=400)

        first_name = guest_name.split()[0] if guest_name else 'Jane'

        try:
            rendered = body.format(
                first_name=first_name,
                guest_name=guest_name,
                location_name=store_name,
                store_name=store_name,
                coupon_value=coupon_value,
                coupon_code=coupon_code,
                booking_link=booking_link,
            )
        except (KeyError, IndexError) as e:
            return Response(
                {'error': f'Template rendering failed: {str(e)}', 'hint': 'Valid placeholders: {first_name}, {guest_name}, {store_name}, {location_name}, {coupon_value}, {coupon_code}, {booking_link}'},
                status=400,
            )

        return Response({
            'rendered': rendered,
            'char_count': len(rendered),
            'sms_segments': (len(rendered) // 160) + (1 if len(rendered) % 160 else 0),
        })


class TwilioStatusCallbackView(APIView):
    """
    Webhook endpoint for Twilio SMS delivery status callbacks.
    Twilio POSTs status updates here as messages progress through delivery.

    Security: Twilio signature validation is ALWAYS enforced.
    Requests without a valid X-Twilio-Signature header are rejected with 403.
    """
    permission_classes = [permissions.AllowAny]
    authentication_classes = []  # No auth required for webhooks

    def post(self, request):
        from .services.twilio_sms_service import handle_status_callback, validate_twilio_signature, get_status_callback_url

        twilio_signature = request.META.get('HTTP_X_TWILIO_SIGNATURE', '')

        if not twilio_signature:
            logger.warning("Twilio status callback rejected: missing X-Twilio-Signature")
            return Response({'error': 'Missing Twilio signature'}, status=403)

        # Use the configured public callback URL for validation, not the internal URL.
        # Behind a reverse proxy/ingress, build_absolute_uri() returns the internal
        # pod URL (e.g. http://backend:8000/...) which doesn't match what Twilio
        # signed against (the public URL like https://dev-api-meetezra.hachiai.com/...).
        url = get_status_callback_url() or request.build_absolute_uri()
        params = request.POST.dict()
        if not validate_twilio_signature(url, params, twilio_signature):
            logger.warning(
                "Twilio status callback rejected: invalid signature "
                f"(validated against url={url})"
            )
            return Response({'error': 'Invalid signature'}, status=403)

        # Process the callback
        data = request.POST.dict() if request.POST else request.data
        result = handle_status_callback(data)

        if 'error' in result:
            return Response(result, status=400)
        return Response(result)


class TwilioMessageLookupView(APIView):
    """Look up the current status of a specific SMS message from Twilio."""

    def get(self, request, message_sid):
        user = request.user
        if user.role not in ('super_admin', 'franchisor_admin'):
            return Response({'error': 'Insufficient permissions'}, status=403)

        from .services.twilio_sms_service import fetch_message_status
        result = fetch_message_status(message_sid)
        if result is None:
            return Response({'error': 'Could not fetch message status'}, status=400)
        return Response(result)


class TwilioInboundSMSView(APIView):
    """
    Webhook endpoint for inbound SMS from Twilio.
    Handles opt-out (STOP) and opt-in (START) keywords.
    Configure this URL as the "A MESSAGE COMES IN" webhook on your Twilio phone number.
    """
    permission_classes = [permissions.AllowAny]
    authentication_classes = []

    def post(self, request):
        from .services.twilio_sms_service import handle_inbound_sms, validate_twilio_signature, get_inbound_sms_url

        twilio_signature = request.META.get('HTTP_X_TWILIO_SIGNATURE', '')
        if not twilio_signature:
            logger.warning("Twilio inbound SMS rejected: missing X-Twilio-Signature")
            return Response({'error': 'Missing Twilio signature'}, status=403)

        url = get_inbound_sms_url() or request.build_absolute_uri()
        params = request.POST.dict()
        if not validate_twilio_signature(url, params, twilio_signature):
            logger.warning("Twilio inbound SMS rejected: invalid signature")
            return Response({'error': 'Invalid signature'}, status=403)

        data = request.POST.dict() if request.POST else request.data
        result = handle_inbound_sms(data)

        # Twilio expects a TwiML response (empty is fine — no auto-reply)
        from django.http import HttpResponse
        return HttpResponse('<Response></Response>', content_type='text/xml')



# ===========================================
# EZRA SCHEDULING - API Views
# ===========================================

class SchedulingReportDownloadView(APIView):
    """Download Scheduling report as Excel with location rankings."""

    def get(self, request):
        from django.http import HttpResponse
        from .services.scheduling_service import SchedulingService
        from .constants import DEFAULT_SCHEDULING_CONFIG
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        user = request.user
        tenant = getattr(user, 'tenant', None)
        if tenant is None and user.role != 'super_admin':
            return Response({'error': 'No tenant assigned'}, status=400)

        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        config = AppConfig.get_config_value('scheduling_config', DEFAULT_SCHEDULING_CONFIG)
        service = SchedulingService(tenant=tenant, config=config)
        rankings = service.get_store_rankings(tenant=tenant, start_date=start_date_str, end_date=end_date_str)

        # Build Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Scheduling Report'

        headers = [
            'Rank', 'Location', 'Store Code', 'State', 'Revenue', 'Labor Hours',
            'Idle Hours', 'Idle %', 'TSTH', 'Tix/Hr', 'Labor $', 'OT Hours', 'OT Flag',
        ]
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        widths = [6, 35, 15, 8, 14, 12, 12, 10, 10, 10, 12, 10, 8]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        red_font = Font(color='FF0000', bold=True)
        for rank, loc in enumerate(rankings, 1):
            row = rank + 1
            ws.cell(row=row, column=1, value=rank).border = thin_border
            ws.cell(row=row, column=2, value=loc['locationName']).border = thin_border
            ws.cell(row=row, column=3, value=loc['storeCode']).border = thin_border
            ws.cell(row=row, column=4, value=loc['state']).border = thin_border
            ws.cell(row=row, column=5, value=float(loc['revenue'])).border = thin_border
            ws.cell(row=row, column=5).number_format = '$#,##0'
            ws.cell(row=row, column=6, value=float(loc['laborHours'])).border = thin_border
            ws.cell(row=row, column=6).number_format = '#,##0.0'
            ws.cell(row=row, column=7, value=float(loc['idleHours'])).border = thin_border
            ws.cell(row=row, column=7).number_format = '#,##0.0'
            pct_cell = ws.cell(row=row, column=8, value=float(loc['idlePercent']) / 100)
            pct_cell.number_format = '0.0%'
            pct_cell.border = thin_border
            if loc['idlePercent'] >= 75:
                pct_cell.font = red_font
            ws.cell(row=row, column=9, value=float(loc['tsth'])).border = thin_border
            ws.cell(row=row, column=9).number_format = '$#,##0'
            ws.cell(row=row, column=10, value=float(loc['ticketsPerLaborHour'])).border = thin_border
            ws.cell(row=row, column=10).number_format = '0.0'
            ws.cell(row=row, column=11, value=float(loc['laborCost'])).border = thin_border
            ws.cell(row=row, column=11).number_format = '$#,##0'
            ws.cell(row=row, column=12, value=float(loc['overtimeHours'])).border = thin_border
            ws.cell(row=row, column=12).number_format = '#,##0.0'
            ot_cell = ws.cell(row=row, column=13, value='OT' if loc['hasOvertimeFlag'] else '')
            ot_cell.border = thin_border
            if loc['hasOvertimeFlag']:
                ot_cell.font = red_font

        # Determine filename
        s = start_date_str or ''
        e = end_date_str or ''
        filename = f'Scheduling_Report_{s}_to_{e}.xlsx' if s and e else 'Scheduling_Report.xlsx'

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ExponentialReportDownloadView(APIView):
    """Download Exponential report as Excel with campaign data and location summaries."""

    def get(self, request):
        from django.http import HttpResponse
        from .services.exponential_service import ExponentialService
        from .constants import DEFAULT_EXPONENTIAL_CONFIG
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        user = request.user
        tenant = getattr(user, 'tenant', None)
        if tenant is None and user.role != 'super_admin':
            return Response({'error': 'No tenant assigned'}, status=400)

        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        config = AppConfig.get_config_value('exponential_config', DEFAULT_EXPONENTIAL_CONFIG)
        service = ExponentialService(tenant=tenant, config=config)
        overview = service.get_overview_metrics(start_date=start_date_str, end_date=end_date_str)

        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # --- Sheet 1: Location Summary ---
        ws = wb.active
        ws.title = 'Location Summary'
        loc_headers = [
            'Rank', 'Location', 'Store Code', 'State', 'Guests MTD',
            'Customers Last Mo.', '4-Week', '6-Week', '8-Week',
            'Follow-ups Sent', 'Uptake %', 'Retention Risk',
        ]
        for col_idx, h in enumerate(loc_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        loc_widths = [6, 35, 15, 8, 12, 16, 10, 10, 10, 14, 10, 14]
        for i, w in enumerate(loc_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        red_font = Font(color='FF0000', bold=True)
        for rank, loc in enumerate(overview.get('locationSummaries', []), 1):
            row = rank + 1
            ws.cell(row=row, column=1, value=rank).border = thin_border
            ws.cell(row=row, column=2, value=loc.get('locationName', '')).border = thin_border
            ws.cell(row=row, column=3, value=loc.get('storeCode', '')).border = thin_border
            ws.cell(row=row, column=4, value=loc.get('state', '')).border = thin_border
            ws.cell(row=row, column=5, value=loc.get('guestsMTD', 0)).border = thin_border
            ws.cell(row=row, column=6, value=loc.get('customersLastMonth', 0)).border = thin_border
            ws.cell(row=row, column=7, value=loc.get('fourWeekCount', 0)).border = thin_border
            ws.cell(row=row, column=8, value=loc.get('sixWeekCount', 0)).border = thin_border
            ws.cell(row=row, column=9, value=loc.get('eightWeekCount', 0)).border = thin_border
            ws.cell(row=row, column=10, value=loc.get('followUpsSent', 0)).border = thin_border
            pct_cell = ws.cell(row=row, column=11, value=loc.get('overallUptake', 0) / 100)
            pct_cell.number_format = '0.0%'
            pct_cell.border = thin_border
            risk_cell = ws.cell(row=row, column=12, value=loc.get('retentionRiskScore', 0))
            risk_cell.border = thin_border
            if loc.get('retentionRiskScore', 0) >= 50:
                risk_cell.font = red_font

        # --- Sheet 2: Segment Summary ---
        ws2 = wb.create_sheet('Segment Summary')
        seg_headers = ['Segment', 'Customers', 'Risk Level', 'Messages Sent', 'Uptake %']
        for col_idx, h in enumerate(seg_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        seg_widths = [15, 12, 12, 14, 10]
        for i, w in enumerate(seg_widths, 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for idx, seg in enumerate(overview.get('segments', []), 1):
            row = idx + 1
            ws2.cell(row=row, column=1, value=seg.get('name', '')).border = thin_border
            ws2.cell(row=row, column=2, value=seg.get('customerCount', 0)).border = thin_border
            ws2.cell(row=row, column=3, value=seg.get('riskLevel', '')).border = thin_border
            ws2.cell(row=row, column=4, value=seg.get('messagesSent', 0)).border = thin_border
            pct_cell = ws2.cell(row=row, column=5, value=seg.get('uptakePercent', 0) / 100)
            pct_cell.number_format = '0.0%'
            pct_cell.border = thin_border

        # --- Sheet 3: Daily Campaign Activity ---
        ws3 = wb.create_sheet('Daily Campaigns')
        daily_headers = ['Date', '4-Week Sends', '6-Week Sends', '8-Week Sends', 'Total Sends']
        for col_idx, h in enumerate(daily_headers, 1):
            cell = ws3.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        daily_widths = [14, 14, 14, 14, 12]
        for i, w in enumerate(daily_widths, 1):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for idx, day in enumerate(overview.get('dailyCampaigns', []), 1):
            row = idx + 1
            ws3.cell(row=row, column=1, value=day.get('date', '')).border = thin_border
            ws3.cell(row=row, column=2, value=day.get('fourWeekSends', 0)).border = thin_border
            ws3.cell(row=row, column=3, value=day.get('sixWeekSends', 0)).border = thin_border
            ws3.cell(row=row, column=4, value=day.get('eightWeekSends', 0)).border = thin_border
            ws3.cell(row=row, column=5, value=day.get('totalSends', 0)).border = thin_border

        s = start_date_str or ''
        e = end_date_str or ''
        filename = f'Exponential_Report_{s}_to_{e}.xlsx' if s and e else 'Exponential_Report.xlsx'

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class SchedulingOverviewView(APIView):
    """Scheduling Overview - Returns labor optimization metrics."""
    def get(self, request):
        from .services.scheduling_service import SchedulingService
        from .constants import DEFAULT_SCHEDULING_CONFIG
        user = request.user
        tenant = getattr(user, 'tenant', None)
        if tenant is None and user.role != 'super_admin':
            return Response({'error': 'No tenant assigned'}, status=400)

        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        config = AppConfig.get_config_value('scheduling_config', DEFAULT_SCHEDULING_CONFIG)
        service = SchedulingService(tenant=tenant, config=config)
        overview = service.get_overview_metrics(start_date=start_date_str, end_date=end_date_str)
        overview['config'] = config
        return Response(overview)


class SchedulingStoreRankingsView(APIView):
    """Stores ranked by % idle hours."""
    def get(self, request):
        from .services.scheduling_service import SchedulingService
        from .constants import DEFAULT_SCHEDULING_CONFIG
        user = request.user
        tenant = getattr(user, 'tenant', None)
        if tenant is None and user.role != 'super_admin':
            return Response({'error': 'No tenant assigned'}, status=400)
        config = AppConfig.get_config_value('scheduling_config', DEFAULT_SCHEDULING_CONFIG)
        service = SchedulingService(tenant=tenant, config=config)
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        return Response(service.get_store_rankings(tenant=tenant, start_date=start_date, end_date=end_date))


class SchedulingStoreDrilldownView(APIView):
    """Detailed time-bucket data for a specific store."""
    def get(self, request, store_id):
        from .services.scheduling_service import SchedulingService
        from .constants import DEFAULT_SCHEDULING_CONFIG
        user = request.user
        tenant = getattr(user, 'tenant', None)
        if tenant is None and user.role != 'super_admin':
            return Response({'error': 'No tenant assigned'}, status=400)
        config = AppConfig.get_config_value('scheduling_config', DEFAULT_SCHEDULING_CONFIG)
        service = SchedulingService(tenant=tenant, config=config)
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        result = service.get_store_drilldown(store_id, start_date=start_date, end_date=end_date)
        if result is None:
            return Response({'error': 'Store not found'}, status=404)
        return Response(result)


class SMSTemplateListView(APIView):
    """CRUD for SMS message templates."""

    def _get_tenant(self, request):
        user = request.user
        tenant = user.tenant
        if not tenant and user.role == 'super_admin':
            tenant = Tenant.objects.filter(is_active=True).first()
        return tenant

    def get(self, request):
        """List all templates, optionally filter by bucket or is_active."""
        tenant = self._get_tenant(request)
        if not tenant:
            return Response({'error': 'No tenant assigned'}, status=400)

        qs = SMSTemplate.objects.filter(tenant=tenant).order_by('bucket', '-is_active', 'name')
        bucket = request.query_params.get('bucket')
        if bucket:
            qs = qs.filter(bucket=bucket)
        active_only = request.query_params.get('active')
        if active_only == 'true':
            qs = qs.filter(is_active=True)

        serializer = SMSTemplateSerializer(qs, many=True)
        return Response({
            'templates': serializer.data,
            'available_placeholders': [
                {'key': '{first_name}', 'description': 'Guest first name'},
                {'key': '{guest_name}', 'description': 'Guest full name'},
                {'key': '{store_name}', 'description': 'Store/location name'},
                {'key': '{location_name}', 'description': 'Store/location name (alias)'},
                {'key': '{coupon_value}', 'description': 'Coupon dollar amount'},
                {'key': '{coupon_code}', 'description': 'Coupon code'},
                {'key': '{booking_link}', 'description': 'Booking URL'},
            ],
            'buckets': [
                {'value': '4wk', 'label': '4-Week (Low Churn Risk)'},
                {'value': '6wk', 'label': '6-Week (Medium Churn Risk)'},
                {'value': '8wk', 'label': '8-Week (High Churn Risk)'},
            ],
        })

    def post(self, request):
        """Create a new SMS template."""
        tenant = self._get_tenant(request)
        if not tenant:
            return Response({'error': 'No tenant assigned'}, status=400)

        serializer = SMSTemplateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({'error': 'Validation failed', 'details': serializer.errors}, status=400)

        # Check unique template_id per tenant
        template_id = serializer.validated_data.get('template_id')
        if SMSTemplate.objects.filter(tenant=tenant, template_id=template_id).exists():
            return Response({'error': f'Template with id "{template_id}" already exists'}, status=400)

        serializer.save(tenant=tenant)
        return Response({'status': 'success', 'template': serializer.data}, status=201)


class SMSTemplateDetailView(APIView):
    """Retrieve, update, or delete a single SMS template."""

    def _get_tenant(self, request):
        user = request.user
        tenant = user.tenant
        if not tenant and user.role == 'super_admin':
            tenant = Tenant.objects.filter(is_active=True).first()
        return tenant

    def _get_template(self, tenant, template_id):
        try:
            return SMSTemplate.objects.get(tenant=tenant, id=template_id)
        except SMSTemplate.DoesNotExist:
            return None

    def get(self, request, template_id):
        tenant = self._get_tenant(request)
        if not tenant:
            return Response({'error': 'No tenant assigned'}, status=400)
        template = self._get_template(tenant, template_id)
        if not template:
            return Response({'error': 'Template not found'}, status=404)
        serializer = SMSTemplateSerializer(template)
        return Response(serializer.data)

    def put(self, request, template_id):
        tenant = self._get_tenant(request)
        if not tenant:
            return Response({'error': 'No tenant assigned'}, status=400)
        template = self._get_template(tenant, template_id)
        if not template:
            return Response({'error': 'Template not found'}, status=404)

        serializer = SMSTemplateSerializer(template, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response({'error': 'Validation failed', 'details': serializer.errors}, status=400)
        serializer.save()
        return Response({'status': 'success', 'template': serializer.data})

    def delete(self, request, template_id):
        tenant = self._get_tenant(request)
        if not tenant:
            return Response({'error': 'No tenant assigned'}, status=400)
        template = self._get_template(tenant, template_id)
        if not template:
            return Response({'error': 'Template not found'}, status=404)
        template.delete()
        return Response({'status': 'deleted'}, status=204)

