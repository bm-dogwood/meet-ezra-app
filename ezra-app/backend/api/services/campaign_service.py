"""
Ezra Exponential - Campaign Service

Handles campaign CRUD, execution, audience estimation, guest import (file upload),
scheduling, multi-location targeting, and message log queries.
"""
import csv
import io
import logging
from datetime import date, timedelta, datetime
from decimal import Decimal

from django.db import transaction
from django.db.models import Count, Q, F, Sum
from django.db.models.functions import TruncDate
from django.utils import timezone

from api.models import (
    Store, Tenant, ExponentialCustomer, ExponentialVisit,
    ExponentialCampaign, ExponentialSMSLog, ExponentialUptake,
    ExponentialSegment, SMSTemplate, AppConfig, SegmentConfig,
)
from api.constants import DEFAULT_EXPONENTIAL_CONFIG

logger = logging.getLogger(__name__)

SEGMENT_MAP = {
    '4wk': '4-week',
    '6wk': '6-week',
    '8wk': '8-week',
    '4-week': '4wk',
    '6-week': '6wk',
    '8-week': '8wk',
}

# Max file upload size: 10 MB
MAX_UPLOAD_SIZE = 10 * 1024 * 1024


class CampaignService:
    """Service for campaign management operations."""

    def __init__(self, tenant=None, config=None):
        self.tenant = tenant
        self.config = config or DEFAULT_EXPONENTIAL_CONFIG

    # ------------------------------------------------------------------
    # Campaign List
    # ------------------------------------------------------------------
    def list_campaigns(self, status=None, segment=None, search=None, page=0, limit=20, service_filter=None):
        """List campaigns with optional filters and pagination. OPTIMIZED VERSION."""
        qs = ExponentialCampaign.objects.all().order_by('-created_at')
        if self.tenant:
            qs = qs.filter(tenant=self.tenant)
        if status and status != 'all':
            status_map = {
                'sending': 'active',
                'sent': 'completed',
                'partially_sent': 'completed',
                'scheduled': 'scheduled',
                'active': 'active',
                'completed': 'completed',
                'paused': 'paused',
            }
            qs = qs.filter(status=status_map.get(status, status))
        if segment and segment != 'all':
            bucket = SEGMENT_MAP.get(segment, segment)
            qs = qs.filter(target_bucket=bucket)
        if service_filter and service_filter != 'all':
            qs = qs.filter(service_filter__icontains=service_filter)
        if search:
            qs = qs.filter(name__icontains=search)

        total = qs.count()
        offset = page * limit
        campaigns_page = list(qs[offset:offset + limit])
        
        # Bulk query: Get SMS stats for all campaigns on this page
        campaign_ids = [c.id for c in campaigns_page]
        sms_stats = self._get_bulk_sms_stats(campaign_ids)
        
        # Serialize with pre-fetched stats
        campaigns = [self._serialize_campaign(c, sms_stats.get(c.id, {})) for c in campaigns_page]
        return {'campaigns': campaigns, 'total': total, 'page': page, 'limit': limit}

    def get_campaign_detail(self, campaign_id):
        """Get a single campaign with full details."""
        try:
            qs = ExponentialCampaign.objects.all()
            if self.tenant:
                qs = qs.filter(tenant=self.tenant)
            c = qs.get(id=campaign_id)
            return self._serialize_campaign(c)
        except ExponentialCampaign.DoesNotExist:
            return None

    def get_campaign_messages(self, campaign_id, status=None, search=None, page=0, limit=50):
        """Get paginated message log for a campaign."""
        qs = ExponentialSMSLog.objects.filter(
            campaign_id=campaign_id
        ).select_related('customer', 'customer__store').order_by('-sent_at')
        if self.tenant:
            qs = qs.filter(campaign__tenant=self.tenant)

        if status and status != 'all':
            qs = qs.filter(status=status)
        if search:
            qs = qs.filter(
                Q(customer__guest_name__icontains=search) |
                Q(customer__phone__icontains=search) |
                Q(customer__store__name__icontains=search)
            )

        total = qs.count()
        offset = page * limit
        logs = qs[offset:offset + limit]

        messages = []
        for log in logs:
            cust = log.customer
            store = cust.store if cust else None
            seg_name = SEGMENT_MAP.get(log.segment_at_send, log.segment_at_send)

            messages.append({
                'id': str(log.id),
                'campaignId': str(log.campaign_id),
                'guestId': cust.guest_code if cust else '',
                'guestPhone': cust.phone or '' if cust else '',
                'locationId': str(store.id) if store else '',
                'locationName': store.name if store else '',
                'segment': seg_name,
                'couponValue': f'${log.coupon_value:.0f}' if log.coupon_value and log.coupon_value == int(log.coupon_value) else (f'${log.coupon_value:.2f}' if log.coupon_value else ''),
                'couponCode': log.campaign.coupon_code if log.campaign else '',
                'messageContent': log.message_body or '',
                'scheduledAt': None,
                'sentAt': log.sent_at.isoformat() if log.sent_at else None,
                'deliveredAt': log.delivered_at.isoformat() if log.delivered_at else None,
                'status': self._map_message_status(log.status),
                'twilioMessageSid': log.twilio_message_sid or None,
                'failedReason': log.error_message or None,
            })

        return {
            'messages': messages,
            'total': total,
            'page': page,
            'limit': limit,
        }

    def get_campaign_stats(self):
        """Get summary stats across all campaigns."""
        qs = ExponentialCampaign.objects.all()
        if self.tenant:
            qs = qs.filter(tenant=self.tenant)

        total_campaigns = qs.count()
        active_campaigns = qs.filter(status='active').count()
        scheduled_campaigns = qs.filter(status='scheduled').count()

        sms_qs = ExponentialSMSLog.objects.all()
        if self.tenant:
            sms_qs = sms_qs.filter(campaign__tenant=self.tenant)

        total_sent = sms_qs.count()
        total_delivered = sms_qs.filter(status='delivered').count()
        avg_delivery = round(total_delivered / total_sent * 100, 1) if total_sent > 0 else 0

        return {
            'totalCampaigns': total_campaigns,
            'activeCampaigns': active_campaigns,
            'scheduledCampaigns': scheduled_campaigns,
            'totalMessagesSent': total_sent,
            'totalDelivered': total_delivered,
            'avgDeliveryRate': avg_delivery,
        }

    def _get_bulk_sms_stats(self, campaign_ids):
        """Get SMS stats for multiple campaigns in bulk. OPTIMIZATION."""
        if not campaign_ids:
            return {}
        
        from django.db.models import Count, Q, Case, When, IntegerField
        
        # Single query to get all stats for all campaigns
        stats = ExponentialSMSLog.objects.filter(
            campaign_id__in=campaign_ids
        ).values('campaign_id').annotate(
            total=Count('id'),
            delivered=Count('id', filter=Q(status='delivered')),
            failed=Count('id', filter=Q(status__in=['failed', 'undelivered'])),
            sent=Count('id', filter=Q(status='sent')),
            queued=Count('id', filter=Q(status='queued')),
        )
        
        # Convert to dict for easy lookup
        stats_map = {}
        for stat in stats:
            campaign_id = stat['campaign_id']
            stats_map[campaign_id] = {
                'total': stat['total'],
                'delivered': stat['delivered'],
                'failed': stat['failed'],
                'sent': stat['sent'],
                'queued': stat['queued'],
            }
        
        return stats_map

    # ------------------------------------------------------------------
    # Campaign Create / Update / Delete
    # ------------------------------------------------------------------
    @staticmethod
    def _to_utc(dt_str, tz_name):
        """Convert a naive datetime string from the given timezone to UTC."""
        import zoneinfo
        if not dt_str:
            return dt_str
        # If already timezone-aware (has +/- offset or Z), return as-is
        if isinstance(dt_str, str) and ('+' in dt_str[10:] or dt_str.endswith('Z')):
            return dt_str
        try:
            from datetime import datetime as _dt
            naive = _dt.fromisoformat(str(dt_str))
            local_tz = zoneinfo.ZoneInfo(tz_name)
            local_dt = naive.replace(tzinfo=local_tz)
            utc_dt = local_dt.astimezone(zoneinfo.ZoneInfo('UTC'))
            return utc_dt
        except Exception:
            return dt_str

    def create_campaign(self, data, tenant):
        """Create a new campaign from request data."""
        # Map segment name to bucket — keep as-is for dynamic segments
        target_bucket = data.get('target_bucket', 'all')
        # No mapping needed — store the segment name directly

        coupon_key = f'coupon_{target_bucket}'
        default_coupon = self.config.get(coupon_key, 10)

        # Enforce unique campaign name per tenant
        name = data.get('name', f'Campaign - {target_bucket.upper()}')
        if ExponentialCampaign.objects.filter(tenant=tenant, name=name).exists():
            return {'error': f'A campaign with the name "{name}" already exists. Please choose a different name.'}

        # Validate coupon code length
        coupon_code = (data.get('coupon_code', '') or '')[:15].upper()

        store = None
        if data.get('store_id'):
            store = Store.objects.filter(id=data['store_id']).first()

        # Determine scope
        scope = data.get('scope', self.config.get('default_campaign_scope', 'all'))
        location_ids = data.get('location_ids', [])
        guest_ids = data.get('guest_ids', [])
        if guest_ids:
            scope = 'guests'
        elif location_ids and len(location_ids) > 1:
            scope = 'multi'
        elif location_ids and len(location_ids) == 1:
            scope = 'single'
            store = Store.objects.filter(id=location_ids[0]).first()

        # Determine initial status — always 'scheduled' unless immediate
        schedule_type = data.get('schedule_type', 'immediate')
        scheduled_at = data.get('scheduled_at')
        campaign_tz = data.get('campaign_timezone', 'America/New_York')
        initial_status = 'scheduled'
        if schedule_type == 'scheduled' and scheduled_at:
            # Convert scheduled_at from user's timezone to UTC
            scheduled_at = self._to_utc(scheduled_at, campaign_tz)

        # For recurring campaigns, compute scheduled_at from start_date + time
        recurring_start_date = data.get('recurring_start_date')
        recurring_time = data.get('recurring_time', '10:00')
        if data.get('is_recurring') and recurring_start_date and recurring_time:
            # Build a datetime string from start_date + time, then convert to UTC
            start_dt_str = f"{recurring_start_date}T{recurring_time}"
            scheduled_at = self._to_utc(start_dt_str, campaign_tz)

        # Extract service_filter from template_variables if present
        tv = data.get('template_variables', {})
        service_filter = tv.get('service_filter', '') if isinstance(tv, dict) else ''

        # Extract date filters
        visit_date_from = data.get('visit_date_from') or (tv.get('date_from') if isinstance(tv, dict) else None)
        visit_date_to = data.get('visit_date_to') or (tv.get('date_to') if isinstance(tv, dict) else None)

        campaign = ExponentialCampaign.objects.create(
            tenant=tenant,
            store=store,
            name=name,
            target_bucket=target_bucket,
            message_template=data.get('message_template', ''),
            coupon_value=Decimal(str(data.get('coupon_value', 0) or 0)),
            coupon_code=coupon_code,
            booking_link=data.get('booking_link', ''),
            template_variables=data.get('template_variables', {}),
            scope=scope,
            location_ids=[int(x) for x in location_ids] if location_ids else [],
            guest_ids=guest_ids or [],
            status=initial_status,
            scheduled_at=scheduled_at,
            is_recurring=data.get('is_recurring', False),
            recurring_frequency=data.get('recurring_frequency', ''),
            recurring_start_date=data.get('recurring_start_date'),
            recurring_end_date=data.get('recurring_end_date'),
            recurring_time=data.get('recurring_time', ''),
            recurring_day_of_week=data.get('recurring_day_of_week'),
            campaign_timezone=data.get('campaign_timezone', 'America/New_York'),
            service_filter=service_filter,
            visit_date_from=visit_date_from or None,
            visit_date_to=visit_date_to or None,
        )
        return campaign

    def update_campaign(self, campaign_id, data):
        """Update a draft or scheduled campaign."""
        from datetime import date as _date

        try:
            qs = ExponentialCampaign.objects.all()
            if self.tenant:
                qs = qs.filter(tenant=self.tenant)
            c = qs.get(id=campaign_id)
        except ExponentialCampaign.DoesNotExist:
            return {'error': 'Campaign not found'}

        if c.status not in ('scheduled',):
            return {'error': 'Only scheduled campaigns can be edited'}

        # Derive is_recurring from schedule_type if present
        if 'schedule_type' in data:
            data['is_recurring'] = data['schedule_type'] == 'recurring'

        updatable = [
            'name', 'message_template', 'coupon_code', 'booking_link',
            'template_variables',
            'scheduled_at', 'is_recurring', 'recurring_frequency',
            'recurring_start_date', 'recurring_end_date', 'recurring_time',
            'recurring_day_of_week', 'campaign_timezone',
        ]
        update_fields = []

        # Enforce unique name per tenant
        if 'name' in data and data['name'] != c.name:
            if ExponentialCampaign.objects.filter(tenant=c.tenant, name=data['name']).exclude(id=c.id).exists():
                return {'error': f'A campaign with the name "{data["name"]}" already exists.'}

        for field in updatable:
            if field in data:
                val = data[field]
                # Convert date string fields to date objects
                if field in ('recurring_start_date', 'recurring_end_date') and isinstance(val, str) and val:
                    try:
                        val = _date.fromisoformat(val)
                    except (ValueError, TypeError):
                        pass
                # Convert scheduled_at from user's timezone to UTC
                if field == 'scheduled_at' and val:
                    tz_name = data.get('campaign_timezone') or c.campaign_timezone or 'America/New_York'
                    val = self._to_utc(val, tz_name)
                setattr(c, field, val)
                update_fields.append(field)

        if 'coupon_value' in data:
            c.coupon_value = Decimal(str(data['coupon_value']))
            update_fields.append('coupon_value')

        if 'target_bucket' in data:
            c.target_bucket = data['target_bucket']
            update_fields.append('target_bucket')

        if 'location_ids' in data:
            c.location_ids = [int(x) for x in data['location_ids']]
            if len(c.location_ids) > 1:
                c.scope = 'multi'
            elif len(c.location_ids) == 1:
                c.scope = 'single'
                c.store = Store.objects.filter(id=c.location_ids[0]).first()
            update_fields.extend(['location_ids', 'scope', 'store_id'])

        if 'guest_ids' in data:
            c.guest_ids = data['guest_ids']
            if c.guest_ids:
                c.scope = 'guests'
            update_fields.extend(['guest_ids', 'scope'])

        # For recurring campaigns, recompute scheduled_at from start_date + time
        if c.is_recurring and c.recurring_start_date and c.recurring_time:
            campaign_tz = c.campaign_timezone or 'America/New_York'
            start_date_str = str(c.recurring_start_date)
            start_dt_str = f"{start_date_str}T{c.recurring_time}"
            c.scheduled_at = self._to_utc(start_dt_str, campaign_tz)
            if 'scheduled_at' not in update_fields:
                update_fields.append('scheduled_at')

        # Update status based on scheduling
        if c.is_recurring or ('scheduled_at' in data and data['scheduled_at']):
            c.status = 'scheduled'
            update_fields.append('status')
        elif 'schedule_type' in data and data['schedule_type'] == 'immediate':
            c.status = 'scheduled'
            c.scheduled_at = None
            update_fields.extend(['status', 'scheduled_at'])

        if update_fields:
            c.save(update_fields=list(set(update_fields)))

        return self._serialize_campaign(c)

    def delete_campaign(self, campaign_id):
        """Delete a campaign (any status)."""
        try:
            qs = ExponentialCampaign.objects.all()
            if self.tenant:
                qs = qs.filter(tenant=self.tenant)
            c = qs.get(id=campaign_id)
        except ExponentialCampaign.DoesNotExist:
            return {'error': 'Campaign not found'}

        c.delete()
        return {'status': 'deleted'}

    # ------------------------------------------------------------------
    # Campaign Execution
    # ------------------------------------------------------------------
    def execute_campaign(self, campaign_id):
        """
        Execute a campaign: find eligible customers, send SMS, log results.
        Returns dict with execution summary.
        """
        try:
            qs = ExponentialCampaign.objects.all()
            if self.tenant:
                qs = qs.filter(tenant=self.tenant)
            campaign = qs.get(id=campaign_id)
        except ExponentialCampaign.DoesNotExist:
            return {'error': 'Campaign not found'}

        if campaign.status not in ('scheduled', 'active'):
            return {'error': f'Campaign cannot be executed in {campaign.status} status'}

        campaign.status = 'active'
        if not campaign.started_at:
            campaign.started_at = timezone.now()
        campaign.save(update_fields=['status', 'started_at'])

        customers = self._get_eligible_customers(campaign)

        sent = 0
        failed = 0
        skipped_dupe_phone = 0
        phones_sent = set()

        from api.services.twilio_sms_service import send_campaign_sms

        for customer in customers:
            # Dedupe: only send once per phone number per campaign
            phone = (customer.phone or '').strip()
            if phone and phone in phones_sent:
                skipped_dupe_phone += 1
                continue
            if phone:
                phones_sent.add(phone)

            # Resolve store name: customer's store → campaign store → first campaign location
            if customer.store:
                store_name = customer.store.name
            elif campaign.store:
                store_name = campaign.store.name
            elif campaign.location_ids:
                _fallback_store = Store.objects.filter(id=campaign.location_ids[0]).first()
                store_name = _fallback_store.name if _fallback_store else ''
            else:
                store_name = ''
            template_body = campaign.message_template or (
                'Hi {first_name}! Visit {location_name} and save ${coupon_value}. Reply STOP to opt out.'
            )

            result = send_campaign_sms(
                customer=customer,
                campaign=campaign,
                template_body=template_body,
                coupon_value=float(campaign.coupon_value),
                store_name=store_name,
            )

            # Log the SMS - initial status from Twilio (queued/sent), not delivered
            sms_status = 'queued'
            if result['success']:
                sms_status = result.get('status', 'sent')
                # Normalize Twilio statuses
                if sms_status in ('accepted', 'queued'):
                    sms_status = 'queued'
                elif sms_status in ('sending', 'sent'):
                    sms_status = 'sent'
            else:
                sms_status = 'failed'

            ExponentialSMSLog.objects.create(
                campaign=campaign,
                customer=customer,
                segment_at_send=campaign.target_bucket,
                coupon_value=campaign.coupon_value,
                message_body=result.get('message_body', ''),
                status=sms_status,
                twilio_message_sid=result.get('message_sid', ''),
                error_message=result.get('error') or '',
            )

            if result['success']:
                sent += 1
            else:
                failed += 1

        # Update campaign counters
        campaign.refresh_from_db()
        campaign.messages_sent = (campaign.messages_sent or 0) + sent + failed
        campaign.messages_failed = (campaign.messages_failed or 0) + failed
        campaign.status = 'completed'
        campaign.completed_at = timezone.now()
        if campaign.is_recurring:
            campaign.last_recurring_run = timezone.now()
        campaign.save(update_fields=[
            'messages_sent', 'messages_failed',
            'status', 'completed_at', 'last_recurring_run',
        ])

        return {
            'campaign_id': campaign.id,
            'customers_targeted': len(customers),
            'sent': sent,
            'failed': failed,
            'skipped_duplicate_phone': skipped_dupe_phone,
        }

    def _get_eligible_customers(self, campaign):
        """Get customers eligible for this campaign based on scope and bucket. Optimised."""
        qs = ExponentialCustomer.objects.filter(sms_opt_in=True)
        if self.tenant:
            # Include CRM guests (tenant=NULL) alongside tenant-owned ones
            qs = qs.filter(Q(tenant=self.tenant) | Q(tenant__isnull=True))

        # --- Scope filtering ---
        if campaign.scope == 'guests' and campaign.guest_ids:
            qs = qs.filter(guest_code__in=campaign.guest_ids)
        elif campaign.scope == 'multi' and campaign.location_ids:
            qs = qs.filter(store_id__in=campaign.location_ids)
        elif campaign.scope == 'single' and campaign.store:
            qs = qs.filter(store=campaign.store)

        # Must have phone — unless test_phone is configured
        from api.services.twilio_sms_service import get_twilio_config
        twilio_cfg = get_twilio_config()
        test_phone = (twilio_cfg.get('test_phone') or '').strip()
        if not test_phone:
            qs = qs.exclude(Q(phone__isnull=True) | Q(phone=''))

        # Service type filter (supports comma-separated list as OR)
        svc_filter = getattr(campaign, 'service_filter', '') or ''
        if svc_filter:
            services = [s.strip() for s in svc_filter.split(',') if s.strip()]
            if len(services) > 1:
                qs = qs.filter(last_service__in=services)
            elif len(services) == 1:
                qs = qs.filter(last_service=services[0])

        # Last visit date range filter
        visit_from = getattr(campaign, 'visit_date_from', None)
        visit_to = getattr(campaign, 'visit_date_to', None)
        if visit_from:
            qs = qs.filter(last_visit_date__gte=visit_from)
        if visit_to:
            qs = qs.filter(last_visit_date__lte=visit_to)

        # Cooldown
        cooldown_days = self.config.get('cooldown_days', 0)
        if cooldown_days > 0:
            cooldown_cutoff = timezone.now() - timedelta(days=cooldown_days)
            recently_messaged = ExponentialSMSLog.objects.filter(
                customer__in=qs,
                sent_at__gte=cooldown_cutoff,
            ).values_list('customer_id', flat=True)
            qs = qs.exclude(id__in=recently_messaged)

        # Skip bucket filtering for guest-targeted campaigns
        if campaign.scope == 'guests' and campaign.guest_ids:
            return list(qs.select_related('store'))

        ref = date.today()
        cfg = self.config
        bucket = campaign.target_bucket
        thirty_days_ago = ref - timedelta(days=30)

        # Filter by bucket using DB-level queries
        qs = qs.filter(last_visit_date__isnull=False)

        # Annotate visits in last 30 days
        qs = qs.annotate(
            visits_30d=Count(
                'visits',
                filter=Q(
                    visits__visit_date__gte=thirty_days_ago,
                    visits__visit_date__lte=ref,
                )
            )
        )

        # Try dynamic segment config first
        segment_config = None
        if bucket and bucket != 'all':
            try:
                segment_config = SegmentConfig.objects.filter(
                    Q(tenant=self.tenant) | Q(tenant__isnull=True),
                    Q(name=bucket) | Q(slug=bucket),
                    is_active=True
                ).first()
            except Exception:
                pass

        if segment_config:
            # Use dynamic segment config for filtering
            if segment_config.max_days:
                cutoff_min = ref - timedelta(days=segment_config.max_days)
                cutoff_max = ref - timedelta(days=segment_config.min_days)
                qs = qs.filter(
                    last_visit_date__gte=cutoff_min,
                    last_visit_date__lte=cutoff_max,
                )
            else:
                cutoff = ref - timedelta(days=segment_config.min_days)
                qs = qs.filter(last_visit_date__lte=cutoff)
        elif bucket == '4wk':
            qs = qs.filter(visits_30d__gte=cfg.get('bucket_4wk_min_visits', 2))
        elif bucket == '6wk':
            min_days = cfg.get('bucket_6wk_min_days', 31)
            max_days = cfg.get('bucket_6wk_max_days', 42)
            cutoff_max = ref - timedelta(days=min_days)
            cutoff_min = ref - timedelta(days=max_days)
            qs = qs.filter(
                last_visit_date__gte=cutoff_min,
                last_visit_date__lte=cutoff_max,
            )
        elif bucket == '8wk':
            min_days = cfg.get('bucket_8wk_min_days', 43)
            cutoff = ref - timedelta(days=min_days)
            qs = qs.filter(last_visit_date__lte=cutoff)
        # If bucket is 'all' or unrecognized, no segment filter — send to all

        return list(qs.select_related('store'))

    # ------------------------------------------------------------------
    # Scheduled Campaign Execution
    # ------------------------------------------------------------------
    def execute_due_campaigns(self):
        """
        Find and execute campaigns that are scheduled and due.
        Called by management command or celery beat.
        Returns list of execution results.
        """
        now = timezone.now()
        due = ExponentialCampaign.objects.filter(
            status='scheduled',
            scheduled_at__lte=now,
        )
        if self.tenant:
            due = due.filter(tenant=self.tenant)

        results = []
        for campaign in due:
            logger.info(f"Executing scheduled campaign {campaign.id}: {campaign.name}")
            result = self.execute_campaign(campaign.id)
            results.append(result)

        # Handle recurring campaigns that need a new run
        recurring = ExponentialCampaign.objects.filter(
            is_recurring=True,
            status='completed',
        ).exclude(recurring_frequency='')
        if self.tenant:
            recurring = recurring.filter(tenant=self.tenant)

        for campaign in recurring:
            if campaign.recurring_end_date and date.today() > campaign.recurring_end_date:
                continue  # Past end date

            next_run = self._get_next_recurring_run(campaign)
            if next_run and next_run <= now:
                # Clone campaign for new run
                new_campaign = self._clone_campaign_for_recurring(campaign)
                logger.info(f"Executing recurring campaign clone {new_campaign.id} from {campaign.id}")
                result = self.execute_campaign(new_campaign.id)
                results.append(result)

        return results

    def _get_next_recurring_run(self, campaign):
        """Calculate next run time for a recurring campaign."""
        last_run = campaign.last_recurring_run or campaign.completed_at
        if not last_run:
            return None

        freq = campaign.recurring_frequency
        if freq == 'daily':
            return last_run + timedelta(days=1)
        elif freq == 'weekly':
            return last_run + timedelta(weeks=1)
        elif freq == 'biweekly':
            return last_run + timedelta(weeks=2)
        elif freq == 'monthly':
            return last_run + timedelta(days=30)
        return None

    def _clone_campaign_for_recurring(self, original):
        """Create a new campaign instance from a recurring template."""
        new = ExponentialCampaign.objects.create(
            tenant=original.tenant,
            store=original.store,
            name=f"{original.name} (Recurring {date.today().isoformat()})",
            target_bucket=original.target_bucket,
            message_template=original.message_template,
            coupon_value=original.coupon_value,
            coupon_code=original.coupon_code,
            booking_link=original.booking_link,
            scope=original.scope,
            location_ids=original.location_ids,
            guest_ids=original.guest_ids,
            status='scheduled',
            is_recurring=True,
            recurring_frequency=original.recurring_frequency,
            recurring_end_date=original.recurring_end_date,
            service_filter=original.service_filter,
        )
        # Mark original as having run
        original.last_recurring_run = timezone.now()
        original.save(update_fields=['last_recurring_run'])
        return new

    # ------------------------------------------------------------------
    # Audience Estimation
    # ------------------------------------------------------------------
    def estimate_audience(self, segment=None, location_ids=None, sms_status=None, last_service=None, date_from=None, date_to=None):
        """Estimate audience size for campaign creation wizard. Matches _get_eligible_customers logic."""
        ref = date.today()
        cfg = self.config
        thirty_days_ago = ref - timedelta(days=30)

        qs = ExponentialCustomer.objects.filter(sms_opt_in=True)
        if self.tenant:
            # Include CRM guests (tenant=NULL) — same as _get_eligible_customers
            qs = qs.filter(Q(tenant=self.tenant) | Q(tenant__isnull=True))

        # Phone required — unless test_phone is configured (same as execution)
        from api.services.twilio_sms_service import get_twilio_config
        twilio_cfg = get_twilio_config()
        test_phone = (twilio_cfg.get('test_phone') or '').strip()
        if not test_phone:
            qs = qs.exclude(Q(phone__isnull=True) | Q(phone=''))

        # Service type filter (list = OR)
        if last_service:
            if isinstance(last_service, list):
                qs = qs.filter(last_service__in=last_service)
            else:
                qs = qs.filter(last_service__icontains=last_service)

        # Last visit date range filter
        if date_from:
            qs = qs.filter(last_visit_date__gte=date_from)
        if date_to:
            qs = qs.filter(last_visit_date__lte=date_to)

        # SMS status filter
        if sms_status == 'opted_in':
            qs = qs.filter(sms_opt_in=True)
        elif sms_status == 'opted_out':
            qs = qs.filter(sms_opt_in=False)
        elif sms_status == 'all':
            pass
            
        if location_ids:
            qs = qs.filter(store_id__in=location_ids)

        # Must have visit date for segment filtering
        qs = qs.filter(last_visit_date__isnull=False)

        # Cooldown — same as execution
        cooldown_days = cfg.get('cooldown_days', 0)
        if cooldown_days > 0:
            cooldown_cutoff = timezone.now() - timedelta(days=cooldown_days)
            recently_messaged = ExponentialSMSLog.objects.filter(
                customer__in=qs,
                sent_at__gte=cooldown_cutoff,
            ).values_list('customer_id', flat=True)
            qs = qs.exclude(id__in=recently_messaged)

        # Annotate visits in last 30 days
        qs = qs.annotate(
            visits_30d=Count(
                'visits',
                filter=Q(
                    visits__visit_date__gte=thirty_days_ago,
                    visits__visit_date__lte=ref,
                )
            )
        )

        # Get segment configs for dynamic filtering
        segment_configs = list(SegmentConfig.objects.filter(
            Q(tenant=self.tenant) | Q(tenant__isnull=True),
            is_active=True
        ).order_by('sort_order'))

        result = []
        
        if segment and segment != 'all':
            # Filter for specific segment
            segment_config = next((s for s in segment_configs if s.name == segment or s.slug == segment), None)
            if segment_config:
                filtered_qs = qs.filter(last_visit_date__isnull=False)
                if segment_config.max_days:
                    cutoff_min = ref - timedelta(days=segment_config.max_days)
                    cutoff_max = ref - timedelta(days=segment_config.min_days)
                    filtered_qs = filtered_qs.filter(
                        last_visit_date__gte=cutoff_min,
                        last_visit_date__lte=cutoff_max,
                    )
                else:
                    cutoff = ref - timedelta(days=segment_config.min_days)
                    filtered_qs = filtered_qs.filter(last_visit_date__lte=cutoff)
                
                count = filtered_qs.count()
                result.append({'segment': segment_config.name, 'count': count})
            else:
                # Fallback to old hardcoded logic
                min_visits_4wk = cfg.get('bucket_4wk_min_visits', 2)
                min_days_6wk = cfg.get('bucket_6wk_min_days', 31)
                max_days_6wk = cfg.get('bucket_6wk_max_days', 42)
                min_days_8wk = cfg.get('bucket_8wk_min_days', 43)

                if segment == '4-week':
                    count = qs.filter(visits_30d__gte=min_visits_4wk).count()
                    result.append({'segment': '4-week', 'count': count})
                elif segment == '6-week':
                    cutoff_6wk_max = ref - timedelta(days=min_days_6wk)
                    cutoff_6wk_min = ref - timedelta(days=max_days_6wk)
                    count = qs.filter(
                        last_visit_date__gte=cutoff_6wk_min,
                        last_visit_date__lte=cutoff_6wk_max,
                    ).exclude(visits_30d__gte=min_visits_4wk).count()
                    result.append({'segment': '6-week', 'count': count})
                elif segment == '8-week':
                    cutoff_8wk = ref - timedelta(days=min_days_8wk)
                    count = qs.filter(last_visit_date__lte=cutoff_8wk).exclude(visits_30d__gte=min_visits_4wk).count()
                    result.append({'segment': '8-week', 'count': count})
        else:
            # Return counts for all segments
            
            for segment_config in segment_configs:
                filtered_qs = qs.filter(last_visit_date__isnull=False)
                if segment_config.max_days:
                    cutoff_min = ref - timedelta(days=segment_config.max_days)
                    cutoff_max = ref - timedelta(days=segment_config.min_days)
                    filtered_qs = filtered_qs.filter(
                        last_visit_date__gte=cutoff_min,
                        last_visit_date__lte=cutoff_max,
                    )
                else:
                    cutoff = ref - timedelta(days=segment_config.min_days)
                    filtered_qs = filtered_qs.filter(last_visit_date__lte=cutoff)
                
                count = filtered_qs.count()
                result.append({'segment': segment_config.name, 'count': count})
            
            # Add 'all' segment - use actual total count from base queryset
            # This ensures we count ALL guests with phones, not just those in segments
            total_count = qs.count()
            result.append({'segment': 'all', 'count': total_count})

        return result

    # ------------------------------------------------------------------
    # Guest List
    # ------------------------------------------------------------------
    def get_guest_list(self, store_id=None, location_ids=None, bucket=None, search=None, page=0, limit=20, source=None, sms_status=None, last_service=None, guest_type=None, sort=None, date_from=None, date_to=None):
        """Paginated guest list with segment info. Optimised with annotations."""
        from django.db.models import Subquery, OuterRef, Value, CharField, IntegerField, DateTimeField
        from django.db.models.functions import Coalesce

        ref = date.today()
        cfg = self.config
        thirty_days_ago = ref - timedelta(days=30)

        # Base queryset with annotations to avoid N+1
        qs = ExponentialCustomer.objects.all().select_related('store')
        if self.tenant:
            qs = qs.filter(Q(tenant=self.tenant) | Q(tenant__isnull=True))
        if store_id:
            qs = qs.filter(store_id=store_id)
        elif location_ids:
            qs = qs.filter(store_id__in=location_ids)

        # Guest type filter (imported vs normal)
        if guest_type == 'imported':
            qs = qs.filter(guest_code__startswith='IMP-')
        elif guest_type == 'normal':
            qs = qs.exclude(guest_code__startswith='IMP-')
        # Legacy source param
        if source == 'imported':
            qs = qs.filter(guest_code__startswith='IMP-')

        if search:
            qs = qs.filter(
                Q(guest_name__icontains=search) |
                Q(guest_code__icontains=search) |
                Q(phone__icontains=search)
            )
        
        # SMS status filter
        if sms_status == 'opted_in':
            qs = qs.filter(sms_opt_in=True)
        elif sms_status == 'opted_out':
            qs = qs.filter(sms_opt_in=False)

        # Last service filter (list = OR)
        if last_service:
            if isinstance(last_service, list):
                qs = qs.filter(last_service__in=last_service)
            else:
                qs = qs.filter(last_service__icontains=last_service)

        # Date range filter on last_visit_date
        if date_from:
            qs = qs.filter(last_visit_date__gte=date_from)
        if date_to:
            qs = qs.filter(last_visit_date__lte=date_to)

        # Annotate visits in last 30 days (single query instead of per-row)
        qs = qs.annotate(
            visits_30d=Count(
                'visits',
                filter=Q(
                    visits__visit_date__gte=thirty_days_ago,
                    visits__visit_date__lte=ref,
                )
            )
        )

        # Annotate last SMS info via subquery
        latest_sms = ExponentialSMSLog.objects.filter(
            customer=OuterRef('pk')
        ).order_by('-sent_at')
        qs = qs.annotate(
            last_sms_date=Subquery(latest_sms.values('sent_at')[:1], output_field=DateTimeField()),
            last_sms_status=Subquery(latest_sms.values('status')[:1], output_field=CharField()),
        )

        # Bucket filtering using annotations (no Python loop)
        # Skip bucket filtering for CRM guests — they don't have visit data
        if bucket and source != 'imported':
            # Try to find matching segment config first
            try:
                segment_config = SegmentConfig.objects.filter(
                    Q(tenant=self.tenant) | Q(tenant__isnull=True),
                    Q(name=bucket) | Q(slug=bucket),
                    is_active=True
                ).first()
                
                if segment_config:
                    # Use segment config min/max days
                    if segment_config.max_days:
                        cutoff_min = ref - timedelta(days=segment_config.max_days)
                        cutoff_max = ref - timedelta(days=segment_config.min_days)
                        qs = qs.filter(
                            last_visit_date__isnull=False,
                            last_visit_date__gte=cutoff_min,
                            last_visit_date__lte=cutoff_max,
                        )
                    else:
                        # No max_days means "X+ weeks" (e.g., 8+ weeks)
                        cutoff = ref - timedelta(days=segment_config.min_days)
                        qs = qs.filter(
                            last_visit_date__isnull=False,
                            last_visit_date__lte=cutoff,
                        )
                else:
                    # Fallback to old hardcoded logic for backward compatibility
                    bk = SEGMENT_MAP.get(bucket, bucket)
                    if bk == '4wk':
                        qs = qs.filter(
                            last_visit_date__isnull=False,
                            visits_30d__gte=cfg.get('bucket_4wk_min_visits', 2),
                        )
                    elif bk == '6wk':
                        min_days = cfg.get('bucket_6wk_min_days', 31)
                        max_days = cfg.get('bucket_6wk_max_days', 42)
                        cutoff_max = ref - timedelta(days=min_days)
                        cutoff_min = ref - timedelta(days=max_days)
                        qs = qs.filter(
                            last_visit_date__isnull=False,
                            last_visit_date__gte=cutoff_min,
                            last_visit_date__lte=cutoff_max,
                        )
                    elif bk == '8wk':
                        min_days = cfg.get('bucket_8wk_min_days', 43)
                        cutoff = ref - timedelta(days=min_days)
                        qs = qs.filter(
                            last_visit_date__isnull=False,
                            last_visit_date__lte=cutoff,
                        )
            except Exception as e:
                logger.warning(f"Error filtering by bucket {bucket}: {e}")
                # Continue without bucket filtering

        total = qs.count()
        offset = page * limit
        # Sort order
        order_field = '-last_visit_date'
        if sort == 'last_visit_asc':
            order_field = 'last_visit_date'
        elif sort == 'last_visit_desc':
            order_field = '-last_visit_date'
        elif sort == 'name_asc':
            order_field = 'guest_name'
        elif sort == 'name_desc':
            order_field = '-guest_name'
        guests = qs.order_by(order_field)[offset:offset + limit]

        # Pre-build segment config list and service type lookup
        segment_configs_list = []
        try:
            segment_configs_list = list(SegmentConfig.objects.filter(
                Q(tenant=self.tenant) | Q(tenant__isnull=True),
                is_active=True
            ).order_by('min_days'))
        except Exception:
            pass

        # last_service is now a field on ExponentialCustomer — no raw report scanning needed

        result = []
        for g in guests:
            days_since = (ref - g.last_visit_date).days if g.last_visit_date else 999
            v30 = g.visits_30d

            # Determine segment using dynamic configs
            if g.guest_code.startswith('IMP-'):
                seg = 'imported'
            else:
                seg = 'active'
                for sc in segment_configs_list:
                    if sc.max_days:
                        if sc.min_days <= days_since <= sc.max_days:
                            seg = sc.name
                            break
                    else:
                        if days_since >= sc.min_days:
                            seg = sc.name
                            break

            last_service = g.last_service or ''

            result.append({
                'id': g.guest_code,
                'name': g.guest_name,
                'phone': g.phone or '',
                'smsOptIn': g.sms_opt_in,
                'storeName': g.store.name if g.store else '',
                'storeId': str(g.store.id) if g.store else '',
                'lastVisitDate': g.last_visit_date.isoformat() if g.last_visit_date else None,
                'totalVisits': g.total_visits,
                'daysSinceLastVisit': days_since,
                'segment': seg,
                'lastService': last_service,
                'lastMessageDate': g.last_sms_date.isoformat() if g.last_sms_date else None,
                'lastMessageStatus': g.last_sms_status,
            })

        return {
            'guests': result,
            'total': total,
            'page': page,
            'limit': limit,
        }

    # ------------------------------------------------------------------
    # Guest Import - File Upload (CSV / Excel)
    # ------------------------------------------------------------------

    # Our standard fields that the user can map to
    IMPORT_FIELDS = [
        {'key': 'first_name', 'label': 'First Name', 'required': True},
        {'key': 'last_name', 'label': 'Last Name', 'required': False},
        {'key': 'phone', 'label': 'Phone Number', 'required': True},
        {'key': 'email', 'label': 'Email', 'required': False},
        {'key': 'location_code', 'label': 'Location / Store Code', 'required': False},
        {'key': 'segment', 'label': 'Segment (4-week/6-week/8-week)', 'required': False},
    ]

    def parse_file_headers(self, uploaded_file):
        """
        Step 1 of 2-step import: Parse file and return detected headers + sample rows.
        Does NOT import anything yet. Returns data for the mapping UI.
        """
        from api.models import GuestImport

        if uploaded_file.size > MAX_UPLOAD_SIZE:
            return {'error': f'File too large. Max size is {MAX_UPLOAD_SIZE // (1024*1024)} MB.'}

        filename = uploaded_file.name.lower()
        try:
            if filename.endswith('.csv'):
                headers, raw_rows = self._parse_csv_raw(uploaded_file)
            elif filename.endswith(('.xlsx', '.xls')):
                headers, raw_rows = self._parse_excel_raw(uploaded_file)
            else:
                return {'error': 'Unsupported file format. Use CSV or Excel (.xlsx/.xls).'}
        except Exception as e:
            logger.error(f"File parse error: {e}")
            return {'error': f'Failed to parse file: {str(e)}'}

        if not headers:
            return {'error': 'No headers found in file.'}
        if not raw_rows:
            return {'error': 'No data rows found in file.'}

        # Auto-suggest mapping based on header names
        suggested_mapping = self._auto_suggest_mapping(headers)

        # Check if auto-mapping covers required fields
        auto_mapped_keys = set(suggested_mapping.values())
        needs_manual_mapping = not ({'first_name', 'phone'} <= auto_mapped_keys)

        # Create GuestImport record to track this upload
        guest_import = GuestImport.objects.create(
            tenant=self.tenant,
            file_name=uploaded_file.name,
            file_size=uploaded_file.size,
            status='pending',
            total_rows=len(raw_rows),
            detected_headers=headers,
            raw_rows=raw_rows[:5000],  # cap stored rows at 5000
        )

        return {
            'import_id': guest_import.id,
            'file_name': uploaded_file.name,
            'total_rows': len(raw_rows),
            'detected_headers': headers,
            'suggested_mapping': suggested_mapping,
            'needs_manual_mapping': needs_manual_mapping,
            'our_fields': self.IMPORT_FIELDS,
            'sample_rows': raw_rows[:5],  # first 5 rows for preview
        }

    def import_with_mapping(self, import_id, column_mapping, user=None):
        """
        Step 2 of 2-step import: Apply user-provided column mapping and import guests.
        """
        from api.models import GuestImport

        try:
            gi = GuestImport.objects.get(id=import_id)
        except GuestImport.DoesNotExist:
            return {'error': 'Import not found'}

        if self.tenant and gi.tenant_id != self.tenant.id:
            return {'error': 'Import not found'}

        if gi.status not in ('pending', 'mapped'):
            return {'error': f'Import already {gi.status}'}

        gi.column_mapping = column_mapping
        gi.status = 'processing'
        gi.uploaded_by = user
        gi.save(update_fields=['column_mapping', 'status', 'uploaded_by'])

        # Apply mapping to raw rows
        mapped_rows = []
        for raw_row in gi.raw_rows:
            mapped = {}
            for source_col, target_field in column_mapping.items():
                if target_field and source_col in raw_row:
                    mapped[target_field] = raw_row[source_col]
            # Handle name splitting if only first_name mapped from a full name column
            if mapped.get('first_name') and not mapped.get('last_name'):
                parts = mapped['first_name'].strip().split()
                if len(parts) > 1:
                    mapped['first_name'] = parts[0]
                    mapped['last_name'] = ' '.join(parts[1:])
            mapped_rows.append(mapped)

        # Import using existing logic
        result = self.import_guests(mapped_rows)

        # Update GuestImport record
        # Set status to 'failed' if there are errors, otherwise 'completed'
        error_count = len(result.get('errors', []))
        gi.status = 'failed' if error_count > 0 else 'completed'
        gi.created_count = result.get('created', 0)
        gi.updated_count = result.get('updated', 0)
        gi.error_count = error_count
        gi.errors = result.get('errors', [])[:100]  # cap stored errors
        gi.completed_at = timezone.now()
        gi.raw_rows = []  # clear raw data to save space
        gi.save(update_fields=[
            'status', 'created_count', 'updated_count', 'error_count',
            'errors', 'completed_at', 'raw_rows',
        ])

        result['import_id'] = gi.id
        return result

    def import_guests_from_file(self, uploaded_file, user=None):
        """
        Direct import (auto-mapping). For files matching our standard template.
        Creates a GuestImport record and processes immediately.
        """
        from api.models import GuestImport

        if uploaded_file.size > MAX_UPLOAD_SIZE:
            return {'error': f'File too large. Max size is {MAX_UPLOAD_SIZE // (1024*1024)} MB.'}

        filename = uploaded_file.name.lower()
        rows = []

        try:
            if filename.endswith('.csv'):
                rows = self._parse_csv(uploaded_file)
            elif filename.endswith(('.xlsx', '.xls')):
                rows = self._parse_excel(uploaded_file)
            else:
                return {'error': 'Unsupported file format. Use CSV or Excel (.xlsx/.xls).'}
        except Exception as e:
            logger.error(f"File parse error: {e}")
            return {'error': f'Failed to parse file: {str(e)}'}

        if not rows:
            return {'error': 'No data rows found in file.'}

        result = self.import_guests(rows)

        # Track in GuestImport
        # Set status to 'failed' if there are errors, otherwise 'completed'
        error_count = len(result.get('errors', []))
        GuestImport.objects.create(
            tenant=self.tenant,
            uploaded_by=user,
            file_name=uploaded_file.name,
            file_size=uploaded_file.size,
            status='failed' if error_count > 0 else 'completed',
            total_rows=len(rows),
            created_count=result.get('created', 0),
            updated_count=result.get('updated', 0),
            error_count=error_count,
            errors=result.get('errors', [])[:100],
            column_mapping={},  # auto-mapped
            completed_at=timezone.now(),
        )

        return result

    def list_imports(self, page=0, limit=20):
        """List guest import history for the tenant."""
        from api.models import GuestImport
        qs = GuestImport.objects.all()
        if self.tenant:
            qs = qs.filter(tenant=self.tenant)

        total = qs.count()
        offset = page * limit
        imports = qs[offset:offset + limit]

        return {
            'imports': [
                {
                    'id': gi.id,
                    'fileName': gi.file_name,
                    'fileSize': gi.file_size,
                    'status': gi.status,
                    'totalRows': gi.total_rows,
                    'created': gi.created_count,
                    'updated': gi.updated_count,
                    'errors': gi.error_count,
                    'uploadedBy': gi.uploaded_by.username if gi.uploaded_by else '',
                    'createdAt': gi.created_at.isoformat(),
                    'completedAt': gi.completed_at.isoformat() if gi.completed_at else None,
                }
                for gi in imports
            ],
            'total': total,
            'page': page,
            'limit': limit,
        }

    def _parse_csv_raw(self, uploaded_file):
        """Parse CSV and return (headers, raw_rows) without mapping."""
        content = uploaded_file.read()
        try:
            text = content.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = content.decode('latin-1')

        reader = csv.DictReader(io.StringIO(text))
        headers = [h.strip() for h in (reader.fieldnames or []) if h and h.strip()]
        raw_rows = []
        for row in reader:
            clean = {k.strip(): (v or '').strip() for k, v in row.items() if k and k.strip()}
            raw_rows.append(clean)
        return headers, raw_rows

    def _parse_excel_raw(self, uploaded_file):
        """Parse Excel and return (headers, raw_rows) without mapping."""
        import openpyxl
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        header_row = next(rows_iter, None)
        if not header_row:
            wb.close()
            return [], []

        headers = [str(h).strip() if h else '' for h in header_row]
        headers = [h for h in headers if h]

        raw_rows = []
        for row_values in rows_iter:
            if not any(row_values):
                continue
            row_dict = {}
            for i, v in enumerate(row_values):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = str(v).strip() if v is not None else ''
            raw_rows.append(row_dict)
        wb.close()
        return headers, raw_rows

    def _auto_suggest_mapping(self, headers):
        """Auto-suggest column mapping based on common header names."""
        mapping = {}
        # Normalized header → our field
        KNOWN_MAPPINGS = {
            'first_name': 'first_name', 'first name': 'first_name', 'firstname': 'first_name',
            'first': 'first_name', 'fname': 'first_name', 'given name': 'first_name',
            'name': 'first_name', 'guest name': 'first_name', 'client name': 'first_name',
            'customer name': 'first_name', 'full name': 'first_name',
            'last_name': 'last_name', 'last name': 'last_name', 'lastname': 'last_name',
            'last': 'last_name', 'lname': 'last_name', 'surname': 'last_name',
            'family name': 'last_name',
            'phone': 'phone', 'phone_number': 'phone', 'phone number': 'phone',
            'mobile': 'phone', 'mobile_number': 'phone', 'mobile number': 'phone',
            'cell': 'phone', 'cell_phone': 'phone', 'cell phone': 'phone',
            'telephone': 'phone', 'tel': 'phone', 'contact number': 'phone',
            'email': 'email', 'email_address': 'email', 'email address': 'email',
            'e-mail': 'email',
            'location_code': 'location_code', 'location code': 'location_code',
            'store_code': 'location_code', 'store code': 'location_code',
            'location': 'location_code', 'store': 'location_code',
            'center name': 'location_code', 'center_name': 'location_code',
            'salon': 'location_code', 'branch': 'location_code',
            'segment': 'segment', 'bucket': 'segment', 'group': 'segment',
        }

        used_fields = set()
        for header in headers:
            normalized = header.strip().lower()
            if normalized in KNOWN_MAPPINGS:
                target = KNOWN_MAPPINGS[normalized]
                if target not in used_fields:
                    mapping[header] = target
                    used_fields.add(target)
            else:
                mapping[header] = ''  # unmapped — user needs to assign

        return mapping

    def _parse_csv(self, uploaded_file):
        """Parse CSV file into list of dicts."""
        content = uploaded_file.read()
        try:
            text = content.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = content.decode('latin-1')

        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for row in reader:
            normalized = {
                k.strip().lower().replace(' ', '_'): (v or '').strip()
                for k, v in row.items() if k
            }
            rows.append(self._map_import_row(normalized))
        return rows

    def _parse_excel(self, uploaded_file):
        """Parse Excel file into list of dicts."""
        import openpyxl
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        # First row = headers
        headers = next(rows_iter, None)
        if not headers:
            return []

        headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in headers]
        rows = []
        for row_values in rows_iter:
            if not any(row_values):
                continue  # skip empty rows
            row_dict = {headers[i]: str(v).strip() if v is not None else '' for i, v in enumerate(row_values) if i < len(headers)}
            rows.append(self._map_import_row(row_dict))
        wb.close()
        return rows

    def _map_import_row(self, row):
        """Map various column name formats to our standard fields."""
        name_val = (row.get('name') or '').strip()
        name_parts = name_val.split() if name_val else []

        first = row.get('first_name') or row.get('firstname') or row.get('first') or ''
        if not first and name_parts:
            first = name_parts[0]

        last = row.get('last_name') or row.get('lastname') or row.get('last') or ''
        if not last and len(name_parts) > 1:
            last = name_parts[-1]

        return {
            'first_name': first,
            'last_name': last,
            'phone': row.get('phone') or row.get('phone_number') or row.get('mobile') or row.get('cell') or '',
            'email': row.get('email') or row.get('email_address') or '',
            'location_code': row.get('location_code') or row.get('store_code') or row.get('location') or row.get('store') or '',
            'segment': row.get('segment') or row.get('bucket') or '',
        }

    def import_guests(self, guests_data):
        """
        Import guests from parsed data.
        Each item: {first_name, last_name, phone, email, location_code, segment}
        """
        created = 0
        updated = 0
        errors = []
        imported_codes = []

        for i, row in enumerate(guests_data):
            first_name = (row.get('first_name') or '').strip()
            last_name = (row.get('last_name') or '').strip()
            phone = (row.get('phone') or '').strip()
            location_code = (row.get('location_code') or '').strip()

            if not first_name or not phone:
                errors.append({'row': i + 1, 'error': 'first_name and phone are required'})
                continue

            phone_clean = ''.join(c for c in phone if c.isdigit() or c == '+')
            if not phone_clean or len(phone_clean) < 10:
                errors.append({'row': i + 1, 'error': f'Invalid phone: {phone}'})
                continue

            # Normalize to E.164: default +1 if no country code
            digits_only = ''.join(c for c in phone_clean if c.isdigit())
            if len(digits_only) == 10:
                phone_clean = '+1' + digits_only
            elif len(digits_only) == 11 and digits_only.startswith('1'):
                phone_clean = '+' + digits_only
            elif not phone_clean.startswith('+'):
                phone_clean = '+' + digits_only

            guest_name = f'{first_name} {last_name}'.strip()
            guest_code = f'IMP-{phone_clean[-10:]}'

            store = None
            if location_code:
                store = Store.objects.filter(
                    Q(external_code=location_code) | Q(name__icontains=location_code)
                ).first()

            # CRM guests: no tenant, no store (unless location_code matched)
            customer, was_created = ExponentialCustomer.objects.update_or_create(
                guest_code=guest_code,
                defaults={
                    'tenant': None,
                    'store': store,
                    'guest_name': guest_name,
                    'phone': phone_clean,
                    'sms_opt_in': True,
                },
            )
            imported_codes.append(guest_code)
            if was_created:
                created += 1
            else:
                updated += 1

        return {
            'created': created,
            'updated': updated,
            'errors': errors,
            'total': len(guests_data),
            'guest_codes': imported_codes,
        }

    # ------------------------------------------------------------------
    # Helpers
    @staticmethod
    def _compute_next_recurring_run(c):
        """Compute the next scheduled run datetime (in campaign timezone) for a recurring campaign."""
        import zoneinfo
        from datetime import datetime as _dt, date as _date, timedelta

        tz_name = c.campaign_timezone or 'America/New_York'
        local_tz = zoneinfo.ZoneInfo(tz_name)
        now_local = _dt.now(local_tz)
        today = now_local.date()

        parts = (c.recurring_time or '10:00').split(':')
        hour, minute = int(parts[0]), int(parts[1]) if len(parts) > 1 else 0

        start = c.recurring_start_date if c.recurring_start_date else today
        end = c.recurring_end_date
        freq = c.recurring_frequency or 'daily'

        # Find the next run date starting from today (or start date if in future)
        candidate = max(today, start)
        for _ in range(400):  # safety limit
            if candidate > end:
                return None
            if freq == 'daily':
                # Every day is valid
                pass
            elif freq == 'weekly':
                # candidate.weekday() must match recurring_day_of_week (0=Mon)
                dow = c.recurring_day_of_week if c.recurring_day_of_week is not None else 0
                if candidate.weekday() != dow:
                    candidate += timedelta(days=1)
                    continue
            elif freq == 'monthly':
                dom = c.recurring_day_of_week if c.recurring_day_of_week and c.recurring_day_of_week >= 1 else 1
                if candidate.day != dom:
                    candidate += timedelta(days=1)
                    continue

            # Check if this candidate is in the future
            candidate_dt = _dt(candidate.year, candidate.month, candidate.day, hour, minute, tzinfo=local_tz)
            if candidate_dt > now_local:
                return candidate_dt.isoformat()

            candidate += timedelta(days=1)

        return None

    # ------------------------------------------------------------------
    @staticmethod
    def _map_message_status(raw_status):
        """Map Twilio/internal SMS status to frontend-friendly status."""
        mapping = {
            'queued': 'pending',
            'accepted': 'pending',
            'sending': 'sent',
            'sent': 'sent',
            'delivered': 'delivered',
            'undelivered': 'failed',
            'failed': 'failed',
            'canceled': 'failed',
            'pending': 'pending',
        }
        return mapping.get(raw_status, 'pending')

    def _serialize_campaign(self, c, sms_stats=None):
        """Serialize a campaign model to frontend-compatible dict. OPTIMIZED VERSION."""
        # Use pre-fetched stats if available, otherwise query (for single campaign detail view)
        if sms_stats is None:
            total = ExponentialSMSLog.objects.filter(campaign=c).count()
            delivered = ExponentialSMSLog.objects.filter(campaign=c, status='delivered').count()
            failed_count = ExponentialSMSLog.objects.filter(campaign=c, status__in=['failed', 'undelivered']).count()
            sent_count = ExponentialSMSLog.objects.filter(campaign=c, status='sent').count()
            queued_count = ExponentialSMSLog.objects.filter(campaign=c, status='queued').count()
        else:
            total = sms_stats.get('total', 0)
            delivered = sms_stats.get('delivered', 0)
            failed_count = sms_stats.get('failed', 0)
            sent_count = sms_stats.get('sent', 0)
            queued_count = sms_stats.get('queued', 0)
        
        delivery_rate = round(delivered / total * 100, 1) if total > 0 else 0

        # Map model status to frontend status
        status_map = {
            'scheduled': 'scheduled',
            'active': 'active',
            'paused': 'paused',
            'completed': 'completed',
        }
        fe_status = status_map.get(c.status, c.status)
        if fe_status == 'completed' and failed_count > 0 and delivered > 0:
            fe_status = 'partially_sent' if failed_count > total * 0.1 else 'completed'

        seg_name = SEGMENT_MAP.get(c.target_bucket, c.target_bucket)

        # Compute next scheduled run for recurring campaigns
        next_scheduled_at = None
        if c.is_recurring and c.status == 'scheduled' and c.recurring_time and c.recurring_end_date:
            next_scheduled_at = self._compute_next_recurring_run(c)

        return {
            'id': str(c.id),
            'name': c.name,
            'templateId': None,
            'templateName': c.message_template[:50] if c.message_template else 'Custom Message',
            'messageContent': c.message_template or '',
            'segment': seg_name,
            'couponValue': f'${c.coupon_value:.0f}' if c.coupon_value and c.coupon_value == int(c.coupon_value) else (f'${c.coupon_value:.2f}' if c.coupon_value else ''),
            'couponCode': c.coupon_code or '',
            'bookingLink': c.booking_link or '',
            'audienceType': 'all_locations' if c.scope == 'all' else ('select_guests' if c.scope == 'guests' else 'select_locations'),
            'locationIds': [str(x) for x in c.location_ids] if c.location_ids else ([str(c.store_id)] if c.store else []),
            'guestIds': c.guest_ids or [],
            'recipientCount': total or c.messages_sent or 0,
            'scheduledAt': c.scheduled_at.isoformat() if c.scheduled_at else None,
            'sentAt': c.started_at.isoformat() if c.started_at else None,
            'completedAt': c.completed_at.isoformat() if c.completed_at else None,
            'status': fe_status,
            'isRecurring': c.is_recurring,
            'recurringFrequency': c.recurring_frequency or None,
            'recurringStartDate': c.recurring_start_date.isoformat() if hasattr(c.recurring_start_date, 'isoformat') else (c.recurring_start_date or None),
            'recurringEndDate': c.recurring_end_date.isoformat() if hasattr(c.recurring_end_date, 'isoformat') else (c.recurring_end_date or None),
            'recurringTime': c.recurring_time or None,
            'recurringDayOfWeek': c.recurring_day_of_week,
            'campaignTimezone': c.campaign_timezone or 'America/New_York',
            'templateVariables': c.template_variables or {},
            'serviceFilter': c.service_filter or '',
            'nextScheduledAt': next_scheduled_at,
            'stats': {
                'total': total,
                'pending': queued_count,
                'sent': sent_count + delivered + failed_count,
                'delivered': delivered,
                'failed': failed_count,
                'deliveryRate': delivery_rate,
            },
            'createdAt': c.created_at.isoformat(),
            'createdBy': '',
        }
