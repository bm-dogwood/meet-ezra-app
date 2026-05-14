"""
Django management command to send scheduled reports.

Invoked by Kubernetes CronJobs with either:
  --cron-expression + --timezone  (new auto-provisioned CronJobs)
  --time + --timezone             (legacy CronJobs, backward compat)

Queries active ReportSchedule records matching the given parameters,
generates all selected report types for each schedule, and emails
the reports as multiple attachments in a single email per schedule.
"""

import logging
from datetime import datetime, timedelta, time as dt_time
from io import BytesIO

from django.core.management.base import BaseCommand
from django.utils import timezone

from api.models import ReportSchedule, Store, ReportMetric
from api.email_service import send_scheduled_report, send_scheduled_reports_multi
from api.excel_generator import ExcelReportGenerator

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Send scheduled reports for a specific cron expression/time and timezone'

    def add_arguments(self, parser):
        parser.add_argument(
            '--time',
            type=str,
            required=False,
            help='Schedule time in HH:MM format (legacy, backward compat)',
        )
        parser.add_argument(
            '--timezone',
            type=str,
            required=True,
            help='IANA timezone string (e.g. America/Chicago)',
        )
        parser.add_argument(
            '--cron-expression',
            type=str,
            required=False,
            help='Cron expression to match schedules (e.g. "0 8 * * *")',
        )

    def handle(self, *args, **options):
        cron_expr = options.get('cron_expression')
        target_tz = options['timezone']

        if cron_expr:
            # New path: match by cron_expression + timezone
            self.stdout.write(
                f'Looking for schedules with cron="{cron_expr}" timezone={target_tz}...'
            )
            schedules = ReportSchedule.objects.filter(
                is_active=True,
                cron_expression=cron_expr,
                timezone=target_tz,
            )
        else:
            # Backward compat: match by time + timezone
            target_time_str = options.get('time')
            if not target_time_str:
                self.stderr.write(self.style.ERROR(
                    'Either --cron-expression or --time must be provided.'
                ))
                return

            try:
                hour, minute = target_time_str.split(':')
                target_time = dt_time(int(hour), int(minute))
            except (ValueError, TypeError):
                self.stderr.write(self.style.ERROR(
                    f'Invalid time format: {target_time_str}. Use HH:MM.'
                ))
                return

            self.stdout.write(
                f'Looking for schedules at {target_time_str} {target_tz}...'
            )
            schedules = ReportSchedule.objects.filter(
                is_active=True,
                schedule_time=target_time,
                timezone=target_tz,
            )

        count = schedules.count()
        if count == 0:
            self.stdout.write('No matching schedules found.')
            return

        self.stdout.write(f'Found {count} schedule(s) to process.')

        success_count = 0
        failure_count = 0

        for schedule in schedules:
            try:
                self._execute_schedule(schedule)
                if schedule.last_run_status == 'success':
                    success_count += 1
                elif schedule.last_run_status == 'partial':
                    success_count += 1  # partial is still a (partial) success
                else:
                    failure_count += 1
            except Exception as exc:
                failure_count += 1
                error_msg = str(exc)
                logger.error(
                    'Schedule %s failed: %s',
                    schedule.id,
                    error_msg,
                    exc_info=True,
                )
                schedule.last_run_at = timezone.now()
                schedule.last_run_status = 'failed'
                schedule.last_run_error = error_msg
                schedule.save(update_fields=['last_run_at', 'last_run_status', 'last_run_error'])

        self.stdout.write(
            self.style.SUCCESS(
                f'Done. {success_count} succeeded, {failure_count} failed.'
            )
        )

    def _execute_schedule(self, schedule):
        """Generate all report types and send as multi-attachment email."""
        self.stdout.write(
            f'  Processing schedule {schedule.id} '
            f'(report_types={schedule.report_types})...'
        )

        tenant = schedule.tenant
        attachments = []
        errors = []

        for report_type in schedule.report_types:
            try:
                excel_bytes, filename, report_name = self._generate_report(
                    report_type, tenant
                )
                attachments.append((filename, excel_bytes, report_name))
            except Exception as e:
                logger.error(
                    'Schedule %s: report type "%s" failed: %s',
                    schedule.id,
                    report_type,
                    str(e),
                    exc_info=True,
                )
                errors.append(f"{report_type}: {str(e)}")

        # Send email with all successful attachments
        if attachments:
            send_scheduled_reports_multi(
                recipients=schedule.recipients,
                attachments=attachments,
            )

        # Record execution status
        if errors and not attachments:
            schedule.last_run_status = 'failed'
            schedule.last_run_error = '; '.join(errors)
        elif errors:
            schedule.last_run_status = 'partial'
            schedule.last_run_error = '; '.join(errors)
        else:
            schedule.last_run_status = 'success'
            schedule.last_run_error = None

        schedule.last_run_at = timezone.now()
        schedule.save(update_fields=['last_run_at', 'last_run_status', 'last_run_error'])

        if schedule.last_run_status == 'success':
            self.stdout.write(f'    ✓ Schedule {schedule.id} completed successfully.')
        elif schedule.last_run_status == 'partial':
            self.stdout.write(
                f'    ⚠ Schedule {schedule.id} completed with partial failures: '
                f'{schedule.last_run_error}'
            )
        else:
            self.stdout.write(
                f'    ✗ Schedule {schedule.id} failed: {schedule.last_run_error}'
            )

    def _generate_report(self, report_type, tenant):
        """
        Generate an Excel report scoped to the given tenant.

        Returns:
            Tuple of (excel_bytes: BytesIO, filename: str, report_name: str)
        """
        yesterday = datetime.now().date() - timedelta(days=1)

        if report_type == 'daily':
            return self._generate_daily_sales_report(tenant, yesterday)
        elif report_type == 'weekly':
            end_date = yesterday
            start_date = end_date - timedelta(days=6)
            return self._generate_weekly_sales_report(tenant, start_date, end_date)
        elif report_type == 'lp':
            return self._generate_lp_report(tenant, yesterday)
        elif report_type == 'scheduling':
            return self._generate_scheduling_report(tenant, yesterday)
        elif report_type == 'exponential':
            return self._generate_exponential_report(tenant, yesterday)
        else:
            raise ValueError(f'Unknown report type: {report_type}')

    def _generate_daily_sales_report(self, tenant, report_date):
        """Generate a daily sales flash Excel report."""
        stores = Store.objects.filter(tenant=tenant, status='active')
        metrics = ReportMetric.objects.filter(
            store__in=stores,
            report_type='sales',
            report_date=report_date,
        ).select_related('store')

        data = self._pivot_metrics_to_rows(metrics)

        generator = ExcelReportGenerator()
        excel_bytes = generator.generate_daily_report(data, report_date)

        filename = f'daily_sales_flash_{report_date.strftime("%Y-%m-%d")}.xlsx'
        return excel_bytes, filename, 'Daily Sales Flash'

    def _generate_weekly_sales_report(self, tenant, start_date, end_date):
        """Generate a weekly sales summary Excel report."""
        stores = Store.objects.filter(tenant=tenant, status='active')
        metrics = ReportMetric.objects.filter(
            store__in=stores,
            report_type='sales',
            report_date__gte=start_date,
            report_date__lte=end_date,
        ).select_related('store')

        data = self._pivot_metrics_to_rows(metrics)

        generator = ExcelReportGenerator()
        excel_bytes = generator.generate_weekly_report(data, start_date, end_date)

        filename = (
            f'weekly_sales_summary_'
            f'{start_date.strftime("%Y-%m-%d")}_to_{end_date.strftime("%Y-%m-%d")}.xlsx'
        )
        return excel_bytes, filename, 'Weekly Sales Summary'

    def _generate_lp_report(self, tenant, report_date):
        """Generate an LP Risk Analysis Excel report.
        Recalculates risk scores from latest raw data before generating."""
        from api.services.lp_service import LPService
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        stores = Store.objects.filter(tenant=tenant, status='active')
        lp_service = LPService(tenant=tenant)

        # Recalculate risk scores from latest data before generating report
        for store in stores:
            try:
                lp_service.calculate_risk_score_for_store(store, report_date)
            except Exception as e:
                logger.warning(
                    'LP recalc failed for store %s date %s: %s',
                    store.id, report_date, str(e),
                )

        report_data = lp_service.generate_lp_report_data(stores, report_date)

        # Build Excel workbook (same format as LPReportDownloadView)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'LP Risk Analysis'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        headers = [
            'Center Name',
            'High/Low Cash to Credit Ratio',
            'High/Low Tip %',
            'High % of Low-Ticket-Value Services',
            'Locations with 2+ Flags',
            '# of High-Risk Locations',
            'Cash %',
            'Tip Percentage',
            '% of Low-Ticket-Value',
            'Service Sales',
            'Tip Amount',
            'Credit Card',
            'Cash Payment',
            'All Services',
            'Low-Ticket-Services',
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        ws.column_dimensions['A'].width = 30
        for col in range(2, 16):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        high_risk_count = sum(1 for d in report_data if d['is_high_risk'])

        for row_idx, data in enumerate(report_data, 2):
            ws.cell(row=row_idx, column=1, value=data['center_name']).border = thin_border

            cell = ws.cell(row=row_idx, column=2, value=data['cash_ratio_risk'].upper())
            cell.border = thin_border
            if data['cash_ratio_risk'] == 'high':
                cell.fill = red_fill
            elif data['cash_ratio_risk'] == 'medium':
                cell.fill = yellow_fill
            else:
                cell.fill = green_fill

            cell = ws.cell(row=row_idx, column=3, value=data['tip_percent_risk'].upper())
            cell.border = thin_border
            if data['tip_percent_risk'] == 'high':
                cell.fill = red_fill
            elif data['tip_percent_risk'] == 'medium':
                cell.fill = yellow_fill
            else:
                cell.fill = green_fill

            cell = ws.cell(row=row_idx, column=4, value=data['low_ticket_risk'].upper())
            cell.border = thin_border
            if data['low_ticket_risk'] == 'high':
                cell.fill = red_fill
            elif data['low_ticket_risk'] == 'medium':
                cell.fill = yellow_fill
            else:
                cell.fill = green_fill

            cell = ws.cell(row=row_idx, column=5, value='Yes' if data['has_2plus_flags'] else '')
            cell.border = thin_border
            if data['has_2plus_flags']:
                cell.fill = yellow_fill

            ws.cell(
                row=row_idx, column=6,
                value=high_risk_count if row_idx == 2 else '',
            ).border = thin_border

            ws.cell(row=row_idx, column=7, value=f"{data['cash_percent']:.2f}%").border = thin_border
            ws.cell(row=row_idx, column=8, value=f"{data['tip_percent']:.2f}%").border = thin_border
            ws.cell(row=row_idx, column=9, value=f"{data['low_ticket_percent']:.2f}%").border = thin_border
            ws.cell(row=row_idx, column=10, value=f"${data['service_sales']:,.2f}").border = thin_border
            ws.cell(row=row_idx, column=11, value=f"${data['tip_amount']:,.2f}").border = thin_border
            ws.cell(row=row_idx, column=12, value=f"${data['credit_card']:,.2f}").border = thin_border
            ws.cell(row=row_idx, column=13, value=f"${data['cash_payment']:,.2f}").border = thin_border
            ws.cell(row=row_idx, column=14, value=data['all_services']).border = thin_border
            ws.cell(row=row_idx, column=15, value=data['low_ticket_services']).border = thin_border

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'lp_risk_analysis_{report_date.strftime("%Y-%m-%d")}.xlsx'
        return output, filename, 'LP Risk Analysis'

    # ------------------------------------------------------------------
    # Helpers reused from SalesReportDownloadView
    # ------------------------------------------------------------------

    METRIC_TO_COLUMN = {
        # Sales report metrics → new template columns
        'Total Net': 'Total Net',
        'Service Net': 'Service Net',
        'Product Net': 'Product Net',
        'Tip Amount': 'Tip Amount',
        'CC': 'CC',
        'Cash': 'Cash',
        # Fallbacks for variant metric names
        'Total Sales': 'Total Net',
        'Service Sales': 'Service Net',
        'Product Sales': 'Product Net',
        'GIft Card Sales': 'Gift Card Sales',
        'Gift Card Sales': 'Gift Card Sales',
        'Gift Net': 'Gift Net',
    }

    def _generate_scheduling_report(self, tenant, report_date):
        """Generate a scheduling report Excel for the last 7 days."""
        from api.services.scheduling_service import SchedulingService
        from api.constants import DEFAULT_SCHEDULING_CONFIG
        from api.models import AppConfig
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        end_date = report_date
        start_date = end_date - timedelta(days=6)

        config = AppConfig.get_config_value('scheduling_config', DEFAULT_SCHEDULING_CONFIG)
        service = SchedulingService(tenant=tenant, config=config)
        rankings = service.get_store_rankings(tenant=tenant,
                                              start_date=start_date.isoformat(),
                                              end_date=end_date.isoformat())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Scheduling Report'

        headers = ['Rank', 'Location', 'Store Code', 'State', 'Revenue',
                    'Labor Hours', 'Idle Hours', 'Idle %', 'TSTH', 'Tix/Hr',
                    'Labor $', 'OT Hours', 'OT Flag']
        hdr_font = Font(bold=True, color='FFFFFF')
        hdr_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center')
            c.border = border

        widths = [6, 35, 15, 8, 14, 12, 12, 10, 10, 10, 12, 10, 8]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        red_font = Font(color='FF0000', bold=True)
        for rank, loc in enumerate(rankings, 1):
            r = rank + 1
            ws.cell(row=r, column=1, value=rank).border = border
            ws.cell(row=r, column=2, value=loc['locationName']).border = border
            ws.cell(row=r, column=3, value=loc['storeCode']).border = border
            ws.cell(row=r, column=4, value=loc['state']).border = border
            ws.cell(row=r, column=5, value=float(loc['revenue'])).border = border
            ws.cell(row=r, column=5).number_format = '$#,##0'
            ws.cell(row=r, column=6, value=float(loc['laborHours'])).border = border
            ws.cell(row=r, column=6).number_format = '#,##0.0'
            ws.cell(row=r, column=7, value=float(loc['idleHours'])).border = border
            ws.cell(row=r, column=7).number_format = '#,##0.0'
            pc = ws.cell(row=r, column=8, value=float(loc['idlePercent']) / 100)
            pc.number_format = '0.0%'
            pc.border = border
            if loc['idlePercent'] >= 75:
                pc.font = red_font
            ws.cell(row=r, column=9, value=float(loc['tsth'])).border = border
            ws.cell(row=r, column=9).number_format = '$#,##0'
            ws.cell(row=r, column=10, value=float(loc['ticketsPerLaborHour'])).border = border
            ws.cell(row=r, column=10).number_format = '0.0'
            ws.cell(row=r, column=11, value=float(loc['laborCost'])).border = border
            ws.cell(row=r, column=11).number_format = '$#,##0'
            ws.cell(row=r, column=12, value=float(loc['overtimeHours'])).border = border
            ws.cell(row=r, column=12).number_format = '#,##0.0'
            ot = ws.cell(row=r, column=13, value='OT' if loc['hasOvertimeFlag'] else '')
            ot.border = border
            if loc['hasOvertimeFlag']:
                ot.font = red_font

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f'Scheduling_Report_{start_date.strftime("%Y%m%d")}_to_{end_date.strftime("%Y%m%d")}.xlsx'
        return buf, filename, 'Scheduling Report'

    def _generate_exponential_report(self, tenant, report_date):
        """Generate an Exponential report Excel with location summary, segments, and daily campaigns."""
        from api.services.exponential_service import ExponentialService
        from api.constants import DEFAULT_EXPONENTIAL_CONFIG
        from api.models import AppConfig
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        end_date = report_date
        start_date = end_date - timedelta(days=29)

        config = AppConfig.get_config_value('exponential_config', DEFAULT_EXPONENTIAL_CONFIG)
        service = ExponentialService(tenant=tenant, config=config)
        overview = service.get_overview_metrics(
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
        )

        wb = openpyxl.Workbook()
        hdr_font = Font(bold=True, color='FFFFFF')
        hdr_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        red_font = Font(color='FF0000', bold=True)

        # --- Sheet 1: Location Summary ---
        ws = wb.active
        ws.title = 'Location Summary'
        loc_headers = [
            'Rank', 'Location', 'Store Code', 'State', 'Guests MTD',
            'Customers Last Mo.', '4-Week', '6-Week', '8-Week',
            'Follow-ups Sent', 'Uptake %', 'Retention Risk',
        ]
        for ci, h in enumerate(loc_headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center')
            c.border = border

        loc_widths = [6, 35, 15, 8, 12, 16, 10, 10, 10, 14, 10, 14]
        for i, w in enumerate(loc_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for rank, loc in enumerate(overview.get('locationSummaries', []), 1):
            r = rank + 1
            ws.cell(row=r, column=1, value=rank).border = border
            ws.cell(row=r, column=2, value=loc.get('locationName', '')).border = border
            ws.cell(row=r, column=3, value=loc.get('storeCode', '')).border = border
            ws.cell(row=r, column=4, value=loc.get('state', '')).border = border
            ws.cell(row=r, column=5, value=loc.get('guestsMTD', 0)).border = border
            ws.cell(row=r, column=6, value=loc.get('customersLastMonth', 0)).border = border
            ws.cell(row=r, column=7, value=loc.get('fourWeekCount', 0)).border = border
            ws.cell(row=r, column=8, value=loc.get('sixWeekCount', 0)).border = border
            ws.cell(row=r, column=9, value=loc.get('eightWeekCount', 0)).border = border
            ws.cell(row=r, column=10, value=loc.get('followUpsSent', 0)).border = border
            pc = ws.cell(row=r, column=11, value=loc.get('overallUptake', 0) / 100)
            pc.number_format = '0.0%'
            pc.border = border
            risk = ws.cell(row=r, column=12, value=loc.get('retentionRiskScore', 0))
            risk.border = border
            if loc.get('retentionRiskScore', 0) >= 50:
                risk.font = red_font

        # --- Sheet 2: Segment Summary ---
        ws2 = wb.create_sheet('Segment Summary')
        seg_headers = ['Segment', 'Customers', 'Risk Level', 'Messages Sent', 'Uptake %']
        for ci, h in enumerate(seg_headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center')
            c.border = border

        seg_widths = [15, 12, 12, 14, 10]
        for i, w in enumerate(seg_widths, 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for idx, seg in enumerate(overview.get('segments', []), 1):
            r = idx + 1
            ws2.cell(row=r, column=1, value=seg.get('name', '')).border = border
            ws2.cell(row=r, column=2, value=seg.get('customerCount', 0)).border = border
            ws2.cell(row=r, column=3, value=seg.get('riskLevel', '')).border = border
            ws2.cell(row=r, column=4, value=seg.get('messagesSent', 0)).border = border
            pc = ws2.cell(row=r, column=5, value=seg.get('uptakePercent', 0) / 100)
            pc.number_format = '0.0%'
            pc.border = border

        # --- Sheet 3: Daily Campaign Activity ---
        ws3 = wb.create_sheet('Daily Campaigns')
        daily_headers = ['Date', '4-Week Sends', '6-Week Sends', '8-Week Sends', 'Total Sends']
        for ci, h in enumerate(daily_headers, 1):
            c = ws3.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center')
            c.border = border

        daily_widths = [14, 14, 14, 14, 12]
        for i, w in enumerate(daily_widths, 1):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for idx, day in enumerate(overview.get('dailyCampaigns', []), 1):
            r = idx + 1
            ws3.cell(row=r, column=1, value=day.get('date', '')).border = border
            ws3.cell(row=r, column=2, value=day.get('fourWeekSends', 0)).border = border
            ws3.cell(row=r, column=3, value=day.get('sixWeekSends', 0)).border = border
            ws3.cell(row=r, column=4, value=day.get('eightWeekSends', 0)).border = border
            ws3.cell(row=r, column=5, value=day.get('totalSends', 0)).border = border

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f'Exponential_Report_{start_date.strftime("%Y%m%d")}_to_{end_date.strftime("%Y%m%d")}.xlsx'
        return buf, filename, 'Exponential Report'

    def _pivot_metrics_to_rows(self, metrics):
        """Transform row-based ReportMetric records into column-based dicts
        matching the new 12-column template."""
        from api.models import StoreTarget

        grouped = {}
        store_dates = set()

        for metric in metrics:
            store_name = metric.store.name if metric.store else 'Unknown'
            report_date = metric.report_date
            key = (metric.store_id, store_name, report_date)
            store_dates.add((metric.store_id, report_date))

            if key not in grouped:
                grouped[key] = {
                    'Center Name': store_name,
                    'Date': report_date.strftime('%Y-%m-%d') if report_date else '',
                    '_store_id': metric.store_id,
                    '_report_date': report_date,
                }

            column_name = self.METRIC_TO_COLUMN.get(metric.metric_name)
            if column_name and metric.metric_value is not None:
                grouped[key][column_name] = float(metric.metric_value)

        # Enrich with Revenue Target, Labor Target from StoreTarget
        for key, row in grouped.items():
            store_id = row.get('_store_id')
            report_date = row.get('_report_date')
            if store_id and report_date:
                try:
                    target = StoreTarget.objects.get(store_id=store_id, target_date=report_date)
                    row['Revenue Target'] = float(target.revenue_target)
                    row['Labor Target'] = float(target.labor_target_hours)
                except StoreTarget.DoesNotExist:
                    pass

        # Enrich with Labor Hours from attendance metrics
        if store_dates:
            attendance_metrics = ReportMetric.objects.filter(
                report_type='attendance',
                metric_name='Actual Hours',
            ).filter(
                store_id__in=[sd[0] for sd in store_dates],
            ).select_related('store')

            labor_by_store_date = {}
            for m in attendance_metrics:
                lkey = (m.store_id, m.report_date)
                labor_by_store_date[lkey] = labor_by_store_date.get(lkey, 0) + float(m.metric_value or 0)

            for key, row in grouped.items():
                store_id = row.get('_store_id')
                report_date = row.get('_report_date')
                lkey = (store_id, report_date)
                if lkey in labor_by_store_date:
                    row['Labor Hours'] = labor_by_store_date[lkey]

        # Compute Average Ticket = Total Net / ticket count (from production)
        if store_dates:
            ticket_metrics = ReportMetric.objects.filter(
                report_type='production',
                metric_name='Total Ticket Count',
            ).filter(
                store_id__in=[sd[0] for sd in store_dates],
            )

            tickets_by_store_date = {}
            for m in ticket_metrics:
                tkey = (m.store_id, m.report_date)
                tickets_by_store_date[tkey] = tickets_by_store_date.get(tkey, 0) + int(m.metric_value or 0)

            for key, row in grouped.items():
                store_id = row.get('_store_id')
                report_date = row.get('_report_date')
                tkey = (store_id, report_date)
                total_net = row.get('Total Net', 0)
                tickets = tickets_by_store_date.get(tkey, 0)
                if tickets > 0 and total_net:
                    row['Average Ticket'] = round(total_net / tickets, 2)

        # Clean up internal keys
        rows = []
        for row in grouped.values():
            row.pop('_store_id', None)
            row.pop('_report_date', None)
            # Remove old columns not in the new template
            row.pop('Gift Card Sales', None)
            row.pop('Gift Net', None)
            rows.append(row)

        rows.sort(key=lambda x: (x.get('Center Name', ''), x.get('Date', '')))
        return rows
